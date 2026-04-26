import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.data_sources.tushare_financial_facts import TushareFinancialFactsAdapter  # type: ignore
from scripts.forecast_tools.rollforward import (  # type: ignore
    WorkbookBlueprintParser,
    build_model_segment_tree,
    build_reconciliation_audit,
    build_segment_mapping_contract,
)


def build_mapping_sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "营收拆分"
    ws["B2"] = "单位：亿元"
    ws["D2"] = 2024
    ws["E2"] = 2025
    ws["F2"] = "2026E"
    ws["B3"] = "手机CIS"
    ws["C3"] = "收入"
    ws["D3"] = 120.0
    ws["E3"] = 140.0
    ws["B4"] = "汽车CIS"
    ws["C4"] = "收入"
    ws["D4"] = 30.0
    ws["E4"] = 48.0
    ws["B5"] = "安防"
    ws["C5"] = "收入"
    ws["D5"] = 20.0
    ws["E5"] = 18.0
    ws["B6"] = "营业收入"
    ws["C6"] = "收入"
    ws["D6"] = 210.0
    ws["E6"] = 240.0
    wb.save(path)


def build_hierarchical_cis_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "营收拆分"
    ws["B2"] = "单位：亿元"
    ws["D2"] = 2024
    ws["E2"] = 2025
    ws["F2"] = "2026E"
    ws["B3"] = "CIS"
    ws["C3"] = "收入"
    ws["D3"] = 150.0
    ws["E3"] = 188.0
    ws["B4"] = "1）手机营收（亿元）"
    ws["C4"] = "收入"
    ws["D4"] = 120.0
    ws["E4"] = 140.0
    ws["B5"] = "2）汽车营收（亿元）"
    ws["C5"] = "收入"
    ws["D5"] = 30.0
    ws["E5"] = 48.0
    ws["B6"] = "安防"
    ws["C6"] = "收入"
    ws["D6"] = 20.0
    ws["E6"] = 18.0
    ws["B7"] = "营业税金及附加/营收收入（%）"
    ws["C7"] = "收入"
    ws["D7"] = 0.01
    ws["E7"] = 0.01
    ws["B8"] = "营业收入"
    ws["C8"] = "收入"
    ws["D8"] = 170.0
    ws["E8"] = 206.0
    wb.save(path)


class TushareFinancialFactsAdapterTests(unittest.TestCase):
    def test_model_segment_tree_excludes_generic_rows_and_tracks_parent_lineage(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "hierarchical_cis.xlsx"
            build_hierarchical_cis_workbook(workbook_path)
            blueprint = WorkbookBlueprintParser().parse(workbook_path)

            tree = build_model_segment_tree(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
            )

            reportable_segments = {item["segment_name"]: item for item in tree["reportable_segments"]}
            self.assertIn("1）手机营收（亿元）", reportable_segments)
            self.assertIn("2）汽车营收（亿元）", reportable_segments)
            self.assertIn("安防", reportable_segments)
            self.assertNotIn("营业税金及附加/营收收入（%）", reportable_segments)
            self.assertNotIn("营业收入", reportable_segments)
            self.assertIn("CIS", reportable_segments["1）手机营收（亿元）"]["lineage_labels"])
            self.assertIsInstance(reportable_segments["1）手机营收（亿元）"]["bridge_tags"], list)

    def test_adapter_builds_workflow_compatible_financial_facts(self) -> None:
        class FakeClient:
            def income(self, **kwargs):
                return [
                    {
                        "ts_code": "603501.SH",
                        "ann_date": "20260331",
                        "end_date": "20251231",
                        "total_revenue": 28854805517.49,
                        "oper_cost": 20015173377.35,
                        "n_income_attr_p": 4045416530.33,
                        "sell_exp": 564341702.89,
                        "admin_exp": 696392890.54,
                        "rd_exp": 2842872872.68,
                        "fin_exp": -73679924.87,
                    }
                ]

            def fina_indicator(self, **kwargs):
                return [
                    {
                        "ts_code": "603501.SH",
                        "ann_date": "20260331",
                        "end_date": "20251231",
                        "grossprofit_margin": 30.6349,
                        "expense_of_gr": 14.4,
                        "saleexp_to_gr": 4.5,
                        "adminexp_of_gr": 4.0,
                        "roe_dt": 0.0,
                        "dt_netprofit": 3909518597.14,
                    }
                ]

            def fina_mainbz(self, **kwargs):
                return [
                    {
                        "ts_code": "603501.SH",
                        "end_date": "20251231",
                        "bz_item": "图像传感器解决方案业务",
                        "bz_sales": 21245839825.85,
                        "bz_cost": 13601778163.68,
                    },
                    {
                        "ts_code": "603501.SH",
                        "end_date": "20251231",
                        "bz_item": "显示解决方案业务",
                        "bz_sales": 941127046.72,
                        "bz_cost": 806606147.97,
                    },
                ]

        result = TushareFinancialFactsAdapter(client=FakeClient()).extract(
            company="豪威集团",
            report_year=2025,
            ts_code="603501.SH",
        )

        self.assertEqual(result["source_type"], "tushare")
        self.assertEqual(result["reported_facts"]["营业收入"], 28854805517.49)
        self.assertAlmostEqual(result["reported_facts"]["毛利"], 8839632140.14, places=2)
        self.assertAlmostEqual(result["reported_facts"]["毛利率"], 0.306349, places=6)
        self.assertEqual(result["reported_facts"]["归母净利润"], 4045416530.33)
        self.assertEqual(result["reported_facts"]["扣非归母净利润"], 3909518597.14)
        self.assertEqual(result["reported_facts"]["销售费用"], 564341702.89)
        self.assertEqual(result["reported_facts"]["财务费用"], -73679924.87)
        self.assertTrue(any(item["metric"] == "营业收入" for item in result["fact_items"]))
        self.assertTrue(any(item["metric"] == "毛利率" and item["source_ref"].startswith("tushare:") for item in result["fact_items"]))
        self.assertTrue(any(item["segment"] == "图像传感器解决方案业务" for item in result["segment_disclosure"]))
        self.assertTrue(any(item["segment"] == "显示解决方案业务" and item["gross_margin"] > 0.14 for item in result["segment_disclosure"]))

    def test_segment_mapping_prefers_workbook_structure_and_tracks_parent_child_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "model.xlsx"
            build_mapping_sample_workbook(workbook_path)
            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            financial_facts = {
                "reported_facts": {"营业收入": 240.0},
                "segment_disclosure": [
                    {"segment": "CIS", "revenue": 188.0, "gross_margin": 0.31, "unit": "亿元"},
                    {"segment": "安防", "revenue": 18.0, "gross_margin": 0.22, "unit": "亿元"},
                ],
            }

            mapping = build_segment_mapping_contract(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
            )
            audit = build_reconciliation_audit(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
                segment_mapping=mapping,
            )

            workbook_segments = {item["workbook_segment"]: item for item in mapping["segment_mappings"]}
            self.assertEqual(workbook_segments["手机CIS"]["mapping_type"], "parent_child")
            self.assertEqual(workbook_segments["手机CIS"]["tushare_segment"], "CIS")
            self.assertEqual(workbook_segments["汽车CIS"]["mapping_type"], "parent_child")
            self.assertEqual(workbook_segments["安防"]["mapping_type"], "exact")
            self.assertIn("revenue_gap", audit)
            self.assertIn("mapped_revenue_total", audit)


    def test_adapter_normalizes_segment_disclosure_and_filters_non_comparable_dimensions(self) -> None:
        class FakeClient:
            def income(self, **kwargs):
                return [{"total_revenue": 240.0, "oper_cost": 180.0}]

            def fina_indicator(self, **kwargs):
                return [{"grossprofit_margin": 25.0}]

            def fina_mainbz(self, **kwargs):
                return [
                    {"bz_item": "图像传感解决方案业务", "bz_sales": 188.0, "bz_cost": 130.0},
                    {"bz_item": "境外地区", "bz_sales": 160.0, "bz_cost": 120.0},
                    {"bz_item": "直销模式", "bz_sales": 120.0, "bz_cost": 90.0},
                    {"bz_item": "显示解决方案业务", "bz_sales": 20.0, "bz_cost": 18.0},
                ]

        result = TushareFinancialFactsAdapter(client=FakeClient()).extract(
            company="豪威集团",
            report_year=2025,
            ts_code="603501.SH",
        )

        segments = {item["segment"]: item for item in result["segment_disclosure"]}
        self.assertIn("图像传感解决方案业务", segments)
        self.assertIn("显示解决方案业务", segments)
        self.assertNotIn("境外地区", segments)
        self.assertNotIn("直销模式", segments)
        self.assertEqual(segments["图像传感解决方案业务"]["segment_dimension"], "business")
        self.assertTrue(segments["图像传感解决方案业务"]["mapping_ready"])

    def test_segment_mapping_collects_workbook_revenue_rows_with_embedded_revenue_labels(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "embedded_revenue_labels.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "营收拆分"
            ws["B2"] = "单位：亿元"
            ws["D2"] = 2024
            ws["E2"] = 2025
            ws["F2"] = "2026E"
            ws["B3"] = "1）手机营收（亿元）"
            ws["C3"] = "收入"
            ws["D3"] = 120.0
            ws["E3"] = 140.0
            ws["B4"] = "2）汽车营收（亿元）"
            ws["C4"] = "收入"
            ws["D4"] = 30.0
            ws["E4"] = 48.0
            ws["B5"] = "营业收入"
            ws["C5"] = "收入"
            ws["D5"] = "=D3+D4"
            ws["E5"] = "=E3+E4"
            wb.save(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            mapping = build_segment_mapping_contract(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts={
                    "reported_facts": {"营业收入": 188.0},
                    "segment_disclosure": [
                        {"segment": "手机", "revenue": 140.0, "mapping_ready": True},
                        {"segment": "汽车", "revenue": 48.0, "mapping_ready": True},
                    ],
                },
            )

            workbook_segments = [item["workbook_segment"] for item in mapping["segment_mappings"]]
            self.assertIn("1）手机营收（亿元）", workbook_segments)
            self.assertIn("2）汽车营收（亿元）", workbook_segments)
            self.assertNotIn("营业收入", workbook_segments)


    def test_segment_mapping_does_not_force_false_anchor_for_broad_disclosure_without_shared_semantics(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "hierarchical_cis.xlsx"
            build_hierarchical_cis_workbook(workbook_path)
            blueprint = WorkbookBlueprintParser().parse(workbook_path)

            mapping = build_segment_mapping_contract(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts={
                    "reported_facts": {"营业收入": 206.0},
                    "segment_disclosure": [
                        {"segment": "消费图像解决方案业务", "revenue": 188.0, "mapping_ready": True},
                        {"segment": "安防", "revenue": 18.0, "mapping_ready": True},
                    ],
                },
                meeting_notes_facts={
                    "bridge_facts": {
                        "消费图像应用": {
                            "claim": "手机和汽车应用主要由消费图像解决方案业务承载，同比增长15%",
                            "growth_rate": 0.15,
                        }
                    }
                },
                evidence_payload={
                    "providers": [
                        {
                            "source_ref": "meeting-note",
                            "source_type": "user_supplied_text",
                            "source_tier": "reference_files",
                            "content": "手机和汽车应用主要由消费图像解决方案业务承载，同比增长15%。",
                            "metadata": {"reference_kind": "meeting_notes"},
                        }
                    ]
                },
            )

            workbook_segments = {item["workbook_segment"]: item for item in mapping["segment_mappings"]}
            self.assertEqual(workbook_segments["1）手机营收（亿元）"]["mapping_type"], "proxy")
            self.assertEqual(workbook_segments["1）手机营收（亿元）"]["tushare_segment"], "")
            self.assertEqual(workbook_segments["2）汽车营收（亿元）"]["mapping_type"], "proxy")
            self.assertEqual(workbook_segments["安防"]["mapping_type"], "exact")
            self.assertGreaterEqual(mapping["anchored_segment_count"], 1)

    def test_reconciliation_audit_requests_candidate_decision_for_limited_mapping_ambiguity(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "model.xlsx"
            build_mapping_sample_workbook(workbook_path)
            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            financial_facts = {
                "reported_facts": {"营业收入": 260.0},
                "segment_disclosure": [],
            }
            mapping = {
                "segment_tree": {
                    "reportable_segments": [
                        {"segment_name": "手机业务", "reported_value": 140.0},
                        {"segment_name": "汽车业务", "reported_value": 48.0},
                        {"segment_name": "安防", "reported_value": 18.0},
                    ]
                },
                "segment_mappings": [
                    {
                        "workbook_segment": "手机业务",
                        "tushare_segment": "消费电子业务",
                        "mapping_type": "proxy",
                        "confidence": "low",
                        "workbook_reported_value": 140.0,
                    },
                    {
                        "workbook_segment": "汽车业务",
                        "tushare_segment": "出行业务",
                        "mapping_type": "proxy",
                        "confidence": "low",
                        "workbook_reported_value": 48.0,
                    },
                    {
                        "workbook_segment": "安防",
                        "tushare_segment": "安防",
                        "mapping_type": "exact",
                        "confidence": "high",
                        "workbook_reported_value": 18.0,
                    },
                ],
                "anchored_segment_count": 1,
                "proxy_segment_count": 2,
                "candidate_clusters": [
                    {
                        "workbook_segment": "手机业务",
                        "candidate_matches": [
                            {"tushare_segment": "消费电子业务", "mapping_type": "proxy", "confidence": "low", "score": 55, "base_score": 20, "tushare_revenue": 140.0, "source_refs": ["t1"]},
                            {"tushare_segment": "移动终端业务", "mapping_type": "proxy", "confidence": "low", "score": 50, "base_score": 18, "tushare_revenue": 140.0, "source_refs": ["t2"]},
                        ],
                    }
                ],
                "unmapped_tushare_segments": ["消费电子业务", "移动终端业务"],
            }
            audit = build_reconciliation_audit(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
                segment_mapping=mapping,
            )

            self.assertFalse(audit["within_tolerance"])
            self.assertEqual(audit["resolution_mode"], "candidate_decision_required")
            self.assertEqual(len(audit["candidate_options"]), 3)
            self.assertTrue(any(option["recommended"] for option in audit["candidate_options"]))
            self.assertTrue(all(option["segment_assignments"] for option in audit["candidate_options"]))
            self.assertTrue(any(item["anchored_revenue_total"] > 0 for item in audit["candidate_options"]))

    def test_reconciliation_audit_hard_stops_when_no_safe_mapping_options_exist(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "summary_only.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "营收拆分"
            ws["B2"] = "单位：亿元"
            ws["D2"] = 2024
            ws["E2"] = 2025
            ws["F2"] = "2026E"
            ws["B3"] = "营业收入"
            ws["C3"] = "收入"
            ws["D3"] = 210.0
            ws["E3"] = 240.0
            wb.save(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            financial_facts = {"reported_facts": {"营业收入": 240.0}, "segment_disclosure": []}
            mapping = build_segment_mapping_contract(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
            )
            audit = build_reconciliation_audit(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
                segment_mapping=mapping,
            )

            self.assertFalse(audit["within_tolerance"])
            self.assertEqual(audit["resolution_mode"], "hard_stop")
            self.assertIn("no_workbook_segments_identified", audit["fail_reasons"])
            self.assertEqual(audit["candidate_options"], [])

    def test_reconciliation_audit_still_offers_candidate_decision_for_low_quality_proxy_when_clusters_exist(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "model.xlsx"
            build_mapping_sample_workbook(workbook_path)
            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            financial_facts = {
                "reported_facts": {"营业收入": 260.0},
                "segment_disclosure": [],
            }
            mapping = {
                "segment_tree": {
                    "reportable_segments": [
                        {"segment_name": "手机业务", "reported_value": 140.0},
                        {"segment_name": "汽车业务", "reported_value": 48.0},
                    ]
                },
                "segment_mappings": [
                    {"workbook_segment": "手机业务", "tushare_segment": "", "mapping_type": "proxy", "confidence": "low", "workbook_reported_value": 140.0},
                    {"workbook_segment": "汽车业务", "tushare_segment": "", "mapping_type": "proxy", "confidence": "low", "workbook_reported_value": 48.0},
                ],
                "anchored_segment_count": 1,
                "proxy_segment_count": 2,
                "candidate_clusters": [
                    {
                        "workbook_segment": "手机业务",
                        "candidate_matches": [
                            {"tushare_segment": "消费电子业务", "mapping_type": "proxy", "confidence": "low", "score": 5, "base_score": 0, "tushare_revenue": 140.0, "source_refs": []},
                            {"tushare_segment": "移动终端业务", "mapping_type": "proxy", "confidence": "low", "score": 5, "base_score": 0, "tushare_revenue": 140.0, "source_refs": []},
                        ],
                    }
                ],
                "unmapped_tushare_segments": ["消费电子业务", "移动终端业务"],
            }
            audit = build_reconciliation_audit(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
                segment_mapping=mapping,
            )

            self.assertEqual(audit["resolution_mode"], "candidate_decision_required")
            self.assertIn("all_candidate_options_low_quality_proxy", audit["fail_reasons"])
            self.assertEqual(len(audit["candidate_options"]), 3)

    def test_reconciliation_audit_uses_anchored_revenue_total_not_workbook_total(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "model.xlsx"
            build_mapping_sample_workbook(workbook_path)
            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            financial_facts = {
                "reported_facts": {"营业收入": 240.0},
                "segment_disclosure": [
                    {"segment": "安防", "revenue": 18.0, "mapping_ready": True, "source_ref": "tushare:security"},
                ],
            }

            mapping = build_segment_mapping_contract(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
            )
            audit = build_reconciliation_audit(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
                segment_mapping=mapping,
            )

            self.assertEqual(audit["anchored_revenue_total"], 18.0)
            self.assertEqual(audit["mapped_revenue_total"], 18.0)
            self.assertEqual(audit["proxy_revenue_total"], 188.0)

    def test_exact_high_confidence_anchor_does_not_escalate_to_candidate_decision(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "power.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "营收拆分"
            ws["B2"] = "单位：亿元"
            ws["D2"] = 2024
            ws["E2"] = 2025
            ws["F2"] = "2026E"
            ws["B3"] = "新能源电力"
            ws["C3"] = "收入"
            ws["D3"] = 24.0
            ws["E3"] = 28.0
            ws["B4"] = "营业收入"
            ws["C4"] = "收入"
            ws["D4"] = 24.0
            ws["E4"] = 28.0
            wb.save(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            financial_facts = {
                "reported_facts": {"营业收入": 28.748485},
                "segment_disclosure": [
                    {"segment": "新能源电力", "revenue": 28.748485, "mapping_ready": True, "source_ref": "tushare:power"},
                    {"segment": "电力产品", "revenue": 10.272628, "mapping_ready": True, "source_ref": "tushare:power_sub"},
                    {"segment": "新能源汽车", "revenue": 17.06069, "mapping_ready": True, "source_ref": "tushare:ev"},
                ],
            }

            mapping = build_segment_mapping_contract(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
            )
            audit = build_reconciliation_audit(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                financial_facts=financial_facts,
                segment_mapping=mapping,
            )

            self.assertEqual(mapping["anchored_segment_count"], 1)
            self.assertEqual(mapping["candidate_clusters"], [])
            self.assertEqual(audit["resolution_mode"], "automatic_pass")


if __name__ == "__main__":
    unittest.main()
