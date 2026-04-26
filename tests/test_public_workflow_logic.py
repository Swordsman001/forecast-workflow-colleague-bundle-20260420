import sys
import tempfile
import unittest
import zipfile
import re
from pathlib import Path
from unittest import mock

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.forecast_tools.providers import AlphaPaiProvider, ProviderResult  # type: ignore
from scripts.forecast_tools.rollforward import (  # type: ignore
    AnnualReportFactExtractor,
    CompanyConfigManager,
    WorkbookBlueprintParser,
    build_evidence_payload,
    build_forecast_architecture,
    build_forecast_basis_payload,
    build_workbook_map_contract,
    materialize_forecast_basis_sheet,
    review_forecast_architecture,
    review_forecast_inputs,
    render_forecast_basis_markdown,
)
from scripts.forecast_tools.build_cell_instructions import build_cell_instructions, validate_forecast_basis, validate_workbook_map  # type: ignore


def build_contract_sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "营收拆分"
    ws["B2"] = "单位：亿元"
    ws["C2"] = 2023
    ws["D2"] = 2024
    ws["E2"] = "2025E"
    ws["F2"] = "2026E"
    ws["G2"] = "2027E"
    labels = [
        "营业收入",
        "毛利率",
        "归母净利润",
        "销售费用率（%）",
        "1）手机营收（亿元）",
        "2）汽车营收（亿元）",
    ]
    for row, label in enumerate(labels, start=3):
        ws.cell(row, 2).value = label
    values = {
        "营业收入": [210.0, 220.0, 230.0, 250.0, 270.0],
        "毛利率": [0.22, 0.23, 0.24, 0.245, 0.25],
        "归母净利润": [5.2, 6.1, 7.0, 8.2, 9.5],
        "销售费用率（%）": [0.031, 0.03, 0.029, 0.028, 0.027],
        "1）手机营收（亿元）": [120.0, 124.0, 128.0, 135.0, 141.0],
        "2）汽车营收（亿元）": [20.0, 24.0, 30.0, 38.0, 45.0],
    }
    for row in range(3, 9):
        label = ws.cell(row, 2).value
        for col, value in enumerate(values[label], start=3):
            ws.cell(row, col).value = value
    wb.save(path)


class PublicWorkflowLogicTests(unittest.TestCase):
    def test_rollforward_exports_one_canonical_segment_mapping_public_helper(self) -> None:
        source = (ROOT / "scripts" / "forecast_tools" / "rollforward.py").read_text(encoding="utf-8")
        self.assertEqual(len(re.findall(r"^def build_model_segment_tree\(", source, flags=re.MULTILINE)), 1)
        self.assertEqual(len(re.findall(r"^def build_segment_mapping_contract\(", source, flags=re.MULTILINE)), 1)
        self.assertEqual(len(re.findall(r"^def build_reconciliation_audit\(", source, flags=re.MULTILINE)), 1)

    def test_annual_report_extractor_returns_structured_financial_facts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            report = Path(tmpdir) / "annual_report.txt"
            report.write_text(
                "\n".join(
                    [
                        "第8页 主要财务指标",
                        "营业收入 288.55亿元",
                        "毛利 98.10亿元",
                        "毛利率 34.0%",
                        "归属于上市公司股东的净利润 45.20亿元",
                        "扣除非经常性损益后的净利润 40.10亿元",
                        "销售费用 12.30亿元 销售费用率 4.3%",
                        "管理费用 8.40亿元 管理费用率 2.9%",
                        "研发费用 21.10亿元 研发费用率 7.3%",
                        "财务费用 -0.80亿元 财务费用率 -0.3%",
                        "实际税率 15.0%",
                        "分业务情况",
                        "手机 180.00亿元 毛利率 32.0%",
                        "汽车 52.00亿元 毛利率 28.0%",
                    ]
                ),
                encoding="utf-8",
            )

            result = AnnualReportFactExtractor().extract(report)

            self.assertEqual(result["reported_facts"]["营业收入"], 288.55)
            self.assertEqual(result["reported_facts"]["毛利"], 98.10)
            self.assertEqual(result["reported_facts"]["毛利率"], 0.34)
            self.assertEqual(result["reported_facts"]["销售费用"], 12.30)
            self.assertEqual(result["reported_facts"]["研发费用率（%）"], 0.073)
            self.assertEqual(result["reported_facts"]["所得税税率（%）"], 0.15)
            fact_items = result.get("fact_items", [])
            self.assertTrue(any(item.get("metric") == "营业收入" and item.get("unit") == "亿元" for item in fact_items))
            self.assertTrue(any(item.get("page_reference") == "第8页" for item in fact_items))
            segment_items = result.get("segment_disclosure", [])
            self.assertTrue(any(item.get("segment") == "手机" and item.get("revenue") == 180.0 for item in segment_items))
            self.assertTrue(any(item.get("segment") == "汽车" and item.get("gross_margin") == 0.28 for item in segment_items))

    def test_annual_report_extractor_reads_segment_disclosure_from_pdf_tables(self) -> None:
        class FakePage:
            def __init__(self, text: str, tables: list[list[list[str | None]]]) -> None:
                self._text = text
                self._tables = tables

            def extract_text(self) -> str:
                return self._text

            def extract_tables(self) -> list[list[list[str | None]]]:
                return self._tables

        class FakePdf:
            def __init__(self, pages: list[FakePage]) -> None:
                self.pages = pages

            def __enter__(self) -> "FakePdf":
                return self

            def __exit__(self, exc_type, exc, tb) -> bool:
                return False

        with tempfile.TemporaryDirectory() as tmpdir:
            report = Path(tmpdir) / "annual_report.pdf"
            report.write_bytes(b"%PDF-1.4")
            fake_pdf = FakePdf(
                [
                    FakePage(
                        "第8页\n主要会计数据",
                        [
                            [
                                ["主要会计数据", "2025年", "2024年", "本期比上年同期增减(%)", "2023年"],
                                ["营业收入", "28,854,805,517.49", "25,730,639,138.12", "12.14", "21,020,641,622.86"],
                                ["利润总额", "4,600,820,158.17", "3,278,468,271.39", "40.33", "691,446,265.33"],
                                ["归属于上市公司股东的净利润", "4,045,416,530.33", "3,323,242,749.90", "21.73", "555,623,916.73"],
                                ["归属于上市公司股东的扣除非经常性损益的净利润", "3,909,518,597.14", "3,056,526,805.13", "27.91", "138,009,609.54"],
                            ]
                        ],
                    ),
                    FakePage(
                        "第35页\n科目",
                        [
                            [
                                ["科目", "本期数", "上年同期数", "变动比例（%）"],
                                ["营业收入", "28,854,805,517.49", "25,730,639,138.12", "12.14"],
                                ["营业成本", "20,015,173,377.35", "18,154,402,520.87", "10.25"],
                                ["销售费用", "564,341,702.89", "556,748,830.55", "1.36"],
                                ["管理费用", "696,392,890.54", "748,445,193.45", "-6.95"],
                                ["研发费用", "2,842,872,872.68", "2,622,086,780.18", "8.42"],
                                ["财务费用", "-73,679,924.87", "-12,924,592.39", "-470.08"],
                            ]
                        ],
                    ),
                    FakePage(
                        "第37页\n主营业务分产品情况\n图像传感器解决方案业务",
                        [
                            [
                                ["主营业务分产品情况", None, None, None, None, None, None],
                                ["分产品", "营业收入", "营业成本", "毛利率（%）", "营业收入比上年增减（%）", "营业成本比上年增减（%）", "毛利率比上年增减（%）"],
                                ["图像传感器解决方案业务", "21,245,839,825.85", "13,601,778,163.68", "35.98", "10.71", "8.24", "增加1.46个百分点"],
                                ["显示解决方案业务", "941,127,046.72", "806,606,147.97", "14.29", "-8.47", "-14.62", "增加6.17个百分点"],
                                ["合计", "28,814,589,212.28", "19,999,857,503.62", "30.59", "12.25", "10.30", "增加1.23个百分点"],
                            ]
                        ],
                    )
                ]
            )

            with mock.patch("scripts.forecast_tools.rollforward.pdfplumber.open", return_value=fake_pdf):
                result = AnnualReportFactExtractor().extract(report)

            self.assertEqual(result["reported_facts"]["归母净利润"], 4045416530.33)
            self.assertEqual(result["reported_facts"]["扣非归母净利润"], 3909518597.14)
            self.assertAlmostEqual(result["reported_facts"]["毛利"], 8839632140.14, places=2)
            self.assertAlmostEqual(result["reported_facts"]["毛利率"], 0.306349, places=6)
            self.assertAlmostEqual(result["reported_facts"]["销售费用率（%）"], 0.019558, places=6)
            segment_items = result.get("segment_disclosure", [])
            self.assertTrue(any(item.get("segment") == "图像传感器解决方案业务" and item.get("category") == "product" for item in segment_items))
            self.assertTrue(any(item.get("segment") == "图像传感器解决方案业务" and abs(item.get("revenue") - 212.458398) < 0.001 for item in segment_items))
            self.assertTrue(any(item.get("segment") == "显示解决方案业务" and item.get("gross_margin") == 0.1429 for item in segment_items))

    def test_alpha_provider_default_recall_types_restrict_to_notes_and_comments(self) -> None:
        self.assertEqual(
            AlphaPaiProvider.DEFAULT_RECALL_TYPES,
            ["roadShow", "roadShow_ir", "roadShow_us", "comment"],
        )
        self.assertNotIn("report", AlphaPaiProvider.DEFAULT_RECALL_TYPES)
        self.assertNotIn("foreign_report", AlphaPaiProvider.DEFAULT_RECALL_TYPES)
        self.assertNotIn("qa", AlphaPaiProvider.DEFAULT_RECALL_TYPES)

    def test_alpha_provider_returns_recalled_content_when_client_available(self) -> None:
        provider = AlphaPaiProvider()
        fake_items = [
            {
                "id": "alpha-1",
                "contextInfo": "发布时间: 2026-04-01,机构: 某券商,标题: 手机CIS份额提升",
                "chunks": ["手机CIS 高端导入，ASP改善，份额提升。"],
            }
        ]
        with mock.patch(
            "scripts.forecast_tools.providers.AlphaPaiProvider._recall_records",
            return_value=fake_items,
        ):
            result = provider.fetch("韦尔股份")

        self.assertEqual(result.source_type, "alpha_pai")
        self.assertIn("手机CIS", result.content)
        self.assertTrue(result.metadata["available"])
        self.assertEqual(result.metadata["record_count"], 1)

    def test_alpha_provider_retries_timeout_before_succeeding(self) -> None:
        provider = AlphaPaiProvider()
        fake_items = [
            {
                "id": "alpha-1",
                "contextInfo": "标题: 手机CIS",
                "chunks": ["手机CIS ASP改善"],
            }
        ]
        with mock.patch(
            "scripts.forecast_tools.providers.AlphaPaiProvider._recall_records",
            side_effect=[TimeoutError("timeout"), fake_items],
        ) as recall_mock:
            result = provider.fetch("豪威集团 手机CIS", timeout_schedule=(1, 2))

        self.assertTrue(result.metadata["available"])
        self.assertEqual(recall_mock.call_count, 2)
        self.assertEqual(result.metadata["attempts"][0]["status"], "error")
        self.assertEqual(result.metadata["attempts"][1]["status"], "ok")

    def test_build_evidence_payload_reads_docx_reference_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            meeting = tmp / "meeting.docx"
            research = tmp / "research.txt"
            with zipfile.ZipFile(meeting, "w") as archive:
                archive.writestr(
                    "word/document.xml",
                    (
                        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
                        "<w:document xmlns:w='http://schemas.openxmlformats.org/wordprocessingml/2006/main'>"
                        "<w:body><w:p><w:r><w:t>手机CIS 2025年增长20%</w:t></w:r></w:p></w:body>"
                        "</w:document>"
                    ),
                )
            research.write_text("research note", encoding="utf-8")

            with mock.patch(
                "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
            ):
                payload = build_evidence_payload(
                    repo_root=ROOT,
                    company="韦尔股份",
                    meeting_notes=meeting,
                    research_report=research,
                )

            meeting_provider = next(
                item
                for item in payload["providers"]
                if item.get("metadata", {}).get("reference_kind") == "meeting_notes"
            )
            self.assertIn("手机CIS", meeting_provider["content"])

    def test_build_evidence_payload_records_priority_and_source_tiers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            meeting = tmp / "meeting.txt"
            research = tmp / "research.txt"
            meeting.write_text("meeting note", encoding="utf-8")
            research.write_text("research note", encoding="utf-8")

            with mock.patch(
                "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
            ):
                payload = build_evidence_payload(
                    repo_root=ROOT,
                    company="韦尔股份",
                    meeting_notes=meeting,
                    research_report=research,
                )

            self.assertEqual(payload["evidence_priority"], ["reference_files", "local_kb", "alpha_pai"])
            self.assertIn("provider_decisions", payload)
            self.assertTrue(any(item["source_tier"] == "reference_files" for item in payload["providers"]))

    def test_build_evidence_payload_mandatorily_recalls_alpha_pai_and_records_stage_checks(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            meeting = tmp / "meeting.txt"
            research = tmp / "research.txt"
            meeting.write_text("meeting note", encoding="utf-8")
            research.write_text("research note", encoding="utf-8")

            with mock.patch(
                "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
            ) as alpha_fetch:
                payload = build_evidence_payload(
                    repo_root=ROOT,
                    company="闊﹀皵鑲′唤",
                    meeting_notes=meeting,
                    research_report=research,
                )

            alpha_fetch.assert_called_once()
            self.assertEqual(alpha_fetch.call_args.args[0], "闊﹀皵鑲′唤")
            self.assertEqual(
                alpha_fetch.call_args.kwargs.get("recall_types"),
                ["roadShow", "roadShow_ir", "roadShow_us", "comment"],
            )
            alpha_decision = next(item for item in payload["provider_decisions"] if item["source_tier"] == "alpha_pai")
            self.assertEqual(alpha_decision["decision"], "mandatory_segment_recall_executed")
            recall_checks = payload.get("recall_checks", {}).get("source_prep", [])
            self.assertTrue(any(item["name"] == "alpha_pai_recalled" and item["passed"] for item in recall_checks))
            self.assertTrue(any(item["name"] == "alpha_pai_has_content" and not item["passed"] for item in recall_checks))

    def test_build_evidence_payload_batches_alpha_queries_for_qualifying_segments(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook = tmp / "model.xlsx"
            meeting = tmp / "meeting.txt"
            research = tmp / "research.txt"
            meeting.write_text("meeting note", encoding="utf-8")
            research.write_text("research note", encoding="utf-8")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "segment"
            ws["B1"] = "metric"
            ws["C1"] = 2024
            ws["D1"] = 2025
            ws["E1"] = "2026E"
            ws["A2"] = "手机CIS"
            ws["B2"] = "收入"
            ws["C2"] = 30.0
            ws["D2"] = 35.0
            ws["A3"] = "汽车CIS"
            ws["B3"] = "收入"
            ws["C3"] = 6.0
            ws["D3"] = 10.0
            ws["A4"] = "安防"
            ws["B4"] = "收入"
            ws["C4"] = 9.0
            ws["D4"] = 10.0
            wb.save(workbook)

            blueprint = WorkbookBlueprintParser().parse(workbook)
            with mock.patch(
                "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                side_effect=lambda query, **kwargs: ProviderResult("alpha_pai", f"alphapai:{query}", "", {"available": False, "query": query, "kwargs": kwargs}),
            ) as alpha_fetch:
                payload = build_evidence_payload(
                    repo_root=ROOT,
                    company="豪威集团",
                    meeting_notes=meeting,
                    research_report=research,
                    workbook_path=workbook,
                    blueprint=blueprint,
                    annual_report_facts={"reported_facts": {"营业收入": 100.0}},
                )

            called_queries = [call.args[0] for call in alpha_fetch.call_args_list]
            self.assertEqual(len(called_queries), 2)
            self.assertTrue(any(item.startswith("豪威集团 手机") for item in called_queries))
            self.assertTrue(any(item.startswith("豪威集团 汽车") for item in called_queries))
            self.assertFalse(any("安防" in item for item in called_queries))
            self.assertTrue(any(item["name"] == "alpha_pai_call_count" and item["detail"] == 2 for item in payload["recall_checks"]["source_prep"]))

    def test_review_forecast_inputs_rejects_alpha_priority_violation(self) -> None:
        logic_review = review_forecast_inputs(
            annual_report_facts={"reported_facts": {"营业收入": 240.0}},
            meeting_notes_facts={
                "bridge_facts": {
                    "1）手机营收（亿元）": {
                        "claim": "alpha only",
                        "source_type": "alpha_pai",
                        "source_ref": "alpha",
                        "used_for": "forecast_growth",
                        "confidence": 0.4,
                        "review_required": False,
                        "growth_rate": 0.1,
                    }
                }
            },
            evidence_payload={
                "providers": [
                    {"source_type": "local_kb", "content": "usable kb", "source_tier": "local_kb"},
                    {"source_type": "alpha_pai", "content": "alpha clue", "source_tier": "alpha_pai"},
                ]
            },
        )
        self.assertFalse(logic_review["passed"])
        self.assertTrue(logic_review["source_priority_warnings"])

    def test_build_forecast_architecture_produces_segment_cards_and_source_tiers(self) -> None:
        architecture = build_forecast_architecture(
            company="韦尔股份",
            report_year=2025,
            annual_report_facts={"reported_facts": {"营业收入": 240.0}},
            meeting_notes_facts={
                "bridge_facts": {
                    "1）手机营收（亿元）": {
                        "claim": "手机CIS 2026年增长10%",
                        "source_type": "meeting_notes",
                        "source_ref": "meeting-note",
                        "used_for": "forecast_growth",
                        "confidence": 0.7,
                        "review_required": False,
                        "growth_rate": 0.1,
                    }
                }
            },
            evidence_payload={
                "providers": [
                    {
                        "source_ref": "meeting-note",
                        "source_type": "user_supplied_text",
                        "source_tier": "reference_files",
                        "content": "手机CIS 新品导入提升 ASP 和份额。",
                        "metadata": {"reference_kind": "meeting_notes"},
                    },
                    {
                        "source_ref": "kb-note",
                        "source_type": "local_kb",
                        "source_tier": "local_kb",
                        "content": "高像素升级带来 ASP 改善。",
                        "metadata": {},
                    },
                ]
            },
        )

        self.assertTrue(architecture["segments"])
        segment = architecture["segments"][0]
        self.assertEqual(segment["source_tier"], "reference_files")
        self.assertTrue(segment["kill_conditions"])
        self.assertIn("volume_logic", segment)
        self.assertIn("asp_logic", segment)
        self.assertIn("share_logic", segment)
        self.assertIn("margin_logic", segment)

    def test_build_forecast_architecture_marks_small_high_growth_segment_as_material(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook = tmp / "model.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "segment"
            ws["B1"] = "metric"
            ws["C1"] = 2024
            ws["D1"] = 2025
            ws["E1"] = "2026E"
            ws["F1"] = "2027E"
            ws["A2"] = "主业务"
            ws["B2"] = "收入"
            ws["C2"] = 80.0
            ws["D2"] = 100.0
            ws["E2"] = 110.0
            ws["F2"] = 121.0
            ws["A3"] = "高增小业务"
            ws["B3"] = "收入"
            ws["C3"] = 5.0
            ws["D3"] = 8.0
            ws["E3"] = 9.6
            ws["F3"] = 11.52
            wb.save(workbook)

            blueprint = WorkbookBlueprintParser().parse(workbook)
            architecture = build_forecast_architecture(
                company="示例公司",
                report_year=2025,
                annual_report_facts={"reported_facts": {"营业收入": 108.0}},
                meeting_notes_facts={"bridge_facts": {}},
                evidence_payload={"providers": []},
                workbook_path=workbook,
                blueprint=blueprint,
            )

            segment = next(item for item in architecture["segments"] if item["segment"] == "高增小业务")
            self.assertLess(segment.get("reported_share") or 0, 0.3)
            self.assertGreater(segment.get("reported_year_growth") or 0, 0.4)
            self.assertTrue(segment.get("material_segment"))

    def test_logic_review_allows_stable_small_low_growth_segment(self) -> None:
        logic_review = review_forecast_architecture(
            forecast_architecture={
                "segments": [
                    {
                        "segment": "其他",
                        "revenue_driver": "documented growth bridge",
                        "years_covered": ["2026E", "2027E", "2028E"],
                        "forecast_values": {"2026E": 1.60, "2027E": 1.60, "2028E": 1.60},
                        "volume_logic": {"mechanism": "维持稳定。"},
                        "asp_logic": {"mechanism": "维持稳定。"},
                        "share_logic": {"mechanism": "维持稳定。", "values": {"2026E": 0.10, "2027E": 0.10, "2028E": 0.10}},
                        "margin_logic": {"mechanism": "维持稳定。", "values": {"2026E": 0.36, "2027E": 0.36, "2028E": 0.36}},
                        "kill_conditions": ["需求进一步走弱"],
                        "reported_share": 0.02,
                        "reported_year_growth": 0.02,
                        "material_segment": False,
                    }
                ]
            },
            annual_report_facts={"reported_facts": {"营业收入": 288.0}},
            evidence_payload={"providers": [{"source_tier": "alpha_pai", "content": "alpha", "source_ref": "alpha"}]},
        )

        self.assertEqual(logic_review["pass_or_fail"], "pass")
        self.assertFalse(
            any("flat FY1-FY3 growth trajectory" in item for item in logic_review["future_year_coverage_warnings"])
        )

    def test_logic_review_rejects_flat_high_growth_segment(self) -> None:
        logic_review = review_forecast_architecture(
            forecast_architecture={
                "segments": [
                    {
                        "segment": "高增小业务",
                        "revenue_driver": "documented growth bridge",
                        "years_covered": ["2026E", "2027E", "2028E"],
                        "forecast_values": {"2026E": 9.6, "2027E": 11.52, "2028E": 13.824},
                        "volume_logic": {"mechanism": "放量。"},
                        "asp_logic": {"mechanism": "ASP改善。"},
                        "share_logic": {"mechanism": "份额提升。", "values": {"2026E": 0.12, "2027E": 0.12, "2028E": 0.12}},
                        "margin_logic": {"mechanism": "高附加值提升。", "values": {"2026E": 0.30, "2027E": 0.30, "2028E": 0.30}},
                        "kill_conditions": ["导入不及预期"],
                        "reported_share": 0.08,
                        "reported_year_growth": 0.60,
                        "material_segment": True,
                    }
                ]
            },
            annual_report_facts={"reported_facts": {"营业收入": 108.0}},
            evidence_payload={"providers": [{"source_tier": "alpha_pai", "content": "alpha", "source_ref": "alpha"}]},
        )

        self.assertEqual(logic_review["pass_or_fail"], "fail")
        self.assertTrue(
            any("flat FY1-FY3 growth trajectory" in item for item in logic_review["future_year_coverage_warnings"])
        )

    def test_build_workbook_map_contract_emits_validator_compatible_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "model.xlsx"
            build_contract_sample_workbook(workbook)
            blueprint = WorkbookBlueprintParser().parse(workbook)

            workbook_map = build_workbook_map_contract(
                workbook_path=workbook,
                blueprint=blueprint,
            )

            validate_workbook_map(workbook_map)
            first_row = workbook_map["row_registry"][0]
            self.assertIn("row_id", first_row)
            self.assertIn("role", first_row)
            self.assertIn("required_years", first_row)
            self.assertIn("must_extend_to_far_year", first_row)
            self.assertIsInstance(workbook_map["current_forecast_window"], list)
            self.assertIsInstance(workbook_map["map_validation_hints"], dict)

    def test_company_config_manager_uses_current_workbook_segments_in_default_config(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            workbook = repo_root / "model.xlsx"
            build_contract_sample_workbook(workbook)
            blueprint = WorkbookBlueprintParser().parse(workbook)

            config_path = CompanyConfigManager(repo_root).ensure_config("样本公司", blueprint)
            content = config_path.read_text(encoding="utf-8")

            self.assertIn("1）手机营收（亿元）", content)
            self.assertIn("2）汽车营收（亿元）", content)
            self.assertNotIn("6）新兴市场/物联网营收（亿元）", content)

    def test_workbook_map_contract_and_basis_compile_into_instructions(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "model.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "营收拆分"
            ws["B2"] = "单位：亿元"
            ws["C2"] = 2023
            ws["D2"] = 2024
            ws["E2"] = "2025E"
            ws["F2"] = "2026E"
            ws["G2"] = "2027E"
            ws["B3"] = "手机业务收入"
            ws["B4"] = "汽车业务收入"
            for col, values in {
                3: (120.0, 20.0),
                4: (124.0, 24.0),
                5: (128.0, 30.0),
                6: (135.0, 38.0),
                7: (141.0, 45.0),
            }.items():
                ws.cell(3, col).value = values[0]
                ws.cell(4, col).value = values[1]
            wb.save(workbook)
            blueprint = WorkbookBlueprintParser().parse(workbook)

            architecture = build_forecast_architecture(
                company="示例公司",
                report_year=2024,
                annual_report_facts={"reported_facts": {"营业收入": 220.0}},
                meeting_notes_facts={
                    "bridge_facts": {
                        "手机业务收入": {
                            "claim": "手机业务 2026 年保持增长。",
                            "source_type": "meeting_notes",
                            "source_ref": "meeting-notes",
                            "used_for": "forecast_growth",
                            "confidence": 0.7,
                            "review_required": False,
                            "growth_rate": 0.08,
                        }
                    }
                },
                evidence_payload={
                    "providers": [
                        {
                            "source_ref": "meeting-notes",
                            "source_type": "user_supplied_text",
                            "source_tier": "reference_files",
                            "content": "手机业务 2026 年保持增长，份额稳定。",
                            "metadata": {"reference_kind": "meeting_notes"},
                        }
                    ]
                },
                workbook_path=workbook,
                blueprint=blueprint,
            )
            basis = build_forecast_basis_payload(
                company="示例公司",
                report_year=2024,
                annual_report_facts={
                    "reported_facts": {"营业收入": 220.0},
                    "fact_items": [
                        {
                            "metric": "营业收入",
                            "value": 220.0,
                            "unit": "亿元",
                            "page_reference": "第8页",
                            "file_reference": "annual_report.pdf",
                            "note": None,
                        }
                    ],
                },
                meeting_notes_facts={
                    "bridge_facts": {
                        "手机业务收入": {
                            "claim": "手机业务 2026 年保持增长。",
                            "source_type": "meeting_notes",
                            "source_ref": "meeting-notes",
                            "used_for": "forecast_growth",
                            "confidence": 0.7,
                            "review_required": False,
                            "growth_rate": 0.08,
                        }
                    }
                },
                evidence_payload={
                    "providers": [
                        {
                            "source_ref": "meeting-notes",
                            "source_type": "user_supplied_text",
                            "source_tier": "reference_files",
                            "content": "手机业务 2026 年保持增长，份额稳定。",
                            "metadata": {"reference_kind": "meeting_notes"},
                        }
                    ]
                },
                forecast_architecture=architecture,
            )
            workbook_map = build_workbook_map_contract(
                workbook_path=workbook,
                blueprint=blueprint,
                report_year=2024,
                target_far_year=2027,
            )
            phone_row = next(
                row for row in workbook_map["row_registry"] if "手机" in str(row.get("label") or "")
            )
            workbook_map["row_registry"] = [phone_row]
            workbook_map["writable_driver_targets"] = [phone_row["row_id"]]
            workbook_map["formula_rows"] = []
            workbook_map["display_rows"] = []

            instructions = build_cell_instructions(workbook_map, basis)

            self.assertTrue(instructions["instructions"])
            self.assertTrue(any(item["write_type"] == "value" for item in instructions["instructions"]))
            self.assertEqual(phone_row["required_years"], ["2025E", "2026E", "2027E"])

    def test_build_forecast_architecture_uses_workbook_segment_names_not_legacy_cis_tags(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook = tmp / "model.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "单位：亿元"
            ws["C1"] = "2023A"
            ws["D1"] = "2024A"
            ws["E1"] = 2025
            ws["F1"] = "2026E"
            ws["G1"] = "2027E"
            ws["H1"] = "2028E"
            ws["A2"] = "高速铜缆"
            ws["B2"] = "收入"
            ws["E2"] = 12.0
            ws["A3"] = "新能源汽车"
            ws["B3"] = "收入"
            ws["E3"] = 8.0
            ws["A4"] = "电力"
            ws["B4"] = "收入"
            ws["E4"] = 6.0
            wb.save(workbook)

            blueprint = WorkbookBlueprintParser().parse(workbook)
            architecture = build_forecast_architecture(
                company="沃尔核材",
                report_year=2025,
                annual_report_facts={"reported_facts": {"营业收入": 84.5}},
                meeting_notes_facts={"bridge_facts": {}},
                evidence_payload={
                    "providers": [
                        {
                            "source_ref": "meeting-note",
                            "source_type": "user_supplied_text",
                            "source_tier": "reference_files",
                            "content": (
                                "高速铜缆收入增长接近50%，224G和448G持续放量。"
                                "新能源汽车业务同比增长23.5%。"
                                "电力板块同比增长10%。"
                            ),
                            "metadata": {"reference_kind": "meeting_notes"},
                        }
                    ]
                },
                workbook_path=workbook,
                blueprint=blueprint,
            )

            segment_names = [item["segment"] for item in architecture["segments"]]
            self.assertIn("高速铜缆", segment_names)
            self.assertIn("新能源汽车", segment_names)
            self.assertIn("电力", segment_names)
            self.assertFalse(any("CIS" in name for name in segment_names))

    def test_review_forecast_architecture_flags_missing_mechanism_chain(self) -> None:
        architecture = {
            "company": "韦尔股份",
            "segments": [
                {
                    "segment": "手机CIS",
                    "years_covered": ["2026E", "2027E", "2028E"],
                    "revenue_driver": "",
                    "volume_logic": {"mechanism": "", "evidence_refs": []},
                    "asp_logic": {"mechanism": "", "evidence_refs": []},
                    "share_logic": {"mechanism": "", "evidence_refs": []},
                    "margin_logic": {"mechanism": "", "evidence_refs": []},
                    "kill_conditions": [],
                    "source_tier": "alpha_pai",
                    "source_ref": ["alpha-1"],
                }
            ]
        }
        review = review_forecast_architecture(
            forecast_architecture=architecture,
            annual_report_facts={"reported_facts": {"营业收入": 240.0}},
            evidence_payload={
                "providers": [
                    {"source_tier": "reference_files", "content": "annual", "source_ref": "annual-1"},
                    {"source_tier": "local_kb", "content": "kb", "source_ref": "kb-1"},
                    {"source_tier": "alpha_pai", "content": "alpha", "source_ref": "alpha-1"},
                ]
            },
        )
        self.assertEqual(review["pass_or_fail"], "fail")
        self.assertTrue(review["missing_mechanism_links"])
        self.assertTrue(review["missing_kill_conditions"])
        self.assertTrue(review["source_priority_warnings"])

    def test_review_forecast_architecture_rejects_flat_future_trajectory_without_stability_rationale(self) -> None:
        architecture = {
            "company": "测试公司",
            "segments": [
                {
                    "segment": "segment_a",
                    "years_covered": ["2026E", "2027E", "2028E"],
                    "revenue_driver": "shipments x ASP",
                    "forecast_values": {"2026E": 100.0, "2027E": 110.0, "2028E": 121.0},
                    "volume_logic": {"mechanism": "放量", "evidence_refs": []},
                    "asp_logic": {"mechanism": "ASP提升", "evidence_refs": []},
                    "share_logic": {"mechanism": "客户导入", "evidence_refs": []},
                    "margin_logic": {
                        "mechanism": "结构升级",
                        "evidence_refs": [],
                        "values": {"2026E": 0.30, "2027E": 0.30, "2028E": 0.30},
                    },
                    "dependent_metric_values": {
                        "share": {"2026E": 0.50, "2027E": 0.50, "2028E": 0.50}
                    },
                    "kill_conditions": ["需求不及预期"],
                    "source_tier": "reference_files",
                    "source_ref": ["annual"],
                    "reported_share": 0.35,
                    "reported_year_growth": 0.10,
                    "material_segment": True,
                }
            ],
        }
        review = review_forecast_architecture(
            forecast_architecture=architecture,
            annual_report_facts={"reported_facts": {"营业收入": 240.0}},
            evidence_payload={"providers": [{"source_tier": "reference_files", "content": "annual", "source_ref": "annual"}]},
        )
        self.assertEqual(review["pass_or_fail"], "fail")
        self.assertTrue(any("flat FY1-FY3 growth trajectory" in item for item in review["future_year_coverage_warnings"]))

    def test_build_forecast_architecture_prefers_existing_workbook_schedule_when_only_non_growth_percentages_exist(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook = tmp / "model.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "segment"
            ws["B1"] = "metric"
            ws["C1"] = 2023
            ws["D1"] = 2024
            ws["E1"] = 2025
            ws["F1"] = "2026E"
            ws["G1"] = "2027E"
            ws["A2"] = "新能源电力"
            ws["B2"] = "收入"
            ws["E2"] = 28.0
            ws["F2"] = 32.0
            ws["G2"] = 36.4
            ws["A3"] = None
            ws["B3"] = "yoy"
            ws["F3"] = 0.14
            ws["G3"] = 0.1375
            wb.save(workbook)

            blueprint = WorkbookBlueprintParser().parse(workbook)
            architecture = build_forecast_architecture(
                company="沃尔核材",
                report_year=2025,
                annual_report_facts={"reported_facts": {"营业收入": 84.8}},
                meeting_notes_facts={"bridge_facts": {}},
                evidence_payload={
                    "providers": [
                        {
                            "source_ref": "meeting-note",
                            "source_type": "user_supplied_text",
                            "source_tier": "reference_files",
                            "content": "新能源电力当前市场份额50%，毛利率31%，订单延续。",
                            "metadata": {"reference_kind": "meeting_notes"},
                        }
                    ]
                },
                workbook_path=workbook,
                blueprint=blueprint,
            )

            segment = next(item for item in architecture["segments"] if item["segment"] == "新能源电力")
            self.assertAlmostEqual(segment["forecast_values"]["2026E"], 32.0, places=4)
            self.assertAlmostEqual(segment["forecast_values"]["2027E"], 36.4, places=4)
            self.assertGreater(segment["forecast_values"]["2028E"], 36.4)

    def test_review_forecast_architecture_rejects_explosive_revenue_trajectory(self) -> None:
        architecture = {
            "company": "沃尔核材",
            "segments": [
                {
                    "segment": "高速铜缆",
                    "years_covered": ["2026E", "2027E", "2028E"],
                    "revenue_driver": "shipments x ASP",
                    "forecast_values": {"2026E": 149.83, "2027E": 223.7411, "2028E": 332.9939},
                    "volume_logic": {"mechanism": "订单放量", "evidence_refs": ["meeting-1"]},
                    "asp_logic": {"mechanism": "ASP提升", "evidence_refs": ["meeting-1"]},
                    "share_logic": {"mechanism": "份额提升", "evidence_refs": ["meeting-1"]},
                    "margin_logic": {"mechanism": "规模效应", "evidence_refs": ["meeting-1"]},
                    "kill_conditions": ["客户拉货不及预期"],
                    "source_tier": "reference_files",
                    "source_ref": ["meeting-1"],
                }
            ],
        }
        review = review_forecast_architecture(
            forecast_architecture=architecture,
            annual_report_facts={"reported_facts": {"营业收入": 84.8}},
            evidence_payload={"providers": [{"source_tier": "reference_files", "content": "meeting", "source_ref": "meeting-1"}]},
        )
        self.assertEqual(review["pass_or_fail"], "fail")
        self.assertTrue(any("explosive" in item for item in review["future_year_coverage_warnings"]))

    def test_review_forecast_architecture_rejects_weak_mapping_and_reconciliation_layer(self) -> None:
        architecture = {
            "company": "测试公司",
            "segments": [
                {
                    "segment": "手机CIS",
                    "years_covered": ["2026E", "2027E", "2028E"],
                    "revenue_driver": "market x share x ASP",
                    "forecast_values": {"2026E": 100.0, "2027E": 110.0, "2028E": 121.0},
                    "volume_logic": {"mechanism": "销量增长", "evidence_refs": ["meeting-1"]},
                    "asp_logic": {"mechanism": "ASP提升", "evidence_refs": ["meeting-1"]},
                    "share_logic": {"mechanism": "份额提升", "evidence_refs": ["meeting-1"]},
                    "margin_logic": {"mechanism": "结构优化", "evidence_refs": ["meeting-1"], "values": {"2026E": 0.3, "2027E": 0.31, "2028E": 0.32}},
                    "kill_conditions": ["需求不及预期"],
                    "source_tier": "reference_files",
                    "source_ref": ["meeting-1"],
                    "reported_share": 0.4,
                    "reported_year_growth": 0.12,
                }
            ],
        }
        review = review_forecast_architecture(
            forecast_architecture=architecture,
            annual_report_facts={"reported_facts": {"营业收入": 240.0}},
            evidence_payload={"providers": [{"source_tier": "reference_files", "content": "meeting", "source_ref": "meeting-1"}]},
            segment_mapping={"workbook_segment_count": 3, "anchored_segment_count": 0, "segment_mappings": []},
            reconciliation_audit={
                "within_tolerance": False,
                "resolution_mode": "candidate_decision_required",
                "fail_reasons": ["no_tushare_segment_mappings"],
            },
        )

        self.assertEqual(review["pass_or_fail"], "fail")
        self.assertTrue(any(item["name"] == "segment_mapping_anchored" and not item["passed"] for item in review["checks"]))
        self.assertTrue(any(item["name"] == "reconciliation_ready" and not item["passed"] for item in review["checks"]))
        self.assertTrue(any("reconciliation" in item.lower() or "mapping" in item.lower() for item in review["must_fix_before_phase_b"]))

    def test_materialize_forecast_basis_sheet_writes_granular_driver_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook = tmp / "book.xlsx"
            wb = openpyxl.Workbook()
            wb.save(workbook)
            basis = {
                "company": "韦尔股份",
                "cutoff_date": "2025-05-31",
                "reported_year": "2024A",
                "target_window": ["2025E", "2026E"],
                "facts": [],
                "assumptions": [],
                "segment_assumption_cards": [
                    {
                        "segment": "手机CIS",
                        "year": "2025E",
                        "metric": "revenue",
                        "value": 100.0,
                        "driver_form": "market x share x ASP",
                        "volume_logic": {"mechanism": "销量增长", "evidence_refs": ["ev1"]},
                        "asp_logic": {"mechanism": "ASP提升", "evidence_refs": ["ev2"]},
                        "share_logic": {"mechanism": "份额提升", "evidence_refs": ["ev3"]},
                        "margin_logic": {"mechanism": "毛利率改善", "evidence_refs": ["ev4"]},
                        "kill_conditions": ["需求不及预期"],
                        "weak_assumptions": ["新品导入顺利"],
                        "source_ref": ["ev1", "ev2"],
                        "confidence": "medium",
                    }
                ],
                "consolidated_logic": {},
            }
            evidence = [
                {"fact_id": "ev1", "source_file": "meeting.txt", "page_or_line": "p1", "text_summary": "销量增长"},
                {"fact_id": "ev2", "source_file": "research.txt", "page_or_line": "p2", "text_summary": "ASP提升"},
                {"fact_id": "ev3", "source_file": "kb", "page_or_line": "q", "text_summary": "份额提升"},
                {"fact_id": "ev4", "source_file": "alpha", "page_or_line": "q", "text_summary": "毛利率改善"},
            ]

            materialize_forecast_basis_sheet(workbook_path=workbook, forecast_basis=basis, evidence_store=evidence)
            out = openpyxl.load_workbook(workbook, data_only=False)
            ws = out["Forecast Basis"]
            rows = [[ws[f"{col}{r}"].value for col in "ABCDEFGHIJKL"] for r in range(1, ws.max_row + 1)]
            self.assertTrue(any(row[1] == "手机CIS.revenue" and row[4] == "market x share x ASP" for row in rows))
            self.assertTrue(any(row[1] == "手机CIS.volume" for row in rows))
            self.assertTrue(any(row[1] == "手机CIS.margin" for row in rows))

    def test_build_forecast_basis_payload_normalizes_architecture_to_validator_shape(self) -> None:
        architecture = {
            "company": "韦尔股份",
            "reported_year": "2025A",
            "target_window": ["2026E", "2027E", "2028E"],
            "segments": [
                {
                    "segment": "手机CIS",
                    "years_covered": ["2026E", "2027E", "2028E"],
                    "revenue_driver": "market x share x ASP",
                    "forecast_values": {"2026E": 100.0, "2027E": 110.0, "2028E": 121.0},
                    "volume_logic": {"mechanism": "销量增长", "evidence_refs": ["ev1"]},
                    "asp_logic": {"mechanism": "ASP提升", "evidence_refs": ["ev2"]},
                    "share_logic": {"mechanism": "份额提升", "evidence_refs": ["ev3"]},
                    "margin_logic": {"mechanism": "毛利率改善", "evidence_refs": ["ev4"], "values": {"2026E": 0.3, "2027E": 0.31, "2028E": 0.32}},
                    "kill_conditions": ["需求不及预期"],
                    "weak_assumptions": ["高端导入顺利"],
                    "source_ref": ["ev1", "ev2"],
                    "source_tier": "reference_files",
                    "confidence": "medium",
                    "review_flag": "none",
                }
            ],
            "consolidated_logic": {"selling_expense_logic": "费用率稳定"},
        }
        basis = build_forecast_basis_payload(
            company="韦尔股份",
            report_year=2025,
            annual_report_facts={"reported_facts": {"营业收入": 240.0}},
            meeting_notes_facts={"bridge_facts": {}},
            evidence_payload={"providers": []},
            forecast_architecture=architecture,
        )
        self.assertEqual(basis["language"], "zh-CN")
        self.assertTrue(basis["completeness_audit"]["passed"])
        self.assertTrue(basis["segment_assumption_cards"])
        self.assertTrue(basis["facts"][0]["evidence_items"])
        self.assertTrue(basis["segment_assumption_cards"][0]["evidence_items"])

    def test_validate_forecast_basis_requires_structured_evidence_items(self) -> None:
        invalid_basis = {
            "company": "沃尔核材",
            "cutoff_date": "2026-04-19",
            "reported_year": "2025A",
            "target_window": ["2026E", "2027E", "2028E"],
            "language": "zh-CN",
            "completeness_audit": {"passed": True},
            "facts": [{"key": "reported_revenue_2025A", "value": 84.5}],
            "assumptions": [],
            "segment_assumption_cards": [
                {
                    "segment": "高速铜缆",
                    "year": "2026E",
                    "metric": "revenue",
                    "value": 13.3,
                    "driver_form": "shipments x ASP",
                    "volume_logic": {"mechanism": "放量", "evidence_refs": ["ev1"]},
                    "asp_logic": {"mechanism": "结构升级", "evidence_refs": ["ev1"]},
                    "share_logic": {"mechanism": "客户导入", "evidence_refs": ["ev1"]},
                    "margin_logic": {"mechanism": "高毛利产品占比提升", "evidence_refs": ["ev1"]},
                    "kill_conditions": ["下游需求不及预期"],
                }
            ],
        }
        with self.assertRaises(Exception):
            validate_forecast_basis(invalid_basis)

    def test_render_forecast_basis_markdown_uses_structured_evidence_rows(self) -> None:
        basis = {
            "company": "沃尔核材",
            "facts": [
                {
                    "metric": "营业收入",
                    "year": "2025A",
                    "value": 84.5,
                    "source_ref": ["ev_annual"],
                    "evidence_items": [
                        {
                            "claim": "2025年营业收入84.51亿元。",
                            "source_ref": "ev_annual",
                            "source_tier": "reference_files",
                            "source_label": "年报 | 营业收入",
                        }
                    ],
                }
            ],
            "segment_assumption_cards": [
                {
                    "segment": "高速铜缆",
                    "year": "2026E",
                    "metric": "revenue",
                    "value": 13.3,
                    "driver_form": "shipments x ASP",
                    "volume_logic": {
                        "mechanism": "224G和448G产品继续放量。",
                        "evidence_refs": ["ev_note"],
                        "evidence_items": [
                            {
                                "claim": "224G和448G产品继续放量。",
                                "source_ref": "ev_note",
                                "source_tier": "reference_files",
                                "source_label": "纪要 | 高速线",
                            }
                        ],
                    },
                    "asp_logic": {"mechanism": "高规格产品占比提升。", "evidence_refs": ["ev_note"], "evidence_items": []},
                    "share_logic": {"mechanism": "核心客户订单延续。", "evidence_refs": ["ev_note"], "evidence_items": []},
                    "margin_logic": {"mechanism": "高毛利产品占比提升。", "evidence_refs": ["ev_note"], "evidence_items": []},
                    "kill_conditions": ["客户扩产落地慢于预期"],
                    "source_ref": ["ev_note"],
                    "evidence_items": [
                        {
                            "claim": "高速铜缆收入增长接近50%，224G和448G继续放量。",
                            "source_ref": "ev_note",
                            "source_tier": "reference_files",
                            "source_label": "纪要 | 高速线",
                        }
                    ],
                }
            ],
        }
        rendered = render_forecast_basis_markdown(basis)
        self.assertIn("依据1：年报 | 营业收入", rendered)
        self.assertIn("依据1：纪要 | 高速线", rendered)
        self.assertNotIn("原始长文本堆叠", rendered)


    def test_build_forecast_basis_payload_prefers_annual_report_for_facts_and_filters_stale_research_forecast(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook = tmp / "model.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "segment"
            ws["B1"] = "metric"
            ws["C1"] = 2025
            ws["D1"] = "2026E"
            ws["E1"] = "2027E"
            ws["F1"] = "2028E"
            ws["A2"] = "高速铜缆"
            ws["B2"] = "收入"
            ws["C2"] = 10.0
            wb.save(workbook)

            blueprint = WorkbookBlueprintParser().parse(workbook)
            evidence_payload = {
                "providers": [
                    {
                        "source_ref": "annual.pdf",
                        "source_type": "user_supplied_text",
                        "source_tier": "reference_files",
                        "content": "2025年营业收入84.5亿元。高速铜缆收入实现较快增长。",
                        "metadata": {"reference_kind": "annual_report"},
                    },
                    {
                        "source_ref": "meeting.md",
                        "source_type": "user_supplied_text",
                        "source_tier": "reference_files",
                        "content": "高速铜缆224G和448G持续放量，客户验证推进。",
                        "metadata": {"reference_kind": "meeting_notes"},
                    },
                    {
                        "source_ref": "research.md",
                        "source_type": "user_supplied_text",
                        "source_tier": "reference_files",
                        "content": "预计2026E收入13亿元，2027E收入16亿元。高速铜缆产品升级带动ASP提升。",
                        "metadata": {"reference_kind": "research_report"},
                    },
                ]
            }
            architecture = build_forecast_architecture(
                company="沃尔核材",
                report_year=2025,
                annual_report_facts={"reported_facts": {"营业收入": 84.5}},
                meeting_notes_facts={"bridge_facts": {}},
                evidence_payload=evidence_payload,
                workbook_path=workbook,
                blueprint=blueprint,
            )
            basis = build_forecast_basis_payload(
                company="沃尔核材",
                report_year=2025,
                annual_report_facts={"reported_facts": {"营业收入": 84.5}},
                meeting_notes_facts={"bridge_facts": {}},
                evidence_payload=evidence_payload,
                forecast_architecture=architecture,
            )

            self.assertEqual(len(basis["facts"][0]["evidence_items"]), 1)
            self.assertEqual(basis["facts"][0]["evidence_items"][0]["source_ref"], "annual.pdf")
            rendered = render_forecast_basis_markdown(basis)
            self.assertNotIn("预计2026E收入13亿元", rendered)
            first_card = basis["segment_assumption_cards"][0]
            claims = [item["claim"] for item in first_card["evidence_items"]]
            self.assertEqual(len(claims), len(set(claims)))
            self.assertTrue(any("ASP" in item["claim"] or "提升" in item["claim"] for item in first_card["evidence_items"]))


    def test_review_forecast_architecture_requires_alpha_recall_trace_check(self) -> None:
        architecture = {
            "company": "test_co",
            "segments": [
                {
                    "segment": "segment_a",
                    "years_covered": ["2026E", "2027E", "2028E"],
                    "revenue_driver": "units x ASP",
                    "forecast_values": {"2026E": 100.0, "2027E": 110.0, "2028E": 121.0},
                    "volume_logic": {"mechanism": "volume growth", "evidence_refs": []},
                    "asp_logic": {"mechanism": "asp uplift", "evidence_refs": []},
                    "share_logic": {"mechanism": "share gain", "evidence_refs": []},
                    "margin_logic": {
                        "mechanism": "mix improvement",
                        "evidence_refs": [],
                        "values": {"2026E": 0.30, "2027E": 0.31, "2028E": 0.32},
                    },
                    "kill_conditions": ["demand miss"],
                    "source_tier": "reference_files",
                    "source_ref": ["annual-1"],
                }
            ],
        }
        review = review_forecast_architecture(
            forecast_architecture=architecture,
            annual_report_facts={"reported_facts": {"revenue": 240.0}},
            evidence_payload={
                "providers": [
                    {"source_tier": "reference_files", "content": "annual", "source_ref": "annual-1"},
                    {"source_tier": "local_kb", "content": "kb", "source_ref": "kb-1"},
                ]
            },
        )
        self.assertEqual(review["pass_or_fail"], "fail")
        self.assertTrue(any(item["name"] == "alpha_pai_recalled" and not item["passed"] for item in review["checks"]))

        self.assertTrue(any(item["name"] == "alpha_pai_recalled" and not item["passed"] for item in review["checks"]))


if __name__ == "__main__":
    unittest.main()
