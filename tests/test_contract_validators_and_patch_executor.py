import hashlib
import json
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.forecast_tools.build_cell_instructions import (  # type: ignore
    ContractValidationError,
    build_cell_instructions,
)
from scripts.forecast_tools.contract_validators import (  # type: ignore
    validate_cell_instructions_payload,
    validate_patch_log_payload,
)
from scripts.forecast_tools.patch_executor import execute_patch_from_instructions  # type: ignore


def sample_workbook_map() -> dict:
    return {
        "workbook": "model.xlsx",
        "main_modeling_sheet": "Model",
        "main_modeling_sheet_index": 0,
        "header_row": 2,
        "label_column": "B",
        "historical_columns": {"2025A": "K"},
        "forecast_columns": {"2026E": "L", "2027E": "M"},
        "rollforward_pattern": {
            "reported_col": "K",
            "forecast_start_col": "L",
            "new_far_year_col": "N",
        },
        "current_headers": {"K": "2025A", "L": "2026E", "M": "2027E"},
        "current_forecast_window": ["2026E", "2027E"],
        "summary_extension_status": {"adjacent_forecast_populated_new_far_year_blank": 0},
        "row_registry": [
            {
                "row_id": "phone_asp",
                "sheet": "Model",
                "row": 33,
                "label": "手机 ASP",
                "role": "driver_input",
                "writable": True,
                "required_years": ["2026E", "2027E"],
                "must_extend_to_far_year": True,
                "year_cells": {"2026E": "L33", "2027E": "M33"},
                "basis_paths": {
                    "2026E": "assumptions.phone_asp_2026E.value",
                    "2027E": "assumptions.phone_asp_2027E.value",
                },
                "validation": {},
            },
            {
                "row_id": "phone_revenue",
                "sheet": "Model",
                "row": 29,
                "label": "手机收入",
                "role": "formula_derived",
                "writable": False,
                "required_years": ["2026E", "2027E"],
                "must_extend_to_far_year": True,
                "year_cells": {"2026E": "L29", "2027E": "M29"},
                "formula_template": "={col}34*{col}33/100",
                "source_row_id": "phone_volume",
                "validation": {},
            },
            {
                "row_id": "summary_phone_revenue",
                "sheet": "Model",
                "row": 118,
                "label": "摘要手机收入",
                "role": "summary_display",
                "writable": False,
                "required_years": ["2026E", "2027E"],
                "must_extend_to_far_year": True,
                "year_cells": {"2026E": "L118", "2027E": "M118"},
                "formula_template": "={col}29",
                "source_row_id": "phone_revenue",
                "validation": {},
            },
        ],
        "writable_driver_targets": ["phone_asp"],
        "formula_rows": ["phone_revenue"],
        "display_rows": ["summary_phone_revenue"],
        "map_validation_hints": {"must_match_headers": True},
    }


def sample_forecast_basis() -> dict:
    return {
        "company": "603501.SH",
        "cutoff_date": "2025-05-31",
        "reported_year": "2025A",
        "target_window": ["2026E", "2027E"],
        "language": "zh-CN",
        "completeness_audit": {"passed": True, "missing_segments": [], "missing_years": [], "missing_margin_logic": []},
        "facts": [],
        "assumptions": [
            {"key": "phone_asp_2026E", "value": 112.5},
            {"key": "phone_asp_2027E", "value": 118.0},
        ],
        "segment_assumption_cards": [
            {
                "segment": "手机CIS",
                "year": "2026E",
                "metric": "revenue",
                "value": 92.5,
                "driver_form": "volume x ASP",
                "volume_logic": {"mechanism": "高规格导入", "evidence_items": []},
                "asp_logic": {"mechanism": "像素升级", "evidence_items": []},
                "share_logic": {"mechanism": "高端份额稳定", "evidence_items": []},
                "margin_logic": {"mechanism": "高端 mix 改善", "evidence_items": []},
                "kill_conditions": ["ASP 下行超预期"],
                "evidence_items": [
                    {
                        "claim": "高规格导入带动收入增长。",
                        "source_ref": "meeting-note-1",
                        "source_tier": "reference_files",
                        "source_label": "meeting-notes.docx | meeting_note",
                    }
                ],
            }
        ],
    }


def build_sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["K2"] = "2025A"
    ws["L2"] = "2026E"
    ws["M2"] = "2027E"
    ws["L29"] = "=L34*L33/100"
    ws["M29"] = "=M34*M33/100"
    ws["L33"] = 0
    ws["M33"] = 0
    ws["L118"] = "=L29"
    ws["M118"] = "=M29"
    wb.save(path)


class ContractValidatorsAndPatchExecutorTests(unittest.TestCase):
    def test_validate_cell_instructions_payload_accepts_compiled_output(self) -> None:
        payload = build_cell_instructions(sample_workbook_map(), sample_forecast_basis())
        validate_cell_instructions_payload(payload)

    def test_validate_patch_log_payload_rejects_missing_output_hash(self) -> None:
        with self.assertRaises(ContractValidationError):
            validate_patch_log_payload(
                [
                    {
                        "sheet": "Model",
                        "cell": "L33",
                        "row_id": "phone_asp",
                        "year": "2026E",
                        "before": 0,
                        "after": 112.5,
                        "write_type": "value",
                        "instruction_id": "phone_asp.2026E",
                        "basis_ref": "assumptions.phone_asp_2026E.value",
                        "review_flag": None,
                        "formula_preserved": None,
                        "parity_audit_status": "pending",
                    }
                ]
            )

    def test_execute_patch_from_instructions_writes_candidate_and_patch_log(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "model.xlsx"
            build_sample_workbook(workbook_path)

            workbook_map = sample_workbook_map()
            workbook_map["workbook"] = str(workbook_path)
            forecast_basis = sample_forecast_basis()
            instructions = build_cell_instructions(workbook_map, forecast_basis)

            output_workbook = tmp / "candidate.xlsx"
            output_patch_log = tmp / "patch_log.json"

            execute_patch_from_instructions(
                workbook_path=workbook_path,
                workbook_map=workbook_map,
                cell_instructions=instructions,
                output_workbook=output_workbook,
                patch_log_path=output_patch_log,
            )

            wb = openpyxl.load_workbook(output_workbook, data_only=False)
            ws = wb["Model"]
            self.assertEqual(ws["L33"].value, 112.5)
            self.assertEqual(ws["M33"].value, 118.0)
            self.assertEqual(ws["L29"].value, "=L34*L33/100")
            self.assertEqual(ws["L118"].value, "=L29")

            patch_log = json.loads(output_patch_log.read_text(encoding="utf-8"))
            validate_patch_log_payload(patch_log)
            self.assertEqual(len(patch_log), 6)
            expected_hash = hashlib.sha256(output_workbook.read_bytes()).hexdigest()
            self.assertTrue(all(entry["output_hash"] == expected_hash for entry in patch_log))

    def test_execute_patch_from_instructions_rewrites_display_rows_when_contract_requests_it(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "model.xlsx"
            build_sample_workbook(workbook_path)

            workbook_map = sample_workbook_map()
            workbook_map["workbook"] = str(workbook_path)
            workbook_map["row_registry"][2]["display_write_mode"] = "rewrite"

            ws = openpyxl.load_workbook(workbook_path)["Model"]
            ws["L118"] = "=L57"
            ws["M118"] = "=M57"
            ws.parent.save(workbook_path)

            instructions = build_cell_instructions(workbook_map, sample_forecast_basis())
            output_workbook = tmp / "candidate.xlsx"
            output_patch_log = tmp / "patch_log.json"

            execute_patch_from_instructions(
                workbook_path=workbook_path,
                workbook_map=workbook_map,
                cell_instructions=instructions,
                output_workbook=output_workbook,
                patch_log_path=output_patch_log,
            )

            wb = openpyxl.load_workbook(output_workbook, data_only=False)
            out_ws = wb["Model"]
            self.assertEqual(out_ws["L118"].value, "=L29")
            self.assertEqual(out_ws["M118"].value, "=M29")

            patch_log = json.loads(output_patch_log.read_text(encoding="utf-8"))
            summary_entries = [entry for entry in patch_log if entry["row_id"] == "summary_phone_revenue"]
            self.assertEqual([entry["write_type"] for entry in summary_entries], ["formula", "formula"])
            self.assertEqual(summary_entries[0]["before"], "=L57")
            self.assertEqual(summary_entries[0]["after"], "=L29")

    def test_execute_patch_from_instructions_fails_when_main_sheet_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "model.xlsx"
            build_sample_workbook(workbook_path)

            workbook_map = sample_workbook_map()
            workbook_map["workbook"] = str(workbook_path)
            workbook_map["main_modeling_sheet"] = "NotModel"
            instructions = build_cell_instructions(sample_workbook_map(), sample_forecast_basis())

            with self.assertRaises(ContractValidationError):
                execute_patch_from_instructions(
                    workbook_path=workbook_path,
                    workbook_map=workbook_map,
                    cell_instructions=instructions,
                    output_workbook=tmp / "candidate.xlsx",
                    patch_log_path=tmp / "patch_log.json",
                )


if __name__ == "__main__":
    unittest.main()
