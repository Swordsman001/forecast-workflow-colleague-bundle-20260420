import json
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.forecast_tools.contract_workflow import run_contract_workflow  # type: ignore


def sample_workbook_map() -> dict:
    return {
        "workbook": "model.xlsx",
        "main_modeling_sheet": "Model",
        "main_modeling_sheet_index": 0,
        "header_row": 2,
        "label_column": "B",
        "historical_columns": {"2025A": "K"},
        "forecast_columns": {"2026E": "L", "2027E": "M"},
        "rollforward_pattern": {
            "reported_col": "K",
            "forecast_start_col": "L",
            "new_far_year_col": "N",
        },
        "current_headers": {"K": "2025A", "L": "2026E", "M": "2027E"},
        "current_forecast_window": ["2026E", "2027E"],
        "summary_extension_status": {"adjacent_forecast_populated_new_far_year_blank": 0},
        "row_registry": [
            {
                "row_id": "phone_asp",
                "sheet": "Model",
                "row": 33,
                "label": "手机 ASP",
                "role": "driver_input",
                "writable": True,
                "required_years": ["2026E", "2027E"],
                "must_extend_to_far_year": True,
                "year_cells": {"2026E": "L33", "2027E": "M33"},
                "basis_paths": {
                    "2026E": "assumptions.phone_asp_2026E.value",
                    "2027E": "assumptions.phone_asp_2027E.value",
                },
                "validation": {},
            },
            {
                "row_id": "phone_revenue",
                "sheet": "Model",
                "row": 29,
                "label": "手机收入",
                "role": "formula_derived",
                "writable": False,
                "required_years": ["2026E", "2027E"],
                "must_extend_to_far_year": True,
                "year_cells": {"2026E": "L29", "2027E": "M29"},
                "formula_template": "={col}34*{col}33/100",
                "source_row_id": "phone_volume",
                "validation": {},
            },
        ],
        "writable_driver_targets": ["phone_asp"],
        "formula_rows": ["phone_revenue"],
        "display_rows": [],
        "map_validation_hints": {"must_match_headers": True},
    }


def sample_forecast_basis() -> dict:
    return {
        "company": "603501.SH",
        "cutoff_date": "2025-05-31",
        "reported_year": "2025A",
        "target_window": ["2026E", "2027E"],
        "language": "zh-CN",
        "completeness_audit": {"passed": True, "missing_segments": [], "missing_years": [], "missing_margin_logic": []},
        "facts": [
            {
                "metric": "revenue",
                "year": "2025A",
                "value": 257.31,
                "source_ref": "annual report p.8",
                "confidence": "high",
                "evidence_items": [
                    {
                        "claim": "2025A revenue = 257.31",
                        "source_ref": "annual report p.8",
                        "source_tier": "reference_files",
                        "source_label": "annual-report.pdf | annual_report",
                    }
                ],
            }
        ],
        "assumptions": [
            {"key": "phone_asp_2026E", "value": 112.5},
            {"key": "phone_asp_2027E", "value": 118.0},
        ],
        "segment_assumption_cards": [
            {
                "segment": "手机CIS",
                "year": "2026E",
                "metric": "revenue",
                "value": 92.5,
                "driver_form": "volume x ASP",
                "volume_logic": {"mechanism": "高像素机型导入", "evidence_items": []},
                "asp_logic": {"mechanism": "像素升级", "evidence_items": []},
                "share_logic": {"mechanism": "高端份额稳定", "evidence_items": []},
                "margin_logic": {"mechanism": "高端 mix 改善", "evidence_items": []},
                "kill_conditions": ["ASP 下行超预期"],
                "source_ref": ["meeting notes 2025-04-29", "research summary"],
                "confidence": "medium",
                "evidence_items": [
                    {
                        "claim": "高像素机型导入支撑收入增长。",
                        "source_ref": "meeting notes 2025-04-29",
                        "source_tier": "reference_files",
                        "source_label": "meeting-notes.docx | meeting_note",
                    }
                ],
            }
        ],
    }


def build_sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["K2"] = "2025A"
    ws["L2"] = "2026E"
    ws["M2"] = "2027E"
    ws["L29"] = "=L34*L33/100"
    ws["M29"] = "=M34*M33/100"
    ws["L33"] = 0
    ws["M33"] = 0
    wb.save(path)


def sample_evidence_store_lines() -> list[dict]:
    return [
        {
            "fact_id": "annual_2025_revenue",
            "company": "603501.SH",
            "period": "2025A",
            "metric": "revenue",
            "value": 288.55,
            "unit": "RMB_100m",
            "source_type": "annual_report",
            "source_file": "annual-report.pdf",
            "confidence": "high",
            "text_summary": "2025年营业收入288.55亿元",
        },
        {
            "fact_id": "meeting_2025_phone_asp",
            "company": "603501.SH",
            "period": "2026E",
            "metric": "phone_asp",
            "value": 112.5,
            "unit": "RMB",
            "source_type": "meeting_note",
            "source_file": "meeting-notes.docx",
            "tag": "management_guidance",
            "confidence": "medium",
            "why_it_matters": "用于手机CIS ASP预测",
        },
    ]


class ContractWorkflowRunnerTests(unittest.TestCase):
    def test_run_contract_workflow_produces_all_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "model.xlsx"
            build_sample_workbook(workbook_path)

            workbook_map_path = tmp / "workbook_map.json"
            forecast_basis_path = tmp / "forecast_basis.json"
            evidence_store_path = tmp / "evidence_store.jsonl"
            output_dir = tmp / "output"
            workbook_map = sample_workbook_map()
            workbook_map["workbook"] = str(workbook_path)
            workbook_map_path.write_text(json.dumps(workbook_map, ensure_ascii=False, indent=2), encoding="utf-8")
            forecast_basis_path.write_text(json.dumps(sample_forecast_basis(), ensure_ascii=False, indent=2), encoding="utf-8")
            evidence_store_path.write_text(
                "\n".join(json.dumps(line, ensure_ascii=False) for line in sample_evidence_store_lines()) + "\n",
                encoding="utf-8",
            )

            result = run_contract_workflow(
                workbook_map_path=workbook_map_path,
                forecast_basis_path=forecast_basis_path,
                evidence_store_path=evidence_store_path,
                workbook_path=workbook_path,
                output_dir=output_dir,
            )

            self.assertTrue(result["artifact_hashes_path"].exists())
            self.assertTrue(result["cell_instructions_path"].exists())
            self.assertTrue(result["candidate_workbook_path"].exists())
            self.assertTrue(result["patch_log_path"].exists())
            self.assertTrue(result["verification_report_path"].exists())

            artifact_hashes = json.loads(result["artifact_hashes_path"].read_text(encoding="utf-8"))
            self.assertIn("evidence_store_hash", artifact_hashes)
            self.assertIn("source_workbook_hash", artifact_hashes)

            cell_instructions = json.loads(result["cell_instructions_path"].read_text(encoding="utf-8"))
            self.assertEqual(cell_instructions["workbook_map_hash"], artifact_hashes["workbook_map_hash"])
            self.assertEqual(cell_instructions["forecast_basis_hash"], artifact_hashes["forecast_basis_hash"])
            self.assertEqual(cell_instructions["evidence_store_hash"], artifact_hashes["evidence_store_hash"])
            self.assertEqual(cell_instructions["source_workbook_hash"], artifact_hashes["source_workbook_hash"])

            patch_log = json.loads(result["patch_log_path"].read_text(encoding="utf-8"))
            self.assertTrue(all(entry["instruction_hash"] == artifact_hashes["cell_instructions_hash"] for entry in patch_log))
            self.assertTrue(all(entry["basis_hash"] == artifact_hashes["forecast_basis_hash"] for entry in patch_log))
            self.assertTrue(all(entry["map_hash"] == artifact_hashes["workbook_map_hash"] for entry in patch_log))
            self.assertTrue(all(entry["evidence_hash"] == artifact_hashes["evidence_store_hash"] for entry in patch_log))

            verification = json.loads(result["verification_report_path"].read_text(encoding="utf-8"))
            self.assertTrue(verification["passed"])
            self.assertEqual(verification["instruction_count"], 4)
            self.assertEqual(verification["patch_log_count"], 4)
            self.assertTrue(verification["hash_match"])
            self.assertTrue(verification["instruction_hash_match"])
            self.assertTrue(verification["basis_hash_match"])
            self.assertTrue(verification["map_hash_match"])
            self.assertTrue(verification["evidence_hash_match"])

            candidate_wb = openpyxl.load_workbook(result["candidate_workbook_path"], data_only=False)
            self.assertIn("Forecast Basis", candidate_wb.sheetnames)
            basis_ws = candidate_wb["Forecast Basis"]
            self.assertEqual(basis_ws["A1"].value, "预测依据")
            self.assertEqual(
                [basis_ws[f"{col}3"].value for col in "ABCDEFGHIJ"],
                ["分类", "项目", "年度", "数值", "驱动形式", "依据摘要", "来源", "证据标签", "风险提示", "置信度"],
            )
            rendered_rows = [
                [basis_ws[f"{col}{row}"].value for col in "ABCDEFGHIJ"]
                for row in range(4, basis_ws.max_row + 1)
            ]
            self.assertTrue(
                any(
                    row[0] == "已知事实"
                    and ("annual-report.pdf" in (row[6] or "") or "annual report p.8" in (row[6] or ""))
                    for row in rendered_rows
                )
            )
            self.assertTrue(
                any(
                    row[0] == "预测假设"
                    and ("meeting-notes.docx" in (row[6] or "") or "meeting notes 2025-04-29" in (row[6] or ""))
                    and row[7] == "meeting notes 2025-04-29; research summary"
                    for row in rendered_rows
                )
            )

    def test_run_contract_workflow_normalizes_sparse_evidence_store_lines(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "model.xlsx"
            build_sample_workbook(workbook_path)

            workbook_map_path = tmp / "workbook_map.json"
            forecast_basis_path = tmp / "forecast_basis.json"
            evidence_store_path = tmp / "evidence_store.jsonl"
            output_dir = tmp / "output"

            workbook_map = sample_workbook_map()
            workbook_map["workbook"] = str(workbook_path)
            workbook_map_path.write_text(json.dumps(workbook_map, ensure_ascii=False, indent=2), encoding="utf-8")
            forecast_basis = sample_forecast_basis()
            forecast_basis_path.write_text(json.dumps(forecast_basis, ensure_ascii=False, indent=2), encoding="utf-8")

            sparse_lines = [
                {
                    "fact_id": "annual_2025_revenue",
                    "period": "2025A",
                    "metric": "revenue",
                    "value": 288.55,
                    "source_type": "annual_report",
                }
            ]
            evidence_store_path.write_text(
                "\n".join(json.dumps(line, ensure_ascii=False) for line in sparse_lines) + "\n",
                encoding="utf-8",
            )

            result = run_contract_workflow(
                workbook_map_path=workbook_map_path,
                forecast_basis_path=forecast_basis_path,
                evidence_store_path=evidence_store_path,
                workbook_path=workbook_path,
                output_dir=output_dir,
            )

            self.assertTrue(result["candidate_workbook_path"].exists())
            verification = json.loads(result["verification_report_path"].read_text(encoding="utf-8"))
            self.assertTrue(verification["passed"])
            candidate_wb = openpyxl.load_workbook(result["candidate_workbook_path"], data_only=False)
            self.assertIn("Forecast Basis", candidate_wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
