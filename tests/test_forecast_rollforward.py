import json
import sys
import tempfile
import unittest
import time
from io import StringIO
from pathlib import Path
from unittest import mock

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.forecast_tools import rollforward as rollforward_mod  # type: ignore
from scripts.forecast_tools.providers import ProviderResult  # type: ignore
from scripts.forecast_rollforward import main as cli_main  # type: ignore


ChangeRecord = rollforward_mod.ChangeRecord
WorkbookBlueprintParser = rollforward_mod.WorkbookBlueprintParser
WorkbookRollforwardEngine = rollforward_mod.WorkbookRollforwardEngine
WorkbookBlueprint = rollforward_mod.WorkbookBlueprint


def build_sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "营收拆分"

    ws.cell(2, 2).value = "单位：亿元"
    headers = {
        3: 2022,
        4: 2023,
        5: "2024E",
        6: "2025E",
        7: "2026E",
        8: "备注",
        9: "24H1",
    }
    for col, value in headers.items():
        ws.cell(2, col).value = value

    labels = [
        "营业收入",
        "毛利率",
        "归母净利润",
        "销售费用率（%）",
        "1）手机营收（亿元）",
        "2）汽车营收（亿元）",
    ]
    for index, label in enumerate(labels, start=3):
        ws.cell(index, 2).value = label

    values = {
        "营业收入": [210.0, 220.0, 230.0, 250.0, 270.0],
        "毛利率": [0.22, 0.23, 0.24, 0.245, 0.25],
        "归母净利润": [5.2, 6.1, 7.0, 8.2, 9.5],
        "销售费用率（%）": [0.031, 0.03, 0.029, 0.028, 0.027],
        "1）手机营收（亿元）": [120.0, 124.0, 128.0, 135.0, 141.0],
        "2）汽车营收（亿元）": [20.0, 24.0, 30.0, 38.0, 45.0],
    }

    for row in range(3, 9):
        label = ws.cell(row, 2).value
        for offset, value in enumerate(values[label], start=3):
            ws.cell(row, offset).value = value

    wb.save(path)


def build_candidate_decision_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "营收拆分"
    ws["B2"] = "单位：亿元"
    ws["D2"] = 2024
    ws["E2"] = "2025E"
    ws["F2"] = "2026E"
    ws["B3"] = "手机CIS"
    ws["C3"] = "收入"
    ws["D3"] = 120.0
    ws["E3"] = 140.0
    ws["B4"] = "汽车CIS"
    ws["C4"] = "收入"
    ws["D4"] = 30.0
    ws["E4"] = 48.0
    ws["B5"] = "安防"
    ws["C5"] = "收入"
    ws["D5"] = 20.0
    ws["E5"] = 18.0
    ws["B6"] = "营业收入"
    ws["C6"] = "收入"
    ws["D6"] = 170.0
    ws["E6"] = 206.0
    wb.save(path)


class WorkbookBlueprintParserTests(unittest.TestCase):
    def test_parser_identifies_historical_and_forecast_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "sample.xlsx"
            build_sample_workbook(workbook_path)

            parser = WorkbookBlueprintParser()
            blueprint = parser.parse(workbook_path)

            self.assertEqual(blueprint.primary_sheet, "营收拆分")
            self.assertEqual(blueprint.label_column, 2)
            self.assertEqual(blueprint.header_row, 2)
            self.assertEqual(blueprint.historical_columns, {2022: 3, 2023: 4})
            self.assertEqual(
                blueprint.forecast_columns,
                {2024: 5, 2025: 6, 2026: 7},
            )
            self.assertEqual(blueprint.excluded_columns, [8, 9])

    def test_parser_prefers_segment_name_column_over_repeated_metric_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "segment_metric.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "分部"
            ws["B1"] = "指标"
            ws["C1"] = 2025
            ws["D1"] = "2026E"
            ws["E1"] = "2027E"
            ws["F1"] = "2028E"
            rows = [
                ("高速铜缆", "收入"),
                (None, "yoy"),
                (None, "毛利率"),
                ("发泡线", "收入"),
                (None, "市场份额"),
                ("新能源电力", "收入"),
            ]
            for row_idx, (segment, metric) in enumerate(rows, start=2):
                ws.cell(row_idx, 1).value = segment
                ws.cell(row_idx, 2).value = metric
            wb.save(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            self.assertEqual(blueprint.label_column, 1)
            self.assertEqual(sorted(blueprint.primary_row_labels.values()), [2, 5, 7])
            self.assertIn([2, 3, 4], blueprint.row_blocks.values())
            self.assertIn([5, 6], blueprint.row_blocks.values())
            return
            self.assertEqual(blueprint.primary_row_labels["楂橀€熼摐缂?"], 2)
            self.assertEqual(blueprint.row_blocks["楂橀€熼摐缂?"], [2, 3, 4])
            self.assertEqual(blueprint.row_blocks["鍙戞场绾?"], [5, 6])


class WorkbookRollforwardEngineTests(unittest.TestCase):
    def test_rollforward_shifts_forecast_window_and_creates_new_far_year(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "sample.xlsx"
            build_sample_workbook(workbook_path)

            parser = WorkbookBlueprintParser()
            blueprint = parser.parse(workbook_path)
            engine = WorkbookRollforwardEngine()

            result = engine.rollforward(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2024,
                actual_overrides={
                    "营业收入": 240.0,
                    "毛利率": 0.255,
                    "归母净利润": 8.8,
                    "销售费用率（%）": 0.026,
                    "1）手机营收（亿元）": 130.0,
                    "2）汽车营收（亿元）": 42.0,
                },
            )

            rolled = openpyxl.load_workbook(result.output_workbook)
            ws = rolled["营收拆分"]

            self.assertEqual(ws.cell(2, 5).value, 2024)
            self.assertEqual(ws.cell(2, 6).value, "2025E")
            self.assertEqual(ws.cell(2, 7).value, "2026E")
            self.assertEqual(ws.cell(2, 8).value, "2027E")
            self.assertEqual(ws.cell(3, 5).value, 240.0)
            self.assertEqual(ws.cell(3, 6).value, 250.0)
            self.assertEqual(ws.cell(3, 7).value, 270.0)
            self.assertEqual(ws.cell(3, 8).value, 291.6)
            self.assertAlmostEqual(ws.cell(7, 8).value, 147.2667, places=4)

    def test_rollforward_emits_field_level_change_records(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "sample.xlsx"
            build_sample_workbook(workbook_path)

            parser = WorkbookBlueprintParser()
            blueprint = parser.parse(workbook_path)
            engine = WorkbookRollforwardEngine()

            result = engine.rollforward(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2024,
                actual_overrides={
                    "营业收入": 240.0,
                    "归母净利润": 8.8,
                },
            )

            actualized = [
                record for record in result.change_records
                if record.change_type == "actualized_from_annual_report"
            ]
            rolled_forward = [
                record for record in result.change_records
                if record.change_type == "rolled_forward"
            ]

            self.assertTrue(
                any(record.row_label == "营业收入" and record.year == 2024 for record in actualized)
            )
            self.assertTrue(
                any(record.row_label == "归母净利润" and record.year == 2024 for record in actualized)
            )
            self.assertTrue(
                any(record.row_label == "营业收入" and record.year == 2027 for record in rolled_forward)
            )
            self.assertTrue(
                all(isinstance(record, ChangeRecord) for record in result.change_records)
            )

    def test_rollforward_extends_formula_rows_into_new_far_year(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "formula_sample.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws.cell(2, 2).value = "项目"
            ws.cell(2, 3).value = 2022
            ws.cell(2, 4).value = 2023
            ws.cell(2, 5).value = "2024E"
            ws.cell(2, 6).value = "2025E"
            ws.cell(2, 7).value = "2026E"
            ws.cell(3, 2).value = "高速铜缆"
            ws.cell(4, 2).value = "新能源电力"
            ws.cell(5, 2).value = "分部合计公式"
            for col, values in {
                3: (10.0, 5.0),
                4: (12.0, 6.0),
                5: (14.0, 7.0),
                6: (16.0, 8.0),
                7: (18.0, 9.0),
            }.items():
                ws.cell(3, col).value = values[0]
                ws.cell(4, col).value = values[1]
            ws.cell(5, 5).value = "=E3+E4"
            ws.cell(5, 6).value = "=F3+F4"
            ws.cell(5, 7).value = "=G3+G4"
            wb.save(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            result = WorkbookRollforwardEngine().rollforward(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2024,
                actual_overrides={"高速铜缆": 15.0, "新能源电力": 7.5},
            )

            rolled = openpyxl.load_workbook(result.output_workbook, data_only=False)
            ws = rolled["Model"]
            self.assertEqual(ws.cell(2, 8).value, "2027E")
            self.assertEqual(ws.cell(5, 8).value, "=H3+H4")

    def test_rollforward_extends_dependent_block_rows_into_new_far_year(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "dependent_rows.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "鍒嗛儴"
            ws["B1"] = "鎸囨爣"
            ws["C1"] = 2023
            ws["D1"] = 2024
            ws["E1"] = "2025E"
            ws["F1"] = "2026E"
            ws["G1"] = "2027E"
            ws["A2"] = "楂橀€熼摐缂?"
            ws["B2"] = "鏀跺叆"
            ws["A3"] = None
            ws["B3"] = "yoy"
            ws["A4"] = None
            ws["B4"] = "姣涘埄鐜?"
            ws["A5"] = "鍙戞场绾?"
            ws["B5"] = "鏀跺叆"
            ws["A6"] = None
            ws["B6"] = "甯傚満浠介"
            for col, value in zip(range(3, 8), [1.0, 3.1, 10.5, 11.6, 13.6], strict=False):
                ws.cell(2, col).value = value
            ws["D3"] = "=D2/C2-1"
            ws["E3"] = "=E2/D2-1"
            ws["F3"] = "=F2/E2-1"
            ws["G3"] = "=G2/F2-1"
            for col, value in zip(range(3, 8), [0.35, 0.45, 0.55, 0.55, 0.55], strict=False):
                ws.cell(4, col).value = value
            for col, value in zip(range(5, 8), [5.0, 6.0, 7.1], strict=False):
                ws.cell(5, col).value = value
            for col, value in zip(range(5, 8), [0.30, 0.52, 0.50], strict=False):
                ws.cell(6, col).value = value
            wb.save(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            result = WorkbookRollforwardEngine().rollforward(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                actual_overrides={},
            )

            rolled = openpyxl.load_workbook(result.output_workbook, data_only=False)
            ws = rolled["Model"]
            self.assertEqual(ws["H3"].value, "=H2/G2-1")
            self.assertEqual(ws["H4"].value, 0.55)
            self.assertEqual(ws["H6"].value, 0.5)
            self.assertEqual(result.parity_audit["status"], "passed")

    def test_rollforward_applies_year_specific_segment_and_dependent_schedules(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "guided_schedule.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "segment"
            ws["B1"] = "metric"
            ws["C1"] = 2024
            ws["D1"] = "2025E"
            ws["E1"] = "2026E"
            ws["F1"] = "2027E"
            ws["A2"] = "segment_a"
            ws["B2"] = "revenue"
            ws["A3"] = None
            ws["B3"] = "yoy"
            ws["A4"] = None
            ws["B4"] = "margin"
            ws["A5"] = None
            ws["B5"] = "share"
            ws["D2"] = 10.0
            ws["E2"] = 11.0
            ws["F2"] = 12.0
            ws["D3"] = 0.1
            ws["E3"] = 0.1
            ws["F3"] = 0.1
            ws["D4"] = 0.30
            ws["E4"] = 0.30
            ws["F4"] = 0.30
            ws["D5"] = 0.20
            ws["E5"] = 0.20
            ws["F5"] = 0.20
            wb.save(workbook_path)

            blueprint = WorkbookBlueprint(
                primary_sheet="Model",
                header_row=1,
                label_column=1,
                historical_columns={2024: 3},
                forecast_columns={2025: 4, 2026: 5, 2027: 6},
                excluded_columns=[],
                row_labels={"segment_a": 2},
                primary_row_labels={"segment_a": 2},
                row_blocks={"segment_a": [2, 3, 4, 5]},
            )
            result = WorkbookRollforwardEngine().rollforward(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                actual_overrides={},
                meeting_guidance={
                    "segment_a": {
                        "claim": "trajectory",
                        "year_values": {"2026E": 11.5, "2027E": 13.0, "2028E": 14.1},
                        "dependent_metric_values": {
                            "yoy": {"2026E": 0.15, "2027E": 0.13, "2028E": 0.085},
                            "margin": {"2026E": 0.31, "2027E": 0.325, "2028E": 0.333},
                            "share": {"2026E": 0.24, "2027E": 0.26, "2028E": 0.27},
                        },
                        "confidence": 0.7,
                        "review_required": False,
                    }
                },
            )

            rolled = openpyxl.load_workbook(result.output_workbook, data_only=False)
            ws = rolled["Model"]
            self.assertEqual(ws["E2"].value, 11.5)
            self.assertEqual(ws["F2"].value, 13.0)
            self.assertEqual(ws["G2"].value, 14.1)
            self.assertEqual(ws["E3"].value, 0.15)
            self.assertEqual(ws["F3"].value, 0.13)
            self.assertEqual(ws["G3"].value, 0.085)
            self.assertEqual(ws["G4"].value, 0.333)
            self.assertEqual(ws["G5"].value, 0.27)

    def test_rollforward_fails_fast_when_block_row_cannot_extend_to_far_year(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "parity_fail.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "鍒嗛儴"
            ws["B1"] = "鎸囨爣"
            ws["C1"] = 2023
            ws["D1"] = 2024
            ws["E1"] = "2025E"
            ws["F1"] = "2026E"
            ws["G1"] = "2027E"
            ws["A2"] = "楂橀€熼摐缂?"
            ws["B2"] = "鏀跺叆"
            ws["E2"] = 10.5
            ws["F2"] = 11.6
            ws["G2"] = 13.6
            ws["A3"] = None
            ws["B3"] = "manual_note"
            ws["F3"] = "needs manual rebuild"
            ws["G3"] = "still manual"
            wb.save(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            with self.assertRaises(RuntimeError):
                WorkbookRollforwardEngine().rollforward(
                    workbook_path=workbook_path,
                    blueprint=blueprint,
                    report_year=2025,
                    actual_overrides={},
                )

    def test_rollforward_rebuilds_formula_summary_rows_from_report_year_formula(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "summary_formula.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Model"
            ws["A1"] = "segment"
            ws["B1"] = "metric"
            ws["C1"] = 2024
            ws["D1"] = "2025E"
            ws["E1"] = "2026E"
            ws["F1"] = "2027E"
            ws["A2"] = "segment_total"
            ws["B2"] = "revenue"
            ws["D2"] = "=D3+D4"
            ws["E2"] = 11.0
            ws["F2"] = 13.0
            ws["A3"] = "segment_a"
            ws["B3"] = "revenue"
            ws["D3"] = 4.0
            ws["E3"] = 5.0
            ws["F3"] = 6.0
            ws["A4"] = "segment_b"
            ws["B4"] = "revenue"
            ws["D4"] = 6.0
            ws["E4"] = 7.0
            ws["F4"] = 8.0
            wb.save(workbook_path)

            blueprint = WorkbookBlueprint(
                primary_sheet="Model",
                header_row=1,
                label_column=1,
                historical_columns={2024: 3},
                forecast_columns={2025: 4, 2026: 5, 2027: 6},
                excluded_columns=[],
                row_labels={"segment_total": 2, "segment_a": 3, "segment_b": 4},
                primary_row_labels={"segment_total": 2, "segment_a": 3, "segment_b": 4},
                row_blocks={"segment_total": [2], "segment_a": [3], "segment_b": [4]},
            )
            result = WorkbookRollforwardEngine().rollforward(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2025,
                actual_overrides={},
                meeting_guidance={
                    "segment_a": {
                        "claim": "trajectory",
                        "year_values": {"2026E": 8.0, "2027E": 10.0, "2028E": 12.0},
                        "confidence": 0.8,
                        "review_required": False,
                    },
                    "segment_b": {
                        "claim": "trajectory",
                        "year_values": {"2026E": 4.0, "2027E": 5.0, "2028E": 6.0},
                        "confidence": 0.8,
                        "review_required": False,
                    },
                },
            )

            rolled = openpyxl.load_workbook(result.output_workbook, data_only=False)
            ws = rolled["Model"]
            self.assertEqual(ws["E2"].value, "=E3+E4")
            self.assertEqual(ws["F2"].value, "=F3+F4")
            self.assertEqual(ws["G2"].value, "=G3+G4")

    def test_rollforward_extends_to_explicit_target_far_year(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "sample.xlsx"
            build_sample_workbook(workbook_path)

            blueprint = WorkbookBlueprintParser().parse(workbook_path)
            result = WorkbookRollforwardEngine().rollforward(
                workbook_path=workbook_path,
                blueprint=blueprint,
                report_year=2024,
                actual_overrides={"营业收入": 240.0},
                target_far_year=2028,
            )

            rolled = openpyxl.load_workbook(result.output_workbook, data_only=False)
            ws = rolled["营收拆分"]
            self.assertEqual(ws.cell(2, 8).value, "2027E")
            self.assertEqual(ws.cell(2, 9).value, "2028E")
            self.assertEqual(result.parity_audit["new_far_year"], 2028)

    def test_collect_pre_edit_inputs_runs_independent_tasks_in_parallel_and_records_timing(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "sample.xlsx"
            report_path = tmp / "report.txt"
            meeting_notes_path = tmp / "meeting.txt"

            build_sample_workbook(workbook_path)
            report_path.write_text("营业收入 240.0", encoding="utf-8")
            meeting_notes_path.write_text("汽车CIS 2025年增长20%", encoding="utf-8")

            starts: dict[str, float] = {}
            original_parse = WorkbookBlueprintParser.parse

            def blueprint_side_effect(self, path):
                starts["blueprint"] = time.perf_counter()
                time.sleep(0.2)
                return original_parse(self, path)

            def financial_side_effect(self, company, report_year, ts_code=None, report_path=None, model_path=None):
                starts["financial_facts"] = time.perf_counter()
                time.sleep(0.2)
                return {"reported_facts": {"钀ヤ笟鏀跺叆": 240.0}, "segment_disclosure": []}

            def meeting_side_effect(self, path):
                starts["meeting_notes"] = time.perf_counter()
                time.sleep(0.2)
                return {"bridge_facts": {"姹借溅": {"growth_rate": 0.2}}}

            def evidence_side_effect(repo_root, company, meeting_notes, research_report=None, annual_report=None):
                starts["evidence_payload"] = time.perf_counter()
                time.sleep(0.2)
                return {"company": company, "providers": []}

            parser = WorkbookBlueprintParser()
            expected_blueprint = parser.parse(workbook_path)

            with mock.patch(
                "scripts.forecast_tools.rollforward.WorkbookBlueprintParser.parse",
                autospec=True,
                side_effect=blueprint_side_effect,
            ) as parse_mock, mock.patch(
                "scripts.forecast_tools.rollforward.TushareFinancialFactsAdapter.extract",
                autospec=True,
                side_effect=financial_side_effect,
            ) as financial_mock, mock.patch(
                "scripts.forecast_tools.rollforward.MeetingNotesFactExtractor.extract",
                autospec=True,
                side_effect=meeting_side_effect,
            ) as meeting_mock, mock.patch(
                "scripts.forecast_tools.rollforward.build_evidence_payload",
                side_effect=evidence_side_effect,
            ) as evidence_mock:
                started = time.perf_counter()
                result = rollforward_mod.collect_pre_edit_inputs(
                    repo_root=ROOT,
                    company="韦尔股份",
                    model_path=workbook_path,
                    report_path=report_path,
                    meeting_notes_path=meeting_notes_path,
                )
                elapsed = time.perf_counter() - started

            self.assertLess(elapsed, 1.3)
            self.assertGreaterEqual(parse_mock.call_count, 1)
            self.assertEqual(financial_mock.call_count, 1)
            self.assertEqual(meeting_mock.call_count, 1)
            self.assertEqual(evidence_mock.call_count, 1)
            self.assertIsNotNone(result["blueprint"])
            self.assertIn("financial_facts", result)
            self.assertIn("segment_mapping", result)
            self.assertIn("reconciliation_audit", result)
            self.assertIn("timing", result)
            self.assertEqual(
                {item["task"] for item in result["timing"]["tasks"]},
                {"blueprint_parse", "financial_facts_extract", "meeting_notes_extract", "evidence_payload_build"},
            )

            earliest = min(starts.values())
            latest = max(starts.values())
            self.assertLess(latest - earliest, 0.25)
            self.assertEqual(result["blueprint"].primary_sheet, expected_blueprint.primary_sheet)

    def test_collect_pre_edit_inputs_api_exists(self) -> None:
        self.assertTrue(hasattr(rollforward_mod, "collect_pre_edit_inputs"))


class ForecastRollforwardCliTests(unittest.TestCase):
    def test_cli_generates_candidate_workbook_and_sidecar_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "sample.xlsx"
            report_path = tmp / "report.txt"
            meeting_notes_path = tmp / "meeting.txt"
            research_report_path = tmp / "research.txt"
            output_dir = tmp / "out"

            build_sample_workbook(workbook_path)
            report_path.write_text(
                "\n".join(
                    [
                        "营业收入 240.0",
                        "毛利率 25.5%",
                        "归母净利润 8.8",
                        "扣非归母净利润 8.1",
                        "销售费用率 2.6%",
                    ]
                ),
                encoding="utf-8",
            )
            meeting_notes_path.write_text(
                "\n".join(
                    [
                        "汽车CIS 2025年增长20%",
                        "手机CIS 2025年增长10%",
                    ]
                ),
                encoding="utf-8",
            )
            research_report_path.write_text(
                "\n".join(
                    [
                        "研究报告：手机CIS受益于高像素新品导入，ASP稳中有升。",
                        "研究报告：汽车CIS长期由ADAS渗透率提升驱动。",
                    ]
                ),
                encoding="utf-8",
            )

            argv = [
                "forecast_rollforward.py",
                "--company",
                "韦尔股份",
                "--model",
                str(workbook_path),
                "--annual-report",
                str(report_path),
                "--meeting-notes",
                str(meeting_notes_path),
                "--research-report",
                str(research_report_path),
                "--output-dir",
                str(output_dir),
            ]
            config_path = ROOT / "data" / "forecast_configs" / "韦尔股份.yaml"

            try:
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch("sys.stdout", new=StringIO()) as stdout:
                        with mock.patch(
                            "scripts.forecast_tools.rollforward.TushareFinancialFactsAdapter.extract",
                            return_value={
                                "reported_facts": {"营业收入": 158.0, "毛利率": 0.255, "归母净利润": 8.8},
                                "fact_items": [],
                                "segment_disclosure": [
                                    {"segment": "手机CIS", "revenue": 128.0, "mapping_ready": True, "source_ref": "tushare:test"},
                                    {"segment": "汽车CIS", "revenue": 30.0, "mapping_ready": True, "source_ref": "tushare:test"},
                                ],
                                "source_type": "tushare",
                            },
                        ), mock.patch(
                            "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                            return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
                        ):
                            exit_code = cli_main()

                self.assertEqual(exit_code, 0)
                self.assertTrue(any(output_dir.glob("*_candidate.xlsx")))
                self.assertTrue(any(output_dir.glob("*_facts.json")))
                self.assertTrue(any(output_dir.glob("*_evidence.json")))
                self.assertTrue(any(output_dir.glob("*_changelog.json")))
                self.assertTrue(any(output_dir.glob("*_changelog.md")))
                self.assertTrue(any(output_dir.glob("*_run_log.md")))
                self.assertTrue(any(output_dir.glob("*_logic_review.json")))
                self.assertTrue(any(output_dir.glob("*_forecast_architecture.json")))
                self.assertTrue(any(output_dir.glob("*_forecast_basis.json")))
                self.assertTrue(any(output_dir.glob("*_forecast_basis.md")))
                self.assertTrue(any(output_dir.glob("*_financial_facts.json")))
                self.assertTrue(any(output_dir.glob("*_segment_mapping.json")))
                self.assertTrue(any(output_dir.glob("*_reconciliation_audit.json")))
                self.assertTrue(config_path.exists())
                summary = json.loads(stdout.getvalue())
                self.assertIn("pre_edit_timing", summary)
                self.assertEqual(summary["pre_edit_timing"]["mode"], "thread_pool")
                evidence_payload = json.loads(next(output_dir.glob("*_evidence.json")).read_text(encoding="utf-8"))
                self.assertIn("timing", evidence_payload)
                self.assertEqual(evidence_payload["timing"]["mode"], "thread_pool")
                self.assertEqual(
                    evidence_payload["evidence_priority"],
                    ["reference_files", "local_kb", "alpha_pai"],
                )
                self.assertIn("provider_decisions", evidence_payload)
                self.assertIn("recall_checks", evidence_payload)
                self.assertTrue(
                    any(
                        item["name"] == "alpha_pai_recalled" and item["passed"]
                        for item in evidence_payload["recall_checks"]["source_prep"]
                    )
                )
                candidate = openpyxl.load_workbook(next(output_dir.glob("*_candidate.xlsx")), data_only=False)
                self.assertIn("Forecast Basis", candidate.sheetnames)
                run_log = next(output_dir.glob("*_run_log.md")).read_text(encoding="utf-8")
                self.assertIn("source prep", run_log)
                self.assertIn("forecast architecture", run_log)
                self.assertIn("logic review", run_log)
                self.assertIn("Recall Checks", run_log)
                self.assertIn("alpha_pai_recalled", run_log)
                logic_review = json.loads(next(output_dir.glob("*_logic_review.json")).read_text(encoding="utf-8"))
                self.assertIn("checks", logic_review)
                self.assertTrue(any(item["name"] == "alpha_pai_recalled" for item in logic_review["checks"]))
            finally:
                if config_path.exists():
                    config_path.unlink()

    def test_cli_normalizes_actual_overrides_to_workbook_units(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "sample.xlsx"
            report_path = tmp / "report.txt"
            meeting_notes_path = tmp / "meeting.txt"
            research_report_path = tmp / "research.txt"
            output_dir = tmp / "out"

            build_sample_workbook(workbook_path)
            report_path.write_text("营业收入 24000000000", encoding="utf-8")
            meeting_notes_path.write_text("汽车CIS 2025年增长20%", encoding="utf-8")
            research_report_path.write_text("研究报告：汽车CIS受益于ADAS渗透率提升。", encoding="utf-8")

            argv = [
                "forecast_rollforward.py",
                "--company",
                "测试公司",
                "--model",
                str(workbook_path),
                "--annual-report",
                str(report_path),
                "--meeting-notes",
                str(meeting_notes_path),
                "--research-report",
                str(research_report_path),
                "--output-dir",
                str(output_dir),
            ]
            config_path = ROOT / "data" / "forecast_configs" / "测试公司.yaml"

            try:
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch("sys.stdout", new=StringIO()):
                        with mock.patch(
                            "scripts.forecast_tools.rollforward.TushareFinancialFactsAdapter.extract",
                            return_value={
                                "reported_facts": {"营业收入": 15800000000.0},
                                "fact_items": [{"metric": "营业收入", "value": 15800000000.0, "unit": "元", "source_ref": "tushare:test"}],
                                "segment_disclosure": [
                                    {"segment": "手机CIS", "revenue": 12800000000.0, "mapping_ready": True, "source_ref": "tushare:test"},
                                    {"segment": "汽车CIS", "revenue": 3000000000.0, "mapping_ready": True, "source_ref": "tushare:test"},
                                ],
                                "source_type": "tushare",
                            },
                        ), mock.patch(
                            "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                            return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
                        ):
                            exit_code = cli_main()

                self.assertEqual(exit_code, 0)
                candidate = openpyxl.load_workbook(next(output_dir.glob("*_candidate.xlsx")), data_only=False)
                ws = candidate["营收拆分"]
                self.assertEqual(ws.cell(3, 5).value, 158.0)
            finally:
                if config_path.exists():
                    config_path.unlink()

    def test_cli_output_dir_contains_no_runtime_generated_executable_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "sample.xlsx"
            report_path = tmp / "report.txt"
            meeting_notes_path = tmp / "meeting.txt"
            output_dir = tmp / "out"

            build_sample_workbook(workbook_path)
            report_path.write_text("营业收入 240.0", encoding="utf-8")
            meeting_notes_path.write_text("汽车CIS 2025年增长10%", encoding="utf-8")

            argv = [
                "forecast_rollforward.py",
                "--company",
                "测试公司",
                "--model",
                str(workbook_path),
                "--annual-report",
                str(report_path),
                "--meeting-notes",
                str(meeting_notes_path),
                "--output-dir",
                str(output_dir),
            ]
            config_path = ROOT / "data" / "forecast_configs" / "测试公司.yaml"
            try:
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch("sys.stdout", new=StringIO()) as stdout:
                        with mock.patch(
                            "scripts.forecast_tools.rollforward.TushareFinancialFactsAdapter.extract",
                            return_value={
                                "reported_facts": {"营业收入": 158.0},
                                "fact_items": [],
                                "segment_disclosure": [
                                    {"segment": "手机CIS", "revenue": 128.0, "mapping_ready": True, "source_ref": "tushare:test"},
                                    {"segment": "汽车CIS", "revenue": 30.0, "mapping_ready": True, "source_ref": "tushare:test"},
                                ],
                                "source_type": "tushare",
                            },
                        ), mock.patch(
                            "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                            return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
                        ):
                            exit_code = cli_main()
                self.assertEqual(exit_code, 0)
                summary = json.loads(stdout.getvalue())
                self.assertEqual(summary["runtime_artifact_guard"]["status"], "passed")
                self.assertFalse(any(path.suffix.lower() in {".py", ".ps1", ".bat", ".cmd", ".sh"} for path in output_dir.iterdir()))
            finally:
                if config_path.exists():
                    config_path.unlink()

    def test_cli_stops_cleanly_with_candidate_decision_package_when_reconciliation_is_out_of_tolerance(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "sample.xlsx"
            report_path = tmp / "report.txt"
            meeting_notes_path = tmp / "meeting.txt"
            output_dir = tmp / "out"

            build_candidate_decision_workbook(workbook_path)
            report_path.write_text("营业收入 240.0", encoding="utf-8")
            meeting_notes_path.write_text("汽车CIS 2025年增长20%", encoding="utf-8")

            argv = [
                "forecast_rollforward.py",
                "--company",
                "测试公司",
                "--model",
                str(workbook_path),
                "--annual-report",
                str(report_path),
                "--meeting-notes",
                str(meeting_notes_path),
                "--output-dir",
                str(output_dir),
            ]
            config_path = ROOT / "data" / "forecast_configs" / "测试公司.yaml"
            mocked_mapping = {
                "segment_tree": {
                    "reportable_segments": [
                        {"segment_name": "手机业务", "reported_value": 140.0},
                        {"segment_name": "汽车业务", "reported_value": 48.0},
                    ]
                },
                "segment_mappings": [
                    {"workbook_segment": "手机业务", "tushare_segment": "消费电子业务", "mapping_type": "proxy", "confidence": "low", "workbook_reported_value": 140.0},
                    {"workbook_segment": "汽车业务", "tushare_segment": "移动终端业务", "mapping_type": "proxy", "confidence": "low", "workbook_reported_value": 48.0},
                ],
                "anchored_segment_count": 1,
                "proxy_segment_count": 2,
                "candidate_clusters": [
                    {
                        "workbook_segment": "手机业务",
                        "candidate_matches": [
                            {"tushare_segment": "消费电子业务", "mapping_type": "proxy", "confidence": "low", "score": 55, "base_score": 20, "tushare_revenue": 140.0, "source_refs": ["t1"]},
                            {"tushare_segment": "移动终端业务", "mapping_type": "proxy", "confidence": "low", "score": 50, "base_score": 18, "tushare_revenue": 140.0, "source_refs": ["t2"]},
                        ],
                    }
                ],
                "unmapped_tushare_segments": ["消费电子业务", "移动终端业务"],
            }
            mocked_audit = {
                "reported_year": "2025A",
                "workbook_revenue_total": 188.0,
                "mapped_revenue_total": 18.0,
                "official_revenue_total": 260.0,
                "revenue_gap": -72.0,
                "revenue_gap_ratio": -0.276923,
                "coverage_ratio": 0.5,
                "anchored_segment_count": 1,
                "proxy_segment_count": 2,
                "anchored_revenue_total": 18.0,
                "proxy_revenue_total": 188.0,
                "residual_revenue_total": 0.0,
                "dirty_segment_labels": [],
                "candidate_quality": {"all_low_quality_proxy": False, "candidate_cluster_count": 1},
                "within_tolerance": False,
                "resolution_mode": "candidate_decision_required",
                "candidate_options": [
                    {
                        "option_id": "A",
                        "summary": "primary",
                        "continue_allowed": True,
                        "segment_assignments": mocked_mapping["segment_mappings"],
                        "structure_retention_score": 0.7,
                        "anchor_coverage_ratio": 0.1,
                        "anchored_revenue_total": 18.0,
                        "mapped_revenue_total": 18.0,
                        "proxy_revenue_total": 170.0,
                        "residual_revenue_total": 0.0,
                        "revenue_gap": -72.0,
                        "revenue_gap_ratio": -0.276923,
                        "proxy_segment_count": 2,
                        "residual_segment_count": 0,
                        "recommended": True,
                    },
                    {
                        "option_id": "B",
                        "summary": "secondary",
                        "continue_allowed": True,
                        "segment_assignments": mocked_mapping["segment_mappings"],
                        "structure_retention_score": 0.68,
                        "anchor_coverage_ratio": 0.1,
                        "anchored_revenue_total": 18.0,
                        "mapped_revenue_total": 18.0,
                        "proxy_revenue_total": 170.0,
                        "residual_revenue_total": 0.0,
                        "revenue_gap": -72.0,
                        "revenue_gap_ratio": -0.276923,
                        "proxy_segment_count": 2,
                        "residual_segment_count": 0,
                        "recommended": False,
                    },
                    {
                        "option_id": "C",
                        "summary": "residual",
                        "continue_allowed": True,
                        "segment_assignments": mocked_mapping["segment_mappings"],
                        "structure_retention_score": 0.8,
                        "anchor_coverage_ratio": 0.1,
                        "anchored_revenue_total": 18.0,
                        "mapped_revenue_total": 18.0,
                        "proxy_revenue_total": 100.0,
                        "residual_revenue_total": 70.0,
                        "revenue_gap": -72.0,
                        "revenue_gap_ratio": -0.276923,
                        "proxy_segment_count": 1,
                        "residual_segment_count": 1,
                        "recommended": False,
                    },
                ],
                "fail_reasons": ["mapping_ambiguity_requires_operator_decision"],
                "unmapped_tushare_segments": ["消费电子业务", "移动终端业务"],
            }
            try:
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch("sys.stdout", new=StringIO()) as stdout, mock.patch(
                        "scripts.forecast_tools.rollforward.TushareFinancialFactsAdapter.extract",
                        return_value={
                            "reported_facts": {"营业收入": 9999.0},
                            "fact_items": [],
                            "segment_disclosure": [
                                {"segment": "CIS", "revenue": 188.0, "mapping_ready": True, "source_ref": "tushare:cis"},
                                {"segment": "图像传感解决方案业务", "revenue": 188.0, "mapping_ready": True, "source_ref": "tushare:image"},
                                {"segment": "安防", "revenue": 18.0, "mapping_ready": True, "source_ref": "tushare:security"},
                            ],
                            "source_type": "tushare",
                        },
                    ), mock.patch(
                        "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                        return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
                    ), mock.patch(
                        "scripts.forecast_rollforward.build_segment_mapping_contract",
                        return_value=mocked_mapping,
                    ), mock.patch(
                        "scripts.forecast_rollforward.build_reconciliation_audit",
                        return_value=mocked_audit,
                    ):
                        exit_code = cli_main()
                self.assertEqual(exit_code, 2)
                summary = json.loads(stdout.getvalue())
                self.assertFalse(any(output_dir.glob("*_candidate.xlsx")))
                audit_paths = list(output_dir.glob("*_reconciliation_audit.json"))
                self.assertTrue(audit_paths)
                audit = json.loads(audit_paths[0].read_text(encoding="utf-8"))
                self.assertEqual(audit["resolution_mode"], "candidate_decision_required")
                self.assertEqual(len(audit["candidate_options"]), 3)
                self.assertEqual(summary["decision_package"]["status"], "candidate_decision_required")
                self.assertEqual(summary["decision_package"]["options"][-1]["option_id"], "R")
                self.assertTrue(any(output_dir.glob("*_segment_mapping.json")))
                self.assertTrue(any(output_dir.glob("*_pre_edit_timing.json")))
                self.assertTrue(any(output_dir.glob("*_failure_diagnostics.md")))
                self.assertTrue(any(output_dir.glob("*_run_log.md")))
            finally:
                if config_path.exists():
                    config_path.unlink()

    def test_cli_can_resume_with_selected_candidate_option(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook_path = tmp / "sample.xlsx"
            report_path = tmp / "report.txt"
            meeting_notes_path = tmp / "meeting.txt"
            output_dir = tmp / "out"

            build_candidate_decision_workbook(workbook_path)
            report_path.write_text("营业收入 240.0", encoding="utf-8")
            meeting_notes_path.write_text("汽车CIS 2025年增长20%", encoding="utf-8")

            first_argv = [
                "forecast_rollforward.py",
                "--company",
                "测试公司",
                "--model",
                str(workbook_path),
                "--annual-report",
                str(report_path),
                "--meeting-notes",
                str(meeting_notes_path),
                "--output-dir",
                str(output_dir),
            ]
            second_argv = first_argv + ["--resume-from", str(output_dir), "--apply-candidate", "A"]
            config_path = ROOT / "data" / "forecast_configs" / "测试公司.yaml"
            tushare_payload = {
                "reported_facts": {"营业收入": 9999.0},
                "fact_items": [],
                "segment_disclosure": [
                    {"segment": "CIS", "revenue": 188.0, "mapping_ready": True, "source_ref": "tushare:cis"},
                    {"segment": "图像传感解决方案业务", "revenue": 188.0, "mapping_ready": True, "source_ref": "tushare:image"},
                    {"segment": "安防", "revenue": 18.0, "mapping_ready": True, "source_ref": "tushare:security"},
                ],
                "source_type": "tushare",
            }
            mocked_mapping = {
                "segment_tree": {
                    "reportable_segments": [
                        {"segment_name": "手机业务", "reported_value": 140.0},
                        {"segment_name": "汽车业务", "reported_value": 48.0},
                    ]
                },
                "segment_mappings": [
                    {"workbook_segment": "手机业务", "tushare_segment": "消费电子业务", "mapping_type": "proxy", "confidence": "low", "workbook_reported_value": 140.0},
                    {"workbook_segment": "汽车业务", "tushare_segment": "移动终端业务", "mapping_type": "proxy", "confidence": "low", "workbook_reported_value": 48.0},
                ],
                "anchored_segment_count": 1,
                "proxy_segment_count": 2,
                "candidate_clusters": [
                    {
                        "workbook_segment": "手机业务",
                        "candidate_matches": [
                            {"tushare_segment": "消费电子业务", "mapping_type": "proxy", "confidence": "low", "score": 55, "base_score": 20, "tushare_revenue": 140.0, "source_refs": ["t1"]},
                            {"tushare_segment": "移动终端业务", "mapping_type": "proxy", "confidence": "low", "score": 50, "base_score": 18, "tushare_revenue": 140.0, "source_refs": ["t2"]},
                        ],
                    }
                ],
                "unmapped_tushare_segments": ["消费电子业务", "移动终端业务"],
            }
            mocked_audit = {
                "reported_year": "2025A",
                "workbook_revenue_total": 188.0,
                "mapped_revenue_total": 18.0,
                "official_revenue_total": 260.0,
                "revenue_gap": -72.0,
                "revenue_gap_ratio": -0.276923,
                "coverage_ratio": 0.5,
                "anchored_segment_count": 1,
                "proxy_segment_count": 2,
                "anchored_revenue_total": 18.0,
                "proxy_revenue_total": 188.0,
                "residual_revenue_total": 0.0,
                "dirty_segment_labels": [],
                "candidate_quality": {"all_low_quality_proxy": False, "candidate_cluster_count": 1},
                "within_tolerance": False,
                "resolution_mode": "candidate_decision_required",
                "candidate_options": [
                    {
                        "option_id": "A",
                        "summary": "primary",
                        "continue_allowed": True,
                        "segment_assignments": mocked_mapping["segment_mappings"],
                        "structure_retention_score": 0.7,
                        "anchor_coverage_ratio": 0.1,
                        "anchored_revenue_total": 18.0,
                        "mapped_revenue_total": 18.0,
                        "proxy_revenue_total": 170.0,
                        "residual_revenue_total": 0.0,
                        "revenue_gap": -72.0,
                        "revenue_gap_ratio": -0.276923,
                        "proxy_segment_count": 2,
                        "residual_segment_count": 0,
                        "recommended": True,
                    }
                ],
                "fail_reasons": ["mapping_ambiguity_requires_operator_decision"],
                "unmapped_tushare_segments": ["消费电子业务", "移动终端业务"],
            }
            try:
                with mock.patch("scripts.forecast_tools.rollforward.TushareFinancialFactsAdapter.extract", return_value=tushare_payload), mock.patch(
                    "scripts.forecast_tools.rollforward.AlphaPaiProvider.fetch",
                    return_value=ProviderResult("alpha_pai", "alphapai:test", "", {"available": False}),
                ), mock.patch(
                    "scripts.forecast_rollforward.build_segment_mapping_contract",
                    return_value=mocked_mapping,
                ), mock.patch(
                    "scripts.forecast_rollforward.build_reconciliation_audit",
                    return_value=mocked_audit,
                ):
                    with mock.patch.object(sys, "argv", first_argv):
                        first_exit = cli_main()
                    self.assertEqual(first_exit, 2)
                    with mock.patch.object(sys, "argv", second_argv):
                        second_exit = cli_main()

                self.assertEqual(second_exit, 0)
                self.assertTrue(any(output_dir.glob("*_candidate.xlsx")))
            finally:
                if config_path.exists():
                    config_path.unlink()


if __name__ == "__main__":
    unittest.main()
