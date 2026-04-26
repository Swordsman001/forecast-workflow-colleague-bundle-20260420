# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.forecast_tools.rollforward import (
    CompanyConfigManager,
    WorkbookRollforwardEngine,
    audit_runtime_output_dir,
    build_forecast_architecture,
    build_forecast_basis_payload,
    build_reconciliation_audit,
    build_segment_mapping_contract,
    build_workbook_map_contract,
    collect_pre_edit_inputs,
    ensure_runtime_artifact_path_allowed,
    materialize_selected_candidate_mapping,
    materialize_forecast_basis_sheet,
    render_forecast_basis_markdown,
    render_run_log,
    review_forecast_architecture,
)
from scripts.forecast_tools.build_cell_instructions import validate_forecast_basis, validate_workbook_map


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def infer_report_year(blueprint) -> int:
    return min(blueprint.forecast_columns)


def infer_workbook_unit_scale(*, model_path: Path, blueprint) -> float:
    wb = openpyxl.load_workbook(model_path, data_only=False)
    ws = wb[blueprint.primary_sheet]
    candidates: list[str] = []
    for row in range(1, min(blueprint.header_row, 5) + 1):
        for col in range(1, min(blueprint.label_column + 2, 4) + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.strip():
                candidates.append(value.strip())
    header_blob = " ".join(candidates)
    if "亿元" in header_blob:
        return 1e8
    if "百万元" in header_blob or "百万" in header_blob:
        return 1e6
    if "万元" in header_blob:
        return 1e4
    return 1.0


def _is_ratio_like_metric(label: str) -> bool:
    text = str(label or "")
    return any(token in text for token in ("率", "%", "yoy", "YOY", "同比"))


def normalize_financial_facts_for_workbook_units(*, financial_facts: dict, model_path: Path, blueprint) -> dict:
    scale = infer_workbook_unit_scale(model_path=model_path, blueprint=blueprint)
    if scale == 1.0:
        return financial_facts
    normalized = {
        **financial_facts,
        "reported_facts": dict(financial_facts.get("reported_facts", {})),
        "fact_items": [dict(item) for item in financial_facts.get("fact_items", [])],
        "segment_disclosure": [dict(item) for item in financial_facts.get("segment_disclosure", [])],
    }
    for label, value in list(normalized["reported_facts"].items()):
        if not isinstance(value, (int, float)):
            continue
        if _is_ratio_like_metric(str(label)):
            continue
        if abs(float(value)) >= max(scale / 100, 1_000_000):
            normalized["reported_facts"][label] = round(float(value) / scale, 6)
    for item in normalized.get("fact_items", []):
        value = item.get("value")
        unit = str(item.get("unit") or "")
        if not isinstance(value, (int, float)):
            continue
        if _is_ratio_like_metric(str(item.get("metric") or "")):
            continue
        if unit == "元" or abs(float(value)) >= max(scale / 100, 1_000_000):
            item["value"] = round(float(value) / scale, 6)
            item["unit"] = "亿元" if scale == 1e8 else unit
    for item in normalized.get("segment_disclosure", []):
        revenue = item.get("revenue")
        unit = str(item.get("unit") or "")
        if not isinstance(revenue, (int, float)):
            continue
        if unit == "元" or abs(float(revenue)) >= max(scale / 100, 1_000_000):
            item["revenue"] = round(float(revenue) / scale, 6)
            item["unit"] = "亿元" if scale == 1e8 else unit
    return normalized


normalize_reported_facts_for_workbook_units = normalize_financial_facts_for_workbook_units


def build_patch_guidance_from_architecture(*, model_path: Path, blueprint, report_year: int, architecture: dict) -> dict:
    wb = openpyxl.load_workbook(model_path, data_only=False)
    ws = wb[blueprint.primary_sheet]
    base_col = blueprint.forecast_columns.get(report_year)
    guidance: dict[str, dict] = {}
    if base_col is None:
        return guidance
    first_forecast_year = f"{report_year + 1}E"
    for segment in architecture.get("segments", []):
        row_label = str(segment.get("segment") or "")
        row = blueprint.row_labels.get(row_label)
        if row is None:
            continue
        base_value = ws.cell(row, base_col).value
        forecast_values = segment.get("forecast_values") or {}
        next_value = forecast_values.get(first_forecast_year)
        growth_rate = None
        if isinstance(base_value, (int, float)) and base_value not in {0, 0.0} and isinstance(next_value, (int, float)):
            growth_rate = round(float(next_value) / float(base_value) - 1, 6)
        guidance[row_label] = {
            "claim": segment.get("evidence_summary") or "",
            "growth_rate": growth_rate,
            "year_values": forecast_values,
            "dependent_metric_values": segment.get("dependent_metric_values", {}),
            "confidence": 0.7 if segment.get("source_tier") != "alpha_pai" else 0.4,
            "review_required": segment.get("review_flag") not in {None, "", "none"},
        }
    return guidance


def _build_artifact_paths(*, output_dir: Path, stem: str, report_year: int) -> dict[str, Path]:
    return {
        "forecast_architecture": output_dir / f"{stem}_{report_year}_forecast_architecture.json",
        "financial_facts": output_dir / f"{stem}_{report_year}_financial_facts.json",
        "segment_mapping": output_dir / f"{stem}_{report_year}_segment_mapping.json",
        "reconciliation_audit": output_dir / f"{stem}_{report_year}_reconciliation_audit.json",
        "pre_edit_timing": output_dir / f"{stem}_{report_year}_pre_edit_timing.json",
        "workbook_map": output_dir / f"{stem}_{report_year}_workbook_map.json",
        "forecast_basis_json": output_dir / f"{stem}_{report_year}_forecast_basis.json",
        "forecast_basis_md": output_dir / f"{stem}_{report_year}_forecast_basis.md",
        "logic_review": output_dir / f"{stem}_{report_year}_logic_review.json",
        "run_log": output_dir / f"{stem}_{report_year}_run_log.md",
        "failure_diagnostics": output_dir / f"{stem}_{report_year}_failure_diagnostics.md",
    }


def _write_json_artifact(path: Path, payload: dict) -> None:
    ensure_runtime_artifact_path_allowed(path)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_json_artifact(path: Path) -> dict:
    ensure_runtime_artifact_path_allowed(path)
    return json.loads(path.read_text(encoding="utf-8"))


def _human_candidate_option_description(option: dict) -> dict:
    option_id = str(option.get("option_id") or "")
    if option_id == "A":
        meaning = "尽量按程序当前最像的官方口径去归类这些歧义业务，然后继续往下跑。"
        pros = "更积极地尝试官方对齐，后续自动执行最顺。"
        risks = "如果程序当前理解错了，会把细分业务挂到不合适的大类下面。"
    elif option_id == "B":
        meaning = "不用程序当前第一判断，改用第二套官方归类方式继续跑。"
        pros = "给出另一种官方理解路径，方便和 A 做比较。"
        risks = "本质仍是强行归类，只是换了另一种强行方式。"
    elif option_id == "C":
        meaning = "保留 workbook 原有细分结构，把暂时不能可靠挂官方口径的业务明确留在 residual。"
        pros = "最保留研究结构，也最少假装程序已经完全理解官方口径。"
        risks = "官方对齐会更弱，后面还需要继续补桥接逻辑。"
    else:
        meaning = "取消本次继续执行，先回去重新桥接上游映射。"
        pros = "不用在当前质量不足的桥接上冒险继续，最保守。"
        risks = "这次不会继续生成最终 candidate workbook，需要先修正桥接。"
    return {
        "option_id": option_id,
        "what_it_means": meaning,
        "pros": pros,
        "risks": risks,
        "structure_retention_score": option.get("structure_retention_score"),
        "revenue_gap_ratio": option.get("revenue_gap_ratio"),
        "recommended": bool(option.get("recommended")),
    }


def _build_human_decision_package(*, company: str, reconciliation_audit: dict, segment_mapping: dict) -> dict:
    resolution_mode = str(reconciliation_audit.get("resolution_mode") or "hard_stop")
    candidate_options = list(reconciliation_audit.get("candidate_options", []))
    dirty_segment_labels = [str(item) for item in reconciliation_audit.get("dirty_segment_labels", [])]
    candidate_clusters = list(segment_mapping.get("candidate_clusters", []))
    ambiguous_segments = [str(item.get("workbook_segment") or "") for item in candidate_clusters if str(item.get("workbook_segment") or "").strip()]
    if resolution_mode != "candidate_decision_required":
        return {
            "status": resolution_mode,
            "current_blocker": "当前不适合让执行者做 A/B/C 选择，需要先修正上游识别或桥接质量。",
            "ambiguous_segments": ambiguous_segments,
            "dirty_segment_labels": dirty_segment_labels,
        }
    options = [_human_candidate_option_description(option) for option in candidate_options]
    options.append(
        {
            "option_id": "R",
            "what_it_means": "取消本次继续执行，先回去重新桥接上游映射。",
            "pros": "不用在当前质量不足的桥接上冒险继续，最保守。",
            "risks": "这次不会继续生成最终 candidate workbook，需要先修正桥接。",
            "structure_retention_score": None,
            "revenue_gap_ratio": None,
            "recommended": False,
        }
    )
    return {
        "status": "candidate_decision_required",
        "current_blocker": (
            f"{company} 当前存在少数业务还不能可靠对应到官方口径。程序先给出几种可继续执行的理解方式，你可以选 A/B/C，或者选 R 取消这次继续执行。"
        ),
        "ambiguous_segments": ambiguous_segments,
        "dirty_segment_labels": dirty_segment_labels,
        "options": options,
    }


def _write_failure_diagnostics(
    *,
    path: Path,
    company: str,
    report_year: int,
    resolution_mode: str,
    fail_reasons: list[str],
    candidate_options: list[dict],
    decision_package: dict | None = None,
) -> None:
    ensure_runtime_artifact_path_allowed(path)
    lines = [
        "# Forecast Workflow Diagnostics",
        "",
        f"- company: {company}",
        f"- reported_year: {report_year}A",
        f"- resolution_mode: {resolution_mode}",
        f"- fail_reasons: {', '.join(fail_reasons) if fail_reasons else 'none'}",
        "",
    ]
    if decision_package:
        lines.extend(
            [
                "## Decision Package",
                "",
                f"- status: {decision_package.get('status')}",
                f"- current_blocker: {decision_package.get('current_blocker')}",
            ]
        )
        ambiguous_segments = decision_package.get("ambiguous_segments") or []
        if ambiguous_segments:
            lines.append(f"- ambiguous_segments: {', '.join(str(item) for item in ambiguous_segments)}")
        dirty_segment_labels = decision_package.get("dirty_segment_labels") or []
        if dirty_segment_labels:
            lines.append(f"- dirty_segment_labels: {', '.join(str(item) for item in dirty_segment_labels)}")
        lines.append("")
    if candidate_options:
        lines.extend(["## Candidate Options", ""])
        option_descriptions = {
            str(item.get("option_id")): item for item in (decision_package or {}).get("options", [])
        }
        for option in candidate_options:
            description = option_descriptions.get(str(option.get("option_id")))
            lines.append(
                f"- {option['option_id']}: {option['summary']} | "
                f"structure_retention_score={option['structure_retention_score']} | "
                f"proxy={option['proxy_segment_count']} | "
                f"residual={option['residual_segment_count']} | "
                f"gap={option['revenue_gap_ratio']} | "
                f"recommended={option['recommended']}"
            )
            if description:
                lines.append(f"  - 人话解释: {description.get('what_it_means')}")
                lines.append(f"  - 好处: {description.get('pros')}")
                lines.append(f"  - 风险: {description.get('risks')}")
    manual_option = next(
        (item for item in (decision_package or {}).get("options", []) if str(item.get("option_id") or "") == "R"),
        None,
    )
    if manual_option:
        lines.extend(
            [
                "",
                "## Manual Option",
                "",
                f"- R: {manual_option.get('what_it_means')}",
                f"  - 好处: {manual_option.get('pros')}",
                f"  - 风险: {manual_option.get('risks')}",
            ]
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_failure_run_log(
    *,
    path: Path,
    company: str,
    pre_edit_timing: dict,
    reconciliation_audit: dict,
    provider_decisions: list[dict],
) -> None:
    ensure_runtime_artifact_path_allowed(path)
    lines = [
        "# Forecast Rollforward Run Log",
        "",
        f"- company: {company}",
        f"- generated_at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- resolution_mode: {reconciliation_audit.get('resolution_mode')}",
        "",
        "## Phases",
        "",
        "| phase | status | note |",
        "|---|---|---|",
        f"| source prep | completed | mode={pre_edit_timing.get('mode')} workers={pre_edit_timing.get('max_workers')} |",
        f"| reconciliation | blocked | reasons={', '.join(reconciliation_audit.get('fail_reasons', [])) or 'none'} |",
        "",
        "## Provider Decisions",
        "",
    ]
    for item in provider_decisions:
        lines.append(
            f"- {item.get('source_tier')}: {item.get('decision')} | query={item.get('query') or ''}"
        )
    lines.extend(["", "## Candidate Options", ""])
    for option in reconciliation_audit.get("candidate_options", []):
        lines.append(
            f"- {option.get('option_id')}: recommended={option.get('recommended')} | "
            f"structure_retention_score={option.get('structure_retention_score')} | "
            f"gap={option.get('revenue_gap_ratio')}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="非标盈利预测表滚动更新器")
    parser.add_argument("--company", required=True, help="公司名称")
    parser.add_argument("--model", required=True, help="历史模型 Excel 路径")
    parser.add_argument("--annual-report", required=True, help="最新年报 PDF 或文本路径")
    parser.add_argument("--meeting-notes", required=False, help="业绩交流会纪要文本路径")
    parser.add_argument("--research-report", required=False, help="研究报告文本路径")
    parser.add_argument("--output-dir", required=True, help="输出目录")
    parser.add_argument("--resume-from", required=False, help="从已有决策包目录继续执行")
    parser.add_argument("--apply-candidate", required=False, help="应用候选方案 A/B/C 后继续执行")
    args = parser.parse_args()
    if bool(args.resume_from) != bool(args.apply_candidate):
        parser.error("--resume-from and --apply-candidate must be provided together")

    repo_root = Path(__file__).resolve().parents[1]
    model_path = Path(args.model).resolve()
    report_path = Path(args.annual_report).resolve()
    meeting_notes_path = Path(args.meeting_notes).resolve() if args.meeting_notes else None
    research_report_path = Path(args.research_report).resolve() if args.research_report else None
    output_dir = Path(args.output_dir).resolve()
    resume_dir = Path(args.resume_from).resolve() if args.resume_from else None
    output_dir.mkdir(parents=True, exist_ok=True)

    pre_edit_inputs = collect_pre_edit_inputs(
        repo_root=repo_root,
        company=args.company,
        model_path=model_path,
        report_path=report_path,
        meeting_notes_path=meeting_notes_path,
        research_report_path=research_report_path,
    )
    blueprint = pre_edit_inputs["blueprint"]
    report_year = infer_report_year(blueprint)
    target_far_year = report_year + 3
    stem = model_path.stem
    artifact_paths = _build_artifact_paths(output_dir=output_dir, stem=stem, report_year=report_year)
    resume_artifact_paths = (
        _build_artifact_paths(output_dir=resume_dir, stem=stem, report_year=report_year)
        if resume_dir is not None
        else None
    )

    config_manager = CompanyConfigManager(repo_root)
    config_path = config_manager.ensure_config(args.company, blueprint)

    financial_facts = normalize_financial_facts_for_workbook_units(
        financial_facts=pre_edit_inputs["financial_facts"],
        model_path=model_path,
        blueprint=blueprint,
    )
    meeting_notes_facts = pre_edit_inputs["meeting_notes_facts"]
    evidence_payload = pre_edit_inputs["evidence_payload"]
    if resume_artifact_paths is not None:
        if str(args.apply_candidate).upper() == "R":
            summary = {
                "candidate_workbook": None,
                "report_year": report_year,
                "resolution_mode": "operator_cancelled_rebridge",
                "message": "operator selected R: cancel current continuation and return to upstream re-bridge",
                "resume_from": str(resume_dir),
            }
            print(json.dumps(summary, ensure_ascii=False, indent=2))
            return 3
        prior_mapping_path = resume_artifact_paths["segment_mapping"]
        prior_audit_path = resume_artifact_paths["reconciliation_audit"]
        if not prior_mapping_path.exists() or not prior_audit_path.exists():
            raise FileNotFoundError(
                f"resume artifacts missing: {prior_mapping_path} / {prior_audit_path}"
            )
        loaded_segment_mapping = _read_json_artifact(prior_mapping_path)
        loaded_reconciliation_audit = _read_json_artifact(prior_audit_path)
        segment_mapping, reconciliation_audit = materialize_selected_candidate_mapping(
            segment_mapping=loaded_segment_mapping,
            reconciliation_audit=loaded_reconciliation_audit,
            option_id=str(args.apply_candidate),
        )
    else:
        segment_mapping = build_segment_mapping_contract(
            workbook_path=model_path,
            blueprint=blueprint,
            report_year=report_year,
            financial_facts=financial_facts,
            meeting_notes_facts=meeting_notes_facts,
            evidence_payload=evidence_payload,
        )
        reconciliation_audit = build_reconciliation_audit(
            workbook_path=model_path,
            blueprint=blueprint,
            report_year=report_year,
            financial_facts=financial_facts,
            segment_mapping=segment_mapping,
        )
    _write_json_artifact(artifact_paths["financial_facts"], financial_facts)
    _write_json_artifact(artifact_paths["segment_mapping"], segment_mapping)
    _write_json_artifact(artifact_paths["reconciliation_audit"], reconciliation_audit)
    _write_json_artifact(artifact_paths["pre_edit_timing"], pre_edit_inputs["timing"])
    if reconciliation_audit.get("resolution_mode") != "automatic_pass":
        decision_package = _build_human_decision_package(
            company=args.company,
            reconciliation_audit=reconciliation_audit,
            segment_mapping=segment_mapping,
        )
        _write_failure_diagnostics(
            path=artifact_paths["failure_diagnostics"],
            company=args.company,
            report_year=report_year,
            resolution_mode=str(reconciliation_audit.get("resolution_mode") or "hard_stop"),
            fail_reasons=[str(item) for item in reconciliation_audit.get("fail_reasons", [])],
            candidate_options=list(reconciliation_audit.get("candidate_options", [])),
            decision_package=decision_package,
        )
        _write_failure_run_log(
            path=artifact_paths["run_log"],
            company=args.company,
            pre_edit_timing=pre_edit_inputs["timing"],
            reconciliation_audit=reconciliation_audit,
            provider_decisions=evidence_payload.get("provider_decisions", []),
        )
        summary = {
            "candidate_workbook": None,
            "financial_facts": str(artifact_paths["financial_facts"]),
            "segment_mapping": str(artifact_paths["segment_mapping"]),
            "reconciliation_audit": str(artifact_paths["reconciliation_audit"]),
            "pre_edit_timing_path": str(artifact_paths["pre_edit_timing"]),
            "failure_diagnostics": str(artifact_paths["failure_diagnostics"]),
            "run_log": str(artifact_paths["run_log"]),
            "report_year": report_year,
            "resolution_mode": reconciliation_audit.get("resolution_mode"),
            "candidate_option_count": len(reconciliation_audit.get("candidate_options", [])),
            "decision_package": decision_package,
            "pre_edit_timing": pre_edit_inputs["timing"],
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 2 if reconciliation_audit.get("resolution_mode") == "candidate_decision_required" else 1
    workbook_map = build_workbook_map_contract(
        workbook_path=model_path,
        blueprint=blueprint,
        report_year=report_year,
        target_far_year=target_far_year,
    )
    validate_workbook_map(workbook_map)
    forecast_architecture = build_forecast_architecture(
        company=args.company,
        report_year=report_year,
        financial_facts=financial_facts,
        meeting_notes_facts=meeting_notes_facts,
        evidence_payload=evidence_payload,
        workbook_path=model_path,
        blueprint=blueprint,
    )
    logic_review = review_forecast_architecture(
        forecast_architecture=forecast_architecture,
        financial_facts=financial_facts,
        evidence_payload=evidence_payload,
        segment_mapping=segment_mapping,
        reconciliation_audit=reconciliation_audit,
    )

    actual_overrides = {
        label: value
        for label, value in financial_facts.get("reported_facts", {}).items()
        if value is not None and label in blueprint.row_labels
    }

    patch_guidance = build_patch_guidance_from_architecture(
        model_path=model_path,
        blueprint=blueprint,
        report_year=report_year,
        architecture=forecast_architecture,
    )

    result = WorkbookRollforwardEngine().rollforward(
        workbook_path=model_path,
        blueprint=blueprint,
        report_year=report_year,
        actual_overrides=actual_overrides,
        target_far_year=target_far_year,
        output_dir=output_dir,
        meeting_guidance=patch_guidance,
        evidence_payload=evidence_payload,
        config_path=config_path,
        facts_payload={
            "reported_facts": financial_facts.get("reported_facts", {}),
            "fact_items": financial_facts.get("fact_items", []),
            "segment_disclosure": financial_facts.get("segment_disclosure", []),
            "bridge_facts": patch_guidance,
            "forecast_inputs": {},
        },
    )

    forecast_basis = build_forecast_basis_payload(
        company=args.company,
        report_year=report_year,
        financial_facts=financial_facts,
        meeting_notes_facts=meeting_notes_facts,
        evidence_payload=evidence_payload,
        forecast_architecture=forecast_architecture,
    )
    validate_forecast_basis(forecast_basis)
    materialize_forecast_basis_sheet(
        workbook_path=result.output_workbook,
        forecast_basis=forecast_basis,
        evidence_store=forecast_basis["evidence_store"],
    )

    forecast_architecture_path = artifact_paths["forecast_architecture"]
    financial_facts_path = artifact_paths["financial_facts"]
    segment_mapping_path = artifact_paths["segment_mapping"]
    reconciliation_audit_path = artifact_paths["reconciliation_audit"]
    workbook_map_path = artifact_paths["workbook_map"]
    forecast_basis_json_path = artifact_paths["forecast_basis_json"]
    forecast_basis_md_path = artifact_paths["forecast_basis_md"]
    logic_review_path = artifact_paths["logic_review"]
    run_log_path = artifact_paths["run_log"]
    for artifact_path in (
        forecast_architecture_path,
        financial_facts_path,
        segment_mapping_path,
        reconciliation_audit_path,
        workbook_map_path,
        forecast_basis_json_path,
        forecast_basis_md_path,
        logic_review_path,
        run_log_path,
    ):
        ensure_runtime_artifact_path_allowed(artifact_path)
    forecast_architecture_path.write_text(
        json.dumps(forecast_architecture, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    financial_facts_path.write_text(json.dumps(financial_facts, ensure_ascii=False, indent=2), encoding="utf-8")
    segment_mapping_path.write_text(json.dumps(segment_mapping, ensure_ascii=False, indent=2), encoding="utf-8")
    reconciliation_audit_path.write_text(
        json.dumps(reconciliation_audit, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    workbook_map_path.write_text(json.dumps(workbook_map, ensure_ascii=False, indent=2), encoding="utf-8")
    forecast_basis_json_path.write_text(json.dumps(forecast_basis, ensure_ascii=False, indent=2), encoding="utf-8")
    forecast_basis_md_path.write_text(render_forecast_basis_markdown(forecast_basis), encoding="utf-8")
    logic_review_path.write_text(json.dumps(logic_review, ensure_ascii=False, indent=2), encoding="utf-8")
    recall_checks = {
        "source_prep": evidence_payload.get("recall_checks", {}).get("source_prep", []),
        "forecast_architecture": forecast_architecture.get("recall_checks", []),
        "logic_review": [
            item for item in logic_review.get("checks", [])
            if str(item.get("name") or "").startswith("alpha_pai_")
        ],
    }
    runtime_artifact_guard = {"status": "pending", "checked_dir": str(output_dir), "executable_artifacts": []}
    run_log_path.write_text(
        render_run_log(
            company=args.company,
            pre_edit_timing=pre_edit_inputs["timing"],
            logic_review=logic_review,
            provider_decisions=evidence_payload.get("provider_decisions", []),
            recall_checks=recall_checks,
            output_workbook=result.output_workbook,
            parity_audit=result.parity_audit,
            runtime_artifact_guard=runtime_artifact_guard,
        ),
        encoding="utf-8",
    )
    runtime_artifact_guard = audit_runtime_output_dir(output_dir)
    if runtime_artifact_guard["status"] != "passed":
        raise RuntimeError(
            "runtime artifact guard failed: "
            + ", ".join(runtime_artifact_guard.get("executable_artifacts", []))
        )
    run_log_path.write_text(
        render_run_log(
            company=args.company,
            pre_edit_timing=pre_edit_inputs["timing"],
            logic_review=logic_review,
            provider_decisions=evidence_payload.get("provider_decisions", []),
            recall_checks=recall_checks,
            output_workbook=result.output_workbook,
            runtime_artifact_guard=runtime_artifact_guard,
        ),
        encoding="utf-8",
    )

    summary = {
        "candidate_workbook": str(result.output_workbook),
        "facts": str(result.facts_path),
        "evidence": str(result.evidence_path),
        "changelog_json": str(result.changelog_json),
        "changelog_md": str(result.changelog_md),
        "config": str(result.config_path),
        "forecast_architecture": str(forecast_architecture_path),
        "financial_facts": str(financial_facts_path),
        "segment_mapping": str(segment_mapping_path),
        "reconciliation_audit": str(reconciliation_audit_path),
        "workbook_map": str(workbook_map_path),
        "forecast_basis_json": str(forecast_basis_json_path),
        "forecast_basis_md": str(forecast_basis_md_path),
        "logic_review": str(logic_review_path),
        "run_log": str(run_log_path),
        "report_year": report_year,
        "change_count": len(result.change_records),
        "pre_edit_timing": pre_edit_inputs["timing"],
        "parity_audit": result.parity_audit,
        "runtime_artifact_guard": runtime_artifact_guard,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
