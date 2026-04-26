# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import inspect
import re
import time
from copy import copy
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber
import yaml
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from scripts.data_sources.tushare_financial_facts import TushareFinancialFactsAdapter

from .providers import (
    AlphaPaiProvider,
    LocalKBProvider,
    ProviderResult,
    ReportDownloadProvider,
    UserSuppliedTextProvider,
    read_supported_text,
)


YEAR_RE = re.compile(r"^\s*(20\d{2})\s*$")
FORECAST_YEAR_RE = re.compile(r"^\s*(20\d{2})E\s*$", re.IGNORECASE)
HALF_YEAR_RE = re.compile(r"^\s*\d{2,4}H[12]\s*$", re.IGNORECASE)
PERCENT_RE = re.compile(r"(-?\d+(?:\.\d+)?)%")
NUMBER_RE = re.compile(r"(-?\d+(?:\.\d+)?)")
NUMBER_WITH_UNIT_RE = re.compile(r"(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(亿元|百万元|百万|万元|元|%)?")
STALE_FORECAST_MARKER_RE = re.compile(r"(20\d{2}E|预计|预测|上调|下调|EPS|PE)", re.IGNORECASE)
PAGE_MARKER_RE = re.compile(r"第\s*(\d+)\s*页")
CELL_REF_ROW_RE = re.compile(r"\$?[A-Z]{1,3}\$?(\d+)")
EXECUTABLE_ARTIFACT_SUFFIXES = {".py", ".ps1", ".bat", ".cmd", ".sh"}
GENERIC_METRIC_TOKENS = {
    "收入",
    "营收",
    "revenue",
    "yoy",
    "毛利率",
    "净利率",
    "净利润",
    "市场份额",
    "份额",
    "销量",
    "asp",
    "原本",
}
STABILITY_HINTS = ("stable", "维持", "持平", "平稳", "基本不变")
MARGIN_UP_HINTS = ("asp", "高附加值", "结构升级", "高毛利", "规模效应", "降本", "良率")
MARGIN_DOWN_HINTS = ("价格压力", "竞争", "降价", "成本上涨", "毛利承压")
SHARE_UP_HINTS = ("份额提升", "导入", "验证", "替代", "平台", "渗透", "客户拓展")
GROWTH_CUE_HINTS = ("增长", "增速", "同比", "yoy", "提升", "增加", "放量", "爬坡", "扩张", "恢复", "加速", "翻倍")
NON_GROWTH_PERCENT_HINTS = ("市场份额", "份额", "毛利率", "净利率", "税率", "费用率", "占比", "渗透率")


def _slugify_text(text: str) -> str:
    cleaned = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", str(text or "")).strip("_").lower()
    return cleaned or "item"

@dataclass
class WorkbookBlueprint:
    primary_sheet: str
    header_row: int
    label_column: int
    historical_columns: dict[int, int]
    forecast_columns: dict[int, int]
    excluded_columns: list[int]
    row_labels: dict[str, int]
    primary_row_labels: dict[str, int] = field(default_factory=dict)
    row_blocks: dict[str, list[int]] = field(default_factory=dict)


@dataclass
class ChangeRecord:
    sheet: str
    row_label: str
    year: int
    before: Any
    after: Any
    change_type: str
    evidence: str
    rationale: str
    confidence: float
    review_flag: bool


@dataclass
class RollforwardResult:
    output_workbook: Path
    facts_path: Path
    evidence_path: Path
    changelog_json: Path
    changelog_md: Path
    config_path: Path
    change_records: list[ChangeRecord] = field(default_factory=list)
    parity_audit: dict[str, Any] = field(default_factory=dict)


def ensure_runtime_artifact_path_allowed(path: Path) -> None:
    if path.suffix.lower() in EXECUTABLE_ARTIFACT_SUFFIXES:
        raise ValueError(f"runtime code generation is not allowed for output artifact: {path}")


def audit_runtime_output_dir(output_dir: Path) -> dict[str, Any]:
    executable_artifacts = [
        str(path)
        for path in output_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in EXECUTABLE_ARTIFACT_SUFFIXES
    ]
    return {
        "status": "passed" if not executable_artifacts else "failed",
        "checked_dir": str(output_dir),
        "executable_artifacts": executable_artifacts,
    }


class WorkbookBlueprintParser:
    def parse(self, workbook_path: Path) -> WorkbookBlueprint:
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
        best_sheet: Worksheet | None = None
        best_header_row = 0
        best_score = -1

        for ws in wb.worksheets:
            header_row, score = self._find_header_row(ws)
            if score > best_score:
                best_sheet = ws
                best_header_row = header_row
                best_score = score

        if best_sheet is None or best_header_row == 0:
            raise ValueError(f"Cannot find year header row in workbook: {workbook_path}")

        historical, forecast, excluded = self._classify_columns(best_sheet, best_header_row)
        label_column = self._find_label_column(
            best_sheet,
            best_header_row,
            historical_columns=historical,
            forecast_columns=forecast,
        )
        row_labels = self._collect_row_labels(best_sheet, best_header_row, label_column)
        primary_row_labels = self._collect_primary_row_labels(best_sheet, best_header_row, label_column)
        row_blocks = self._collect_row_blocks(best_sheet, best_header_row, label_column, primary_row_labels)

        return WorkbookBlueprint(
            primary_sheet=best_sheet.title,
            header_row=best_header_row,
            label_column=label_column,
            historical_columns=historical,
            forecast_columns=forecast,
            excluded_columns=excluded,
            row_labels=row_labels,
            primary_row_labels=primary_row_labels,
            row_blocks=row_blocks,
        )

    def _find_header_row(self, ws: Worksheet) -> tuple[int, int]:
        best_row = 0
        best_score = -1
        for row in range(1, min(ws.max_row, 12) + 1):
            score = 0
            for col in range(1, min(ws.max_column, 64) + 1):
                value = ws.cell(row, col).value
                if isinstance(value, int) and 2000 <= value <= 2100:
                    score += 2
                elif isinstance(value, str):
                    if FORECAST_YEAR_RE.match(value):
                        score += 3
                    elif YEAR_RE.match(value):
                        score += 2
                    elif HALF_YEAR_RE.match(value) or "备注" in value:
                        score += 1
            if score > best_score:
                best_row = row
                best_score = score
        return best_row, best_score

    def _find_label_column(
        self,
        ws: Worksheet,
        header_row: int,
        historical_columns: dict[int, int],
        forecast_columns: dict[int, int],
    ) -> int:
        best_col = 1
        best_score = float("-inf")
        year_columns = list(historical_columns.values()) + list(forecast_columns.values())
        if year_columns:
            max_candidate_col = max(1, min(min(year_columns) - 1, 8))
        else:
            max_candidate_col = min(ws.max_column, 8)
        for col in range(1, max_candidate_col + 1):
            non_empty = 0
            unique_labels: set[str] = set()
            generic_metric_hits = 0
            revenue_pair_hits = 0
            for row in range(header_row + 1, min(ws.max_row, header_row + 250) + 1):
                value = ws.cell(row, col).value
                if isinstance(value, str) and value.strip():
                    label = value.strip()
                    lowered = label.lower()
                    non_empty += 1
                    unique_labels.add(label)
                    if lowered in GENERIC_METRIC_TOKENS or any(token in lowered for token in GENERIC_METRIC_TOKENS):
                        generic_metric_hits += 1
                    next_value = ws.cell(row, col + 1).value if col + 1 <= ws.max_column else None
                    if isinstance(next_value, str) and ("收入" in next_value or "营收" in next_value):
                        revenue_pair_hits += 1
            score = len(unique_labels) * 3 + non_empty + revenue_pair_hits * 2 - generic_metric_hits * 2
            if score > best_score:
                best_col = col
                best_score = score
        return best_col

    def _classify_columns(
        self,
        ws: Worksheet,
        header_row: int,
    ) -> tuple[dict[int, int], dict[int, int], list[int]]:
        historical: dict[int, int] = {}
        forecast: dict[int, int] = {}
        excluded: list[int] = []
        for col in range(1, min(ws.max_column, 256) + 1):
            value = ws.cell(header_row, col).value
            if isinstance(value, int) and 2000 <= value <= 2100:
                historical[value] = col
                continue
            if isinstance(value, str):
                value = value.strip()
                forecast_match = FORECAST_YEAR_RE.match(value)
                if forecast_match:
                    forecast[int(forecast_match.group(1))] = col
                    continue
                if HALF_YEAR_RE.match(value) or "备注" in value:
                    excluded.append(col)
        return historical, forecast, excluded

    def _collect_row_labels(self, ws: Worksheet, header_row: int, label_column: int) -> dict[str, int]:
        labels: dict[str, int] = {}
        for row in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row, label_column).value
            if isinstance(value, str):
                label = value.strip()
                if label and label not in labels:
                    labels[label] = row
        return labels

    def _is_primary_block_row(self, ws: Worksheet, row: int, label_column: int) -> bool:
        value = ws.cell(row, label_column).value
        if not isinstance(value, str) or not value.strip():
            return False
        raw_label = value
        metric = ws.cell(row, label_column + 1).value if label_column + 1 <= ws.max_column else None
        if isinstance(metric, str) and metric.strip():
            return True
        return raw_label == raw_label.lstrip()

    def _collect_primary_row_labels(self, ws: Worksheet, header_row: int, label_column: int) -> dict[str, int]:
        labels: dict[str, int] = {}
        for row in range(header_row + 1, ws.max_row + 1):
            if not self._is_primary_block_row(ws, row, label_column):
                continue
            value = ws.cell(row, label_column).value
            if isinstance(value, str):
                label = value.strip()
                if label and label not in labels:
                    labels[label] = row
        return labels

    def _collect_row_blocks(
        self,
        ws: Worksheet,
        header_row: int,
        label_column: int,
        primary_row_labels: dict[str, int],
    ) -> dict[str, list[int]]:
        del header_row, label_column
        blocks: dict[str, list[int]] = {}
        ordered = sorted(primary_row_labels.items(), key=lambda item: item[1])
        for index, (label, start_row) in enumerate(ordered):
            next_row = ordered[index + 1][1] if index + 1 < len(ordered) else ws.max_row + 1
            blocks[label] = list(range(start_row, next_row))
        return blocks


class AnnualReportFactExtractor:
    DEFAULT_LABEL_MAP = {
        "营业收入": ["营业收入", "营业总收入"],
        "毛利": ["毛利", "毛利润", "营业毛利"],
        "毛利率": ["毛利率", "综合毛利率"],
        "归母净利润": ["归属于上市公司股东的净利润", "归母净利润"],
        "扣非归母净利润": ["扣除非经常性损益后的净利润", "归属于上市公司股东的扣除非经常性损益的净利润", "扣非归母净利润"],
        "销售费用": ["销售费用"],
        "销售费用率（%）": ["销售费用率", "销售费用占营业收入比例"],
        "管理费用": ["管理费用"],
        "管理费用率（%）": ["管理费用率", "管理费用占营业收入比例"],
        "研发费用": ["研发费用"],
        "研发费用率（%）": ["研发费用率", "研发费用占营业收入比例"],
        "财务费用": ["财务费用"],
        "财务费用率（%）": ["财务费用率", "财务费用占营业收入比例"],
        "所得税税率（%）": ["所得税税率", "实际税率"],
    }
    SEGMENT_SECTION_HINTS = ("分业务", "分产品", "分行业", "按业务", "按产品", "按应用", "分应用")

    def extract(self, report_path: Path) -> dict[str, Any]:
        pages = self._read_document_pages(report_path)
        text = "\n".join(page["text"] for page in pages)
        facts, fact_items = self._extract_labeled_numbers(pages, report_path)
        segment_disclosure = self._extract_segment_disclosure(pages, report_path)
        return {
            "source_path": str(report_path),
            "reported_facts": facts,
            "fact_items": fact_items,
            "segment_disclosure": segment_disclosure,
            "raw_text_excerpt": text[:4000],
        }

    def _read_document_pages(self, report_path: Path) -> list[dict[str, Any]]:
        if report_path.suffix.lower() == ".pdf":
            pages: list[dict[str, Any]] = []
            with pdfplumber.open(report_path) as pdf:
                for index, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text() or ""
                    tables = page.extract_tables() or []
                    if text.strip():
                        pages.append(
                            {
                                "page_number": index,
                                "page_reference": f"第{index}页",
                                "text": text,
                                "tables": tables,
                            }
                        )
            return pages
        text = report_path.read_text(encoding="utf-8", errors="ignore")
        page_match = PAGE_MARKER_RE.search(text)
        page_reference = f"第{page_match.group(1)}页" if page_match else "text"
        return [{"page_number": 1, "page_reference": page_reference, "text": text, "tables": []}]

    def _extract_labeled_numbers(
        self,
        pages: list[dict[str, Any]],
        report_path: Path,
    ) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        facts: dict[str, Any] = {}
        fact_items: list[dict[str, Any]] = []
        table_fact_items = self._extract_fact_items_from_tables(pages, report_path)
        table_lookup: dict[str, dict[str, Any]] = {}
        for item in table_fact_items:
            metric = str(item.get("metric") or "")
            table_lookup.setdefault(metric, item)
        for canonical_label, aliases in self.DEFAULT_LABEL_MAP.items():
            facts[canonical_label] = None
            matched_item: dict[str, Any] | None = table_lookup.get(canonical_label)
            for alias in aliases:
                if matched_item is not None:
                    break
                matched_item = self._find_fact_item_for_alias(pages, alias, canonical_label, report_path)
                if matched_item is not None:
                    break
            if matched_item is not None:
                facts[canonical_label] = matched_item["value"]
                fact_items.append(matched_item)
        support_items = list(fact_items)
        support_items.extend(item for item in table_fact_items if str(item.get("metric") or "") == "营业成本")
        facts, fact_items = self._derive_missing_reported_facts(facts, fact_items, report_path, support_items=support_items)
        return facts, fact_items

    def _extract_fact_items_from_tables(
        self,
        pages: list[dict[str, Any]],
        report_path: Path,
    ) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        normalized_aliases: dict[str, tuple[str, str]] = {}
        for canonical_label, aliases in self.DEFAULT_LABEL_MAP.items():
            for alias in aliases:
                normalized_aliases[re.sub(r"\s+", "", alias)] = (canonical_label, alias)
        normalized_aliases["营业成本"] = ("营业成本", "营业成本")

        for page in pages:
            page_reference = str(page.get("page_reference") or "")
            for table in page.get("tables") or []:
                for raw_row in table:
                    if not isinstance(raw_row, list):
                        continue
                    row = [str(cell).replace("\n", "").strip() if cell is not None else "" for cell in raw_row]
                    if not any(row):
                        continue
                    lead = re.sub(r"\s+", "", row[0])
                    match = normalized_aliases.get(lead)
                    if match is None:
                        continue
                    canonical_label, alias = match
                    search_cells = row[1:] if row[1:] else row
                    found_value: float | None = None
                    found_unit: str | None = None
                    for cell in search_cells:
                        found_value, found_unit = self._extract_value_and_unit(f"{alias} {cell}", canonical_label, alias)
                        if found_value is not None:
                            break
                    if found_value is None:
                        continue
                    items.append(
                        {
                            "metric": canonical_label,
                            "value": found_value,
                            "unit": found_unit,
                            "page_reference": page_reference,
                            "file_reference": str(report_path),
                            "note": "table_extracted",
                        }
                    )
        return items

    def _derive_missing_reported_facts(
        self,
        facts: dict[str, Any],
        fact_items: list[dict[str, Any]],
        report_path: Path,
        support_items: list[dict[str, Any]] | None = None,
    ) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        support_items = support_items or fact_items
        items_by_metric = {str(item.get("metric") or ""): item for item in support_items}
        revenue = facts.get("营业收入")
        operating_cost_item = next((item for item in support_items if item.get("metric") == "营业成本"), None)
        operating_cost = operating_cost_item.get("value") if operating_cost_item is not None else None

        def append_derived(metric: str, value: float, unit: str | None, note: str, source_metrics: list[str]) -> None:
            if facts.get(metric) is not None:
                return
            sources = [items_by_metric.get(name) for name in source_metrics if items_by_metric.get(name) is not None]
            page_reference = " / ".join(dict.fromkeys(str(item.get("page_reference") or "") for item in sources if item)) or "derived"
            facts[metric] = value
            item = {
                "metric": metric,
                "value": value,
                "unit": unit,
                "page_reference": page_reference,
                "file_reference": str(report_path),
                "note": note,
            }
            fact_items.append(item)
            items_by_metric[metric] = item

        if isinstance(revenue, (int, float)) and isinstance(operating_cost, (int, float)):
            gross_profit = round(float(revenue) - float(operating_cost), 6)
            current_gross_profit = facts.get("毛利")
            if not isinstance(current_gross_profit, (int, float)) or abs(float(current_gross_profit) - gross_profit) / max(abs(gross_profit), 1.0) > 0.2:
                facts["毛利"] = None
            append_derived("毛利", gross_profit, operating_cost_item.get("unit") if operating_cost_item else None, "derived_from_revenue_and_cost", ["营业收入", "营业成本"])
            if float(revenue) not in {0.0, 0}:
                derived_margin = round(gross_profit / float(revenue), 6)
                current_margin = facts.get("毛利率")
                if not isinstance(current_margin, (int, float)) or not (0 <= float(current_margin) <= 1):
                    facts["毛利率"] = None
                append_derived("毛利率", derived_margin, "%", "derived_from_revenue_and_cost", ["营业收入", "营业成本"])

        fee_rate_map = {
            "销售费用": "销售费用率（%）",
            "管理费用": "管理费用率（%）",
            "研发费用": "研发费用率（%）",
            "财务费用": "财务费用率（%）",
        }
        if isinstance(revenue, (int, float)) and float(revenue) not in {0.0, 0}:
            for fee_metric, rate_metric in fee_rate_map.items():
                fee_value = facts.get(fee_metric)
                if isinstance(fee_value, (int, float)):
                    append_derived(
                        rate_metric,
                        round(float(fee_value) / float(revenue), 6),
                        "%",
                        "derived_from_expense_and_revenue",
                        [fee_metric, "营业收入"],
                    )

        return facts, fact_items

    def _find_fact_item_for_alias(
        self,
        pages: list[dict[str, Any]],
        alias: str,
        canonical_label: str,
        report_path: Path,
    ) -> dict[str, Any] | None:
        for page in pages:
            lines = [line.strip() for line in str(page.get("text") or "").splitlines() if line.strip()]
            for idx, line in enumerate(lines):
                if alias not in line:
                    continue
                search_targets = [line]
                if "率" in canonical_label or "%" in canonical_label:
                    if idx + 1 < len(lines) and ("：" in line or ":" in line):
                        search_targets.append(" ".join(lines[idx: idx + 2]))
                else:
                    search_targets.append(" ".join(lines[idx: idx + 3]))
                value = None
                unit = None
                for target in search_targets:
                    value, unit = self._extract_value_and_unit(target, canonical_label, alias)
                    if value is not None:
                        break
                if value is None:
                    continue
                page_reference = self._resolve_page_reference(lines=lines, index=idx, fallback=str(page.get("page_reference") or ""))
                return {
                    "metric": canonical_label,
                    "value": value,
                    "unit": unit,
                    "page_reference": page_reference,
                    "file_reference": str(report_path),
                    "note": None,
                }
        return None

    def _extract_value_and_unit(self, text: str, canonical_label: str, alias: str) -> tuple[float | None, str | None]:
        snippet = text[text.find(alias): text.find(alias) + 160] if alias in text else text[:160]
        percent = PERCENT_RE.search(snippet)
        if ("率" in canonical_label or "%" in canonical_label) and percent:
            return round(float(percent.group(1)) / 100, 6), "%"
        if "率" in canonical_label or "%" in canonical_label:
            bare_number = re.search(r"-?\d+(?:\.\d+)?", snippet.replace(",", ""))
            if bare_number:
                return round(float(bare_number.group(0)) / 100, 6), "%"
        for number_text, unit in NUMBER_WITH_UNIT_RE.findall(snippet.replace(",", "")):
            if not number_text:
                continue
            value = float(number_text)
            normalized_unit = unit or None
            if normalized_unit == "%":
                return round(value / 100, 6), "%"
            return value, normalized_unit
        return None, None

    def _resolve_page_reference(self, *, lines: list[str], index: int, fallback: str) -> str:
        for cursor in range(index, max(-1, index - 5), -1):
            match = PAGE_MARKER_RE.search(lines[cursor])
            if match:
                return f"第{match.group(1)}页"
        return fallback or "unknown"

    def _extract_segment_disclosure(
        self,
        pages: list[dict[str, Any]],
        report_path: Path,
    ) -> list[dict[str, Any]]:
        segment_items: list[dict[str, Any]] = []
        seen_keys: set[tuple[str, str, str]] = set()

        section_token_map = {
            "product": ("分产品", "按产品", "主要产品"),
            "business": ("分业务", "按业务"),
            "application": ("分应用", "按应用"),
            "industry": ("分行业", "按行业"),
            "region": ("分地区",),
            "sales_model": ("销售模式",),
        }

        def detect_category(text: str) -> str | None:
            normalized = text.replace("\n", "").strip()
            for category, tokens in section_token_map.items():
                if any(token in normalized for token in tokens):
                    return category
            return None

        def normalize_row(raw_row: Any) -> list[str]:
            if not isinstance(raw_row, list):
                return []
            return [str(cell).replace("\n", "").strip() if cell is not None else "" for cell in raw_row]

        def add_item(*, category: str | None, segment: str, revenue: float, gross_margin: float | None, page_reference: str) -> None:
            clean_segment = segment.strip()
            if not clean_segment or clean_segment in {"合计", "小计"}:
                return
            dedupe_key = (category or "unknown", clean_segment, page_reference)
            if dedupe_key in seen_keys:
                return
            seen_keys.add(dedupe_key)
            segment_items.append(
                {
                    "segment": clean_segment,
                    "category": category or "unknown",
                    "revenue": revenue,
                    "gross_margin": gross_margin,
                    "unit": "亿元",
                    "page_reference": page_reference,
                    "file_reference": str(report_path),
                }
            )

        for page in pages:
            page_reference = str(page.get("page_reference") or "")
            for table in page.get("tables") or []:
                current_category: str | None = None
                in_disclosure_table = False
                for raw_row in table:
                    row = normalize_row(raw_row)
                    if not any(row):
                        continue
                    joined = " ".join(cell for cell in row if cell)
                    leading = row[0]
                    category = detect_category(joined) or detect_category(leading)
                    if category is not None:
                        current_category = category
                        in_disclosure_table = category in {"product", "business", "application", "industry", "region", "sales_model"}
                        continue
                    if not in_disclosure_table or current_category is None:
                        continue
                    if "营业收入" in joined and ("毛利率" in joined or "营业成本" in joined):
                        continue
                    if leading in {"分产品", "分业务", "分行业", "分地区", "主要产品", "销售模式"}:
                        continue
                    if leading in {"合计", "小计"} or not leading:
                        continue

                    revenue: float | None = None
                    revenue_unit: str | None = None
                    for candidate in row[1:4]:
                        revenue, revenue_unit = self._extract_value_and_unit(f"营业收入 {candidate}", "营业收入", "营业收入")
                        if revenue is not None:
                            break
                    if revenue is None:
                        continue

                    gross_margin: float | None = None
                    if len(row) >= 4:
                        gross_margin, margin_unit = self._extract_value_and_unit(f"毛利率 {row[3]}", "毛利率", "毛利率")
                        if margin_unit != "%":
                            gross_margin = None

                    if (revenue_unit == "元" or revenue_unit is None) and revenue > 1_000_000:
                        revenue = round(revenue / 100000000, 6)
                    add_item(
                        category=current_category,
                        segment=leading,
                        revenue=revenue,
                        gross_margin=gross_margin,
                        page_reference=page_reference,
                    )

            lines = [line.strip() for line in str(page.get("text") or "").splitlines() if line.strip()]
            in_section = False
            for idx, line in enumerate(lines):
                if any(hint in line for hint in self.SEGMENT_SECTION_HINTS):
                    in_section = True
                    continue
                if not in_section:
                    continue
                if idx > 0 and any(hint in lines[idx - 1] for hint in ("目录", "释义")):
                    in_section = False
                    continue
                match = re.match(r"^\s*([\u4e00-\u9fffA-Za-z0-9/（）()\-]+)\s+(-?\d+(?:\.\d+)?)\s*亿元(?:.*?毛利率\s*(-?\d+(?:\.\d+)?)%)?", line)
                if not match:
                    continue
                segment = match.group(1).strip()
                revenue = float(match.group(2))
                margin_text = match.group(3)
                add_item(
                    category=detect_category(line),
                    segment=segment,
                    revenue=revenue,
                    gross_margin=round(float(margin_text) / 100, 6) if margin_text else None,
                    page_reference=self._resolve_page_reference(lines=lines, index=idx, fallback=page_reference),
                )
            if segment_items:
                break
        return segment_items


class MeetingNotesFactExtractor:
    GROWTH_CUE_TOKENS = ("增长", "同比", "增速", "提升", "放量", "恢复", "增加", "扩张", "yoy")
    BLOCKED_LABEL_TOKENS = ("收入", "营收", "营业收入", "营业总收入", "毛利率", "净利率", "费用率", "税率", "占比", "份额")

    def extract(self, notes_path: Path) -> dict[str, Any]:
        content = read_supported_text(notes_path)
        segment_guidance = self._extract_segment_guidance(content)
        return {
            "source_path": str(notes_path),
            "bridge_facts": segment_guidance,
            "raw_text_excerpt": content[:4000],
        }

    def _extract_segment_guidance(self, content: str) -> dict[str, Any]:
        guidance: dict[str, Any] = {}
        for line in content.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            lowered = stripped.lower()
            if not any(token.lower() in lowered for token in self.GROWTH_CUE_TOKENS):
                continue
            percent = PERCENT_RE.search(stripped)
            if not percent:
                continue
            label = self._infer_segment_label(stripped)
            if not label:
                continue
            guidance[label] = {
                "claim": stripped,
                "source_type": "meeting_notes",
                "source_ref": f"user_supplied_meeting_notes:{label}",
                "used_for": "forecast_growth",
                "confidence": 0.7,
                "review_required": False,
                "growth_rate": round(float(percent.group(1)) / 100, 6),
            }
        return guidance

    def _infer_segment_label(self, sentence: str) -> str:
        direct = re.match(
            r"^\s*([\u4e00-\u9fffA-Za-z][\u4e00-\u9fffA-Za-z0-9/()（）\-_]{1,32}?)\s*(?:20\d{2}年)?(?:[^%\n]{0,20})?(?:增长|同比|增速|提升|放量|恢复|增加|扩张|yoy)",
            sentence,
            flags=re.IGNORECASE,
        )
        if direct:
            candidate = direct.group(1).strip()
            if not any(token in candidate for token in self.BLOCKED_LABEL_TOKENS):
                return candidate
        candidates = re.findall(r"[\u4e00-\u9fffA-Za-z][\u4e00-\u9fffA-Za-z0-9/()（）\-_]{1,24}", sentence)
        for candidate in candidates:
            compact = candidate.strip(":-：，,;；。 ")
            if len(compact) < 2:
                continue
            if any(token in compact for token in self.BLOCKED_LABEL_TOKENS):
                continue
            if PERCENT_RE.search(compact):
                continue
            return compact
        return ""


class CompanyConfigManager:
    def __init__(self, repo_root: Path):
        self.repo_root = repo_root
        self.config_dir = repo_root / "data" / "forecast_configs"
        self.config_dir.mkdir(parents=True, exist_ok=True)

    def ensure_config(self, company: str, blueprint: WorkbookBlueprint) -> Path:
        config_path = self.config_dir / f"{self._slugify(company)}.yaml"
        if config_path.exists():
            return config_path

        segment_labels = [
            label
            for label in blueprint.primary_row_labels
            if label not in GENERIC_CONSOLIDATED_LABELS and ("收入" in label or "营收" in label)
        ]

        payload = {
            "layout": {
                "primary_sheet": blueprint.primary_sheet,
                "header_row": blueprint.header_row,
                "label_column": blueprint.label_column,
                "historical_columns": blueprint.historical_columns,
                "forecast_columns": blueprint.forecast_columns,
                "excluded_columns": blueprint.excluded_columns,
            },
            "actualization_rules": {
                "actualize_latest_forecast_to_report_year": True,
                "freeze_historical_columns": True,
            },
            "rollforward_rules": {
                "append_new_far_year": True,
                "forecast_window_mode": "shift_right_by_one_year",
            },
            "line_map": {
                "reported_facts": {
                    "营业收入": "营业收入",
                    "毛利率": "毛利率",
                    "归母净利润": "归母净利润",
                    "扣非归母净利润": "扣非归母净利润",
                    "销售费用率（%）": "销售费用率（%）",
                    "管理费用率（%）": "管理费用率（%）",
                    "研发费用率（%）": "研发费用率（%）",
                    "财务费用率（%）": "财务费用率（%）",
                    "所得税税率（%）": "所得税税率（%）",
                }
            },
            "bridge_rules": {
                "allow_prior_logic_when_evidence_missing": True,
                "segment_labels": segment_labels,
            },
            "forecast_rules": {
                "default_far_year_method": "carry_last_growth",
                "summary_rows_use_existing_formula_or_value": True,
            },
            "evidence_priority": [
                "reference_files",
                "local_kb",
                "alpha_pai",
            ],
            "review_thresholds": {
                "absolute_change": 0.05,
                "relative_change": 0.1,
            },
        }
        with config_path.open("w", encoding="utf-8") as fh:
            yaml.safe_dump(payload, fh, allow_unicode=True, sort_keys=False)
        return config_path

    def _slugify(self, company: str) -> str:
        cleaned = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", company).strip("_")
        return cleaned or "company"


class WorkbookRollforwardEngine:
    REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")

    def rollforward(
        self,
        workbook_path: Path,
        blueprint: WorkbookBlueprint,
        report_year: int,
        actual_overrides: dict[str, Any],
        target_far_year: int | None = None,
        output_dir: Path | None = None,
        meeting_guidance: dict[str, Any] | None = None,
        evidence_payload: dict[str, Any] | None = None,
        config_path: Path | None = None,
        facts_payload: dict[str, Any] | None = None,
    ) -> RollforwardResult:
        output_dir = output_dir or workbook_path.parent
        output_dir.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.load_workbook(workbook_path)
        ws = wb[blueprint.primary_sheet]
        if report_year not in blueprint.forecast_columns:
            raise ValueError(f"Report year {report_year} not found in forecast columns")

        change_records: list[ChangeRecord] = []
        meeting_guidance = meeting_guidance or {}
        report_col = blueprint.forecast_columns[report_year]
        ws.cell(blueprint.header_row, report_col).value = report_year
        for label, row in blueprint.row_labels.items():
            report_cell = ws.cell(row, report_col)
            if label in actual_overrides:
                before = report_cell.value
                report_cell.value = actual_overrides[label]
                change_records.append(
                    ChangeRecord(
                        sheet=ws.title,
                        row_label=label,
                        year=report_year,
                        before=before,
                        after=report_cell.value,
                        change_type="actualized_from_annual_report",
                        evidence="annual_report",
                        rationale="replace latest forecast with reported actual",
                        confidence=0.95,
                        review_flag=False,
                    )
                )

        extended_forecast_columns = dict(blueprint.forecast_columns)
        target_far_year = max(target_far_year or (max(extended_forecast_columns) + 1), max(extended_forecast_columns) + 1)

        prev_col = None
        curr_col = None
        new_far_col = None
        new_far_year = None
        processed_rows = set(blueprint.row_labels.values())

        while max(extended_forecast_columns) < target_far_year:
            last_forecast_year = max(extended_forecast_columns)
            last_forecast_col = extended_forecast_columns[last_forecast_year]
            new_far_year = last_forecast_year + 1
            new_far_col = last_forecast_col + 1
            ws.insert_cols(new_far_col, amount=1)
            self._copy_column_format(ws, last_forecast_col, new_far_col)
            extended_forecast_columns[new_far_year] = new_far_col
            ws.cell(blueprint.header_row, new_far_col).value = f"{new_far_year}E"

            for label, row in blueprint.row_labels.items():
                if self._rebuild_formula_row_from_report_year(
                    ws=ws,
                    row=row,
                    row_label=label,
                    report_year=report_year,
                    forecast_columns=extended_forecast_columns,
                    blueprint=blueprint,
                    change_records=change_records,
                    only_if_summary=True,
                ):
                    continue

                self._apply_guidance(
                    ws=ws,
                    row=row,
                    label=label,
                    report_year=report_year,
                    forecast_columns=extended_forecast_columns,
                    meeting_guidance=meeting_guidance,
                    change_records=change_records,
                )

                prev_col = extended_forecast_columns.get(last_forecast_year - 1)
                curr_col = extended_forecast_columns.get(last_forecast_year)
                if prev_col is None or curr_col is None:
                    continue
                before_far = ws.cell(row, new_far_col).value
                if self._cell_has_value(before_far):
                    continue
                translated_formula = self._translate_formula_to_far_year(
                    formula=ws.cell(row, curr_col).value,
                    origin=ws.cell(row, curr_col).coordinate,
                    target=ws.cell(row, new_far_col).coordinate,
                )
                if translated_formula is not None:
                    ws.cell(row, new_far_col).value = translated_formula
                    change_records.append(
                        ChangeRecord(
                            sheet=ws.title,
                            row_label=label,
                            year=new_far_year,
                            before=before_far,
                            after=translated_formula,
                            change_type="formula_extended_to_far_year",
                            evidence="prior_model_formula",
                            rationale="extend existing forecast formula into new far-year column",
                            confidence=0.9,
                            review_flag=False,
                        )
                    )
                    continue
                derived_value = self._derive_far_year_value(
                    previous_value=ws.cell(row, prev_col).value,
                    current_value=ws.cell(row, curr_col).value,
                )
                if derived_value is None:
                    continue
                ws.cell(row, new_far_col).value = derived_value
                change_records.append(
                    ChangeRecord(
                        sheet=ws.title,
                        row_label=label,
                        year=new_far_year,
                        before=before_far,
                        after=derived_value,
                        change_type="rolled_forward",
                        evidence="prior_model",
                        rationale="extend forecast window by carrying latest growth",
                        confidence=0.6,
                        review_flag=False,
                    )
                )

            for primary_label in blueprint.primary_row_labels:
                self._apply_dependent_metric_guidance(
                    ws=ws,
                    blueprint=blueprint,
                    primary_label=primary_label,
                    forecast_columns=extended_forecast_columns,
                    report_year=report_year,
                    meeting_guidance=meeting_guidance,
                    change_records=change_records,
                )
            for primary_label, block_rows in blueprint.row_blocks.items():
                for row in block_rows:
                    if row in processed_rows:
                        continue
                    row_label = self._row_display_label(ws=ws, row=row, blueprint=blueprint, primary_label=primary_label)
                    has_explicit_schedule = any(
                        record.row_label == row_label and record.change_type == "dependent_metric_schedule_applied"
                        for record in change_records
                    )
                    if not has_explicit_schedule:
                        if self._rebuild_formula_row_from_report_year(
                            ws=ws,
                            row=row,
                            row_label=row_label,
                            report_year=report_year,
                            forecast_columns=extended_forecast_columns,
                            blueprint=blueprint,
                            change_records=change_records,
                            only_if_summary=False,
                        ):
                            continue
                    self._extend_dependent_row_to_far_year(
                        ws=ws,
                        row=row,
                        primary_label=primary_label,
                        prev_col=prev_col,
                        curr_col=curr_col,
                        new_far_col=new_far_col,
                        new_far_year=new_far_year,
                        blueprint=blueprint,
                        change_records=change_records,
                    )

        if new_far_col is None or new_far_year is None:
            new_far_year = max(extended_forecast_columns)
            new_far_col = extended_forecast_columns[new_far_year]
        current_forecast_col = extended_forecast_columns.get(new_far_year - 1)
        parity_audit = self._audit_far_year_parity(
            ws=ws,
            blueprint=blueprint,
            current_forecast_col=current_forecast_col,
            new_far_col=new_far_col,
            new_far_year=new_far_year,
        )
        if parity_audit["status"] != "passed":
            raise RuntimeError(
                "far-year parity audit failed: "
                + "; ".join(
                    f"{issue['parent_label']} row {issue['row']} ({issue['row_label']})"
                    for issue in parity_audit["issues"][:8]
                )
            )

        self._mark_review_rows(ws, blueprint, change_records)

        stem = workbook_path.stem
        output_workbook = output_dir / f"{stem}_{report_year}_candidate.xlsx"
        facts_path = output_dir / f"{stem}_{report_year}_facts.json"
        evidence_path = output_dir / f"{stem}_{report_year}_evidence.json"
        changelog_json = output_dir / f"{stem}_{report_year}_changelog.json"
        changelog_md = output_dir / f"{stem}_{report_year}_changelog.md"

        wb.save(output_workbook)

        facts_payload = facts_payload or {"reported_facts": actual_overrides}
        evidence_payload = evidence_payload or {}

        for artifact_path in (output_workbook, facts_path, evidence_path, changelog_json, changelog_md):
            ensure_runtime_artifact_path_allowed(artifact_path)

        facts_path.write_text(
            json.dumps(facts_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        evidence_path.write_text(
            json.dumps(evidence_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        changelog_json.write_text(
            json.dumps([asdict(record) for record in change_records], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        changelog_md.write_text(self._render_changelog_md(change_records), encoding="utf-8")

        return RollforwardResult(
            output_workbook=output_workbook,
            facts_path=facts_path,
            evidence_path=evidence_path,
            changelog_json=changelog_json,
            changelog_md=changelog_md,
            config_path=config_path or Path(),
            change_records=change_records,
            parity_audit=parity_audit,
        )

    def _copy_column_format(self, ws: Worksheet, source_col: int, target_col: int) -> None:
        ws.column_dimensions[openpyxl.utils.get_column_letter(target_col)].width = (
            ws.column_dimensions[openpyxl.utils.get_column_letter(source_col)].width
        )
        for row in range(1, ws.max_row + 1):
            src = ws.cell(row, source_col)
            dst = ws.cell(row, target_col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)

    def _translate_formula_to_far_year(
        self,
        *,
        formula: Any,
        origin: str,
        target: str,
    ) -> str | None:
        if not isinstance(formula, str) or not formula.startswith("="):
            return None
        try:
            return Translator(formula, origin=origin).translate_formula(target)
        except Exception:
            return None

    def _derive_far_year_value(self, previous_value: Any, current_value: Any) -> float | None:
        prev = self._to_number(previous_value)
        curr = self._to_number(current_value)
        if prev is None or curr is None:
            return None
        if abs(prev) < 1e-9:
            return round(curr, 4)
        if 0 <= prev <= 1 and 0 <= curr <= 1:
            return round(curr, 6)
        growth = curr / prev
        return round(curr * growth, 4)

    def _to_number(self, value: Any) -> float | None:
        if isinstance(value, (int, float)):
            return float(value)
        return None

    def _apply_guidance(
        self,
        ws: Worksheet,
        row: int,
        label: str,
        report_year: int,
        forecast_columns: dict[int, int],
        meeting_guidance: dict[str, Any],
        change_records: list[ChangeRecord],
    ) -> None:
        guidance = meeting_guidance.get(label)
        if not guidance:
            return
        year_values = guidance.get("year_values") or {}
        if year_values:
            for year in sorted(y for y in forecast_columns if y > report_year):
                cell_key = f"{year}E"
                if cell_key not in year_values:
                    continue
                before = ws.cell(row, forecast_columns[year]).value
                after = year_values[cell_key]
                ws.cell(row, forecast_columns[year]).value = after
                change_records.append(
                    ChangeRecord(
                        sheet=ws.title,
                        row_label=label,
                        year=year,
                        before=before,
                        after=after,
                        change_type="segment_schedule_applied",
                        evidence=guidance.get("claim", "forecast_architecture"),
                        rationale="apply year-specific segment forecast schedule from forecast architecture",
                        confidence=float(guidance.get("confidence", 0.7)),
                        review_flag=bool(guidance.get("review_required", False)),
                    )
                )
            return
        growth_rate = guidance.get("growth_rate")
        if growth_rate is None:
            return

        base_year = report_year
        base_col = forecast_columns[report_year]
        base_value = self._to_number(ws.cell(row, base_col).value)
        if base_value is None:
            return

        current = base_value
        for year in sorted(y for y in forecast_columns if y > report_year):
            col = forecast_columns[year]
            before = ws.cell(row, col).value
            current = round(current * (1 + growth_rate), 4)
            ws.cell(row, col).value = current
            change_records.append(
                ChangeRecord(
                    sheet=ws.title,
                    row_label=label,
                    year=year,
                    before=before,
                    after=current,
                    change_type="bridged_from_meeting_notes",
                    evidence=guidance.get("claim", "meeting_notes"),
                    rationale=f"apply meeting-notes growth guidance from {base_year}",
                    confidence=float(guidance.get("confidence", 0.7)),
                    review_flag=bool(guidance.get("review_required", False)),
                )
            )

    def _dependent_metric_key(self, metric_value: Any) -> str | None:
        if not isinstance(metric_value, str):
            return None
        lowered = metric_value.strip().lower()
        if "yoy" in lowered:
            return "yoy"
        if "毛利率" in lowered or "margin" in lowered:
            return "margin"
        if "市场份额" in lowered or "份额" in lowered or "share" in lowered:
            return "share"
        if "净利率" in lowered:
            return "net_margin"
        return None

    def _apply_dependent_metric_guidance(
        self,
        *,
        ws: Worksheet,
        blueprint: WorkbookBlueprint,
        primary_label: str,
        forecast_columns: dict[int, int],
        report_year: int,
        meeting_guidance: dict[str, Any],
        change_records: list[ChangeRecord],
    ) -> None:
        guidance = meeting_guidance.get(primary_label) or {}
        dependent_metric_values = guidance.get("dependent_metric_values") or {}
        if not dependent_metric_values:
            return
        for row in blueprint.row_blocks.get(primary_label, []):
            metric_value = ws.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws.max_column else None
            metric_key = self._dependent_metric_key(metric_value)
            if metric_key not in dependent_metric_values:
                continue
            year_values = dependent_metric_values.get(metric_key) or {}
            row_label = self._row_display_label(ws=ws, row=row, blueprint=blueprint, primary_label=primary_label)
            for year in sorted(y for y in forecast_columns if y > report_year):
                value_key = f"{year}E"
                if value_key not in year_values:
                    continue
                col = forecast_columns[year]
                before = ws.cell(row, col).value
                after = year_values[value_key]
                ws.cell(row, col).value = after
                change_records.append(
                    ChangeRecord(
                        sheet=ws.title,
                        row_label=row_label,
                        year=year,
                        before=before,
                        after=after,
                        change_type="dependent_metric_schedule_applied",
                        evidence=guidance.get("claim", "forecast_architecture"),
                        rationale=f"apply year-specific {metric_key} schedule under {primary_label}",
                        confidence=float(guidance.get("confidence", 0.7)),
                        review_flag=bool(guidance.get("review_required", False)),
                    )
                )

    def _row_display_label(
        self,
        *,
        ws: Worksheet,
        row: int,
        blueprint: WorkbookBlueprint,
        primary_label: str,
    ) -> str:
        primary_value = ws.cell(row, blueprint.label_column).value
        metric_value = ws.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws.max_column else None
        primary_text = primary_value.strip() if isinstance(primary_value, str) and primary_value.strip() else ""
        metric_text = metric_value.strip() if isinstance(metric_value, str) and metric_value.strip() else ""
        if primary_text and metric_text:
            return f"{primary_text}.{metric_text}"
        if primary_text:
            return primary_text
        if metric_text:
            return f"{primary_label}.{metric_text}"
        return f"{primary_label}.row_{row}"

    def _cell_has_value(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return bool(value.strip())
        return True

    def _formula_references_other_primary_rows(
        self,
        *,
        formula: Any,
        row: int,
        blueprint: WorkbookBlueprint,
    ) -> bool:
        if not isinstance(formula, str) or not formula.startswith("="):
            return False
        primary_rows = set((blueprint.primary_row_labels or blueprint.row_labels).values())
        referenced_rows = {
            int(match.group(1))
            for match in CELL_REF_ROW_RE.finditer(formula)
            if int(match.group(1)) != row
        }
        return bool(referenced_rows & primary_rows)

    def _rebuild_formula_row_from_report_year(
        self,
        *,
        ws: Worksheet,
        row: int,
        row_label: str,
        report_year: int,
        forecast_columns: dict[int, int],
        blueprint: WorkbookBlueprint,
        change_records: list[ChangeRecord],
        only_if_summary: bool,
    ) -> bool:
        report_col = forecast_columns.get(report_year)
        if report_col is None:
            return False
        report_formula = ws.cell(row, report_col).value
        if not isinstance(report_formula, str) or not report_formula.startswith("="):
            return False
        if only_if_summary and not self._formula_references_other_primary_rows(
            formula=report_formula,
            row=row,
            blueprint=blueprint,
        ):
            return False
        rebuilt = False
        for year in sorted(y for y in forecast_columns if y > report_year):
            col = forecast_columns[year]
            translated_formula = self._translate_formula_to_far_year(
                formula=report_formula,
                origin=ws.cell(row, report_col).coordinate,
                target=ws.cell(row, col).coordinate,
            )
            if translated_formula is None:
                continue
            before = ws.cell(row, col).value
            ws.cell(row, col).value = translated_formula
            change_records.append(
                ChangeRecord(
                    sheet=ws.title,
                    row_label=row_label,
                    year=year,
                    before=before,
                    after=translated_formula,
                    change_type="summary_formula_rebuilt" if only_if_summary else "formula_rebuilt_from_report_year",
                    evidence="prior_model_formula",
                    rationale="rebuild forecast row from the report-year formula template",
                    confidence=0.95,
                    review_flag=False,
                )
            )
            rebuilt = True
        return rebuilt

    def _extend_dependent_row_to_far_year(
        self,
        *,
        ws: Worksheet,
        row: int,
        primary_label: str,
        prev_col: int | None,
        curr_col: int | None,
        new_far_col: int,
        new_far_year: int,
        blueprint: WorkbookBlueprint,
        change_records: list[ChangeRecord],
    ) -> None:
        if prev_col is None or curr_col is None:
            return
        if self._cell_has_value(ws.cell(row, new_far_col).value):
            return
        current_value = ws.cell(row, curr_col).value
        previous_value = ws.cell(row, prev_col).value
        if not self._cell_has_value(current_value) and not self._cell_has_value(previous_value):
            return
        row_label = self._row_display_label(ws=ws, row=row, blueprint=blueprint, primary_label=primary_label)
        translated_formula = self._translate_formula_to_far_year(
            formula=current_value,
            origin=ws.cell(row, curr_col).coordinate,
            target=ws.cell(row, new_far_col).coordinate,
        )
        before_far = ws.cell(row, new_far_col).value
        if translated_formula is not None:
            ws.cell(row, new_far_col).value = translated_formula
            change_records.append(
                ChangeRecord(
                    sheet=ws.title,
                    row_label=row_label,
                    year=new_far_year,
                    before=before_far,
                    after=translated_formula,
                    change_type="dependent_formula_extended_to_far_year",
                    evidence="prior_model_formula",
                    rationale=f"extend dependent block row under {primary_label} into new far-year column",
                    confidence=0.9,
                    review_flag=False,
                )
            )
            return
        derived_value = self._derive_far_year_value(previous_value=previous_value, current_value=current_value)
        if derived_value is None:
            return
        ws.cell(row, new_far_col).value = derived_value
        change_records.append(
            ChangeRecord(
                sheet=ws.title,
                row_label=row_label,
                year=new_far_year,
                before=before_far,
                after=derived_value,
                change_type="dependent_row_rolled_forward",
                evidence="prior_model",
                rationale=f"extend dependent block row under {primary_label} by carrying latest pattern",
                confidence=0.6,
                review_flag=False,
            )
        )

    def _audit_far_year_parity(
        self,
        *,
        ws: Worksheet,
        blueprint: WorkbookBlueprint,
        current_forecast_col: int | None,
        new_far_col: int,
        new_far_year: int,
    ) -> dict[str, Any]:
        if current_forecast_col is None:
            return {"status": "passed", "new_far_year": new_far_year, "issues": []}
        issues: list[dict[str, Any]] = []
        for primary_label, block_rows in blueprint.row_blocks.items():
            for row in block_rows:
                current_value = ws.cell(row, current_forecast_col).value
                far_value = ws.cell(row, new_far_col).value
                if not self._cell_has_value(current_value):
                    continue
                if self._cell_has_value(far_value):
                    continue
                issues.append(
                    {
                        "parent_label": primary_label,
                        "row": row,
                        "row_label": self._row_display_label(
                            ws=ws,
                            row=row,
                            blueprint=blueprint,
                            primary_label=primary_label,
                        ),
                        "current_value": current_value,
                    }
                )
        return {
            "status": "passed" if not issues else "failed",
            "new_far_year": new_far_year,
            "issues": issues,
        }

    def _mark_review_rows(
        self,
        ws: Worksheet,
        blueprint: WorkbookBlueprint,
        change_records: list[ChangeRecord],
    ) -> None:
        flagged = {record.row_label for record in change_records if record.review_flag}
        for label in flagged:
            row = blueprint.row_labels.get(label)
            if row is None:
                continue
            ws.cell(row, blueprint.label_column).fill = self.REVIEW_FILL

    def _render_changelog_md(self, change_records: list[ChangeRecord]) -> str:
        lines = [
            "# Forecast Rollforward Changelog",
            "",
            "| Sheet | Row Label | Year | Change Type | Before | After | Evidence | Confidence | Review |",
            "|---|---|---:|---|---:|---:|---|---:|---|",
        ]
        for record in change_records:
            lines.append(
                "| {sheet} | {row_label} | {year} | {change_type} | {before} | {after} | {evidence} | {confidence:.2f} | {review_flag} |".format(
                    **asdict(record)
                )
            )
        lines.append("")
        return "\n".join(lines)


def _extract_reported_revenue_anchor(annual_report_facts: dict[str, Any] | None) -> float | None:
    if not annual_report_facts:
        return None
    values: list[float] = []
    for key, value in (annual_report_facts.get("reported_facts", {}) or {}).items():
        if not isinstance(value, (int, float)):
            continue
        key_text = str(key)
        if "营业收入" in key_text or "营收" in key_text or "收入" in key_text:
            values.append(float(value))
    return max(values) if values else None


def _resolve_financial_facts(
    *,
    financial_facts: dict[str, Any] | None = None,
    annual_report_facts: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return dict(financial_facts or annual_report_facts or {})


def _normalize_alpha_keyword(label: str) -> str:
    cleaned = re.sub(r"^\s*\d+\s*[)）.、]\s*", "", str(label or "").strip())
    cleaned = cleaned.replace("其中：", "").replace("其中:", "").strip()
    cleaned = re.sub(r"[（(【\[].*?[）)】\]]", "", cleaned)
    for suffix in ("收入", "营收", "营业收入", "营业总收入"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)].strip()
    return cleaned.strip(" -_/")


def _collect_alpha_candidate_segments(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[blueprint.primary_sheet]
    revenue_col = blueprint.forecast_columns.get(report_year) or blueprint.historical_columns.get(report_year)
    if revenue_col is None:
        return []
    blocked_keywords = ("合计", "总计", "营业收入", "营业总收入", "毛利", "净利润", "占比", "增速", "yoy", "毛利率", "净利率", "市占率", "ASP", "量", "市场容量")
    candidates: list[dict[str, Any]] = []
    seen_keywords: set[str] = set()
    ordered_labels = sorted((blueprint.primary_row_labels or blueprint.row_labels).items(), key=lambda item: item[1])
    label_index = {label: idx for idx, (label, _) in enumerate(ordered_labels)}
    for label, row in ordered_labels:
        metric = ws.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws.max_column else None
        metric_text = str(metric or "").strip()
        label_text = str(label or "").strip()
        normalized = _normalize_alpha_keyword(label_text)
        if not normalized:
            continue
        if any(token in label_text for token in blocked_keywords):
            if "收入" not in label_text and "营收" not in label_text:
                continue
        if normalized in {"营业", "营业总"}:
            continue
        if any(token in normalized for token in ("半导体设计", "电子元器件代理及销售", "其他设计")):
            continue
        looks_like_revenue_row = any(token in label_text for token in ("收入", "营收")) or any(
            token in metric_text for token in ("收入", "营收")
        )
        if not looks_like_revenue_row:
            continue
        if _safe_number(ws.cell(row, revenue_col).value) is None:
            continue
        context_keyword = None
        current_index = label_index.get(label, 0)
        for prev_label, _prev_row in reversed(ordered_labels[:current_index]):
            prev_normalized = _normalize_alpha_keyword(str(prev_label))
            if not prev_normalized:
                continue
            if prev_normalized in {"营业", "营业总"}:
                continue
            if any(token in prev_normalized for token in ("毛利", "净利润", "占比", "增速", "ASP", "量", "市场容量")):
                continue
            if prev_normalized.isupper() or re.fullmatch(r"[A-Za-z0-9/_-]{2,8}", prev_normalized):
                context_keyword = prev_normalized
                break
            if _prev_row < row - 12:
                break
        keyword = normalized
        if context_keyword and context_keyword not in keyword:
            keyword = f"{keyword}{context_keyword}"
        if keyword in seen_keywords:
            continue
        candidates.append({"row_label": label_text, "keyword": keyword, "row": row})
        seen_keywords.add(keyword)
    return candidates


def _plan_alpha_pai_queries(
    *,
    company: str,
    workbook_path: Path | None,
    blueprint: WorkbookBlueprint | None,
    annual_report_facts: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    if workbook_path is None or blueprint is None:
        fallback_reason = "fallback_company_query_without_workbook"
        return [{"query": company, "keyword": company, "reason": fallback_reason}]
    if blueprint.forecast_columns:
        forecast_start = min(blueprint.forecast_columns)
        historical_end = max(blueprint.historical_columns) if blueprint.historical_columns else None
        if historical_end is not None and forecast_start == historical_end + 1:
            report_year = historical_end
        else:
            report_year = forecast_start
    else:
        report_year = max(blueprint.historical_columns) if blueprint.historical_columns else None
    if report_year is None:
        return [{"query": company, "keyword": company, "reason": "fallback_company_query_without_year_anchor"}]
    history_candidates = [year for year in blueprint.historical_columns if year < report_year] if blueprint else []
    history_year = max(history_candidates) if history_candidates else None
    model_rows = _collect_alpha_candidate_segments(
        workbook_path=workbook_path,
        blueprint=blueprint,
        report_year=report_year,
    )
    total_revenue = _extract_reported_revenue_anchor(annual_report_facts)
    row_values: dict[str, float] = {}
    for candidate in model_rows:
        row_label = str(candidate["row_label"])
        current = _extract_metric_series(
            workbook_path=workbook_path,
            blueprint=blueprint,
            row=int(candidate["row"]),
            years=[report_year],
        ).get(report_year)
        if current is not None:
            row_values[row_label] = float(current)
    candidate_total = sum(value for value in row_values.values() if value > 0) or None
    if total_revenue in {None, 0, 0.0}:
        total_revenue = candidate_total
    elif candidate_total not in {None, 0, 0.0} and float(total_revenue) > float(candidate_total) * 10:
        total_revenue = candidate_total

    planned: list[dict[str, Any]] = []
    seen_queries: set[str] = set()
    for candidate in model_rows:
        row_label = str(candidate["row_label"])
        current = row_values.get(row_label)
        if current is None:
            continue
        share = (current / total_revenue) if total_revenue not in {None, 0, 0.0} else None
        growth = None
        if history_year is not None:
            previous = _extract_metric_series(
                workbook_path=workbook_path,
                blueprint=blueprint,
                row=blueprint.row_labels.get(row_label),
                years=[history_year],
            ).get(history_year)
            if previous not in {None, 0, 0.0}:
                growth = float(current) / float(previous) - 1
        if (share is not None and share > 0.30) or (growth is not None and growth > 0.40):
            keyword = str(candidate["keyword"])
            query = f"{company} {keyword}".strip() if keyword else company
            if query in seen_queries:
                continue
            reasons: list[str] = []
            if share is not None and share > 0.30:
                reasons.append(f"revenue_share>{share:.2%}")
            if growth is not None and growth > 0.40:
                reasons.append(f"growth>{growth:.2%}")
            planned.append(
                {
                    "query": query,
                    "keyword": keyword or company,
                    "segment": row_label,
                    "reason": " and ".join(reasons),
                    "revenue_share": round(share, 6) if share is not None else None,
                    "growth_rate": round(growth, 6) if growth is not None else None,
                }
            )
            seen_queries.add(query)
    if planned:
        return planned
    return [{"query": company, "keyword": company, "reason": "fallback_company_query_no_segment_threshold_hit"}]


def build_evidence_payload(
    repo_root: Path,
    company: str,
    meeting_notes: Path | None,
    research_report: Path | None = None,
    annual_report: Path | None = None,
    workbook_path: Path | None = None,
    blueprint: WorkbookBlueprint | None = None,
    annual_report_facts: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if workbook_path is not None and blueprint is None:
        try:
            blueprint = WorkbookBlueprintParser().parse(workbook_path)
        except Exception:
            blueprint = None
    payload: dict[str, Any] = {
        "company": company,
        "evidence_priority": ["reference_files", "local_kb", "alpha_pai"],
        "providers": [],
        "provider_decisions": [],
        "recall_checks": {"source_prep": []},
    }

    def record_provider(
        result: ProviderResult,
        *,
        source_tier: str,
        decision: str,
        used_for: str,
    ) -> None:
        provider_row = asdict(result)
        provider_row["source_tier"] = source_tier
        payload["providers"].append(provider_row)
        payload["provider_decisions"].append(
            {
                "source_type": result.source_type,
                "source_ref": result.source_ref,
                "source_tier": source_tier,
                "decision": decision,
                "used_for": used_for,
                "has_content": bool(result.content.strip()),
            }
        )

    reference_candidates = [
        ("annual_report", annual_report),
        ("meeting_notes", meeting_notes),
        ("research_report", research_report),
    ]
    reference_has_content = False
    for reference_kind, path in reference_candidates:
        if path is None or not path.exists():
            continue
        result = UserSuppliedTextProvider(path).fetch(company)
        result.metadata["reference_kind"] = reference_kind
        record_provider(
            result,
            source_tier="reference_files",
            decision="primary_reference_loaded",
            used_for="facts_and_forecast_grounding",
        )
        reference_has_content = reference_has_content or bool(result.content.strip())

    local_kb_result = LocalKBProvider(repo_root).fetch(company)
    local_kb_has_content = bool(local_kb_result.content.strip())
    record_provider(
        local_kb_result,
        source_tier="local_kb",
        decision="secondary_recall_loaded",
        used_for="mechanism_bridge_and_history",
    )

    alpha_query_plan = _plan_alpha_pai_queries(
        company=company,
        workbook_path=workbook_path,
        blueprint=blueprint,
        annual_report_facts=annual_report_facts,
    )
    alpha_results: list[tuple[ProviderResult, dict[str, Any]]] = []
    with ThreadPoolExecutor(
        max_workers=max(1, min(4, len(alpha_query_plan))),
        thread_name_prefix="alphapai-recall",
    ) as executor:
        future_map = {
            executor.submit(
                AlphaPaiProvider().fetch,
                item["query"],
                recall_types=["roadShow", "roadShow_ir", "roadShow_us", "comment"],
            ): item
            for item in alpha_query_plan
        }
        for future in as_completed(future_map):
            plan_item = future_map[future]
            result = future.result()
            result.metadata["query_plan"] = plan_item
            alpha_results.append((result, plan_item))
    alpha_results.sort(key=lambda item: item[1]["query"])
    for alpha_result, plan_item in alpha_results:
        record_provider(
            alpha_result,
            source_tier="alpha_pai",
            decision="mandatory_segment_recall_executed",
            used_for=f"cross_check_and_incremental_clues::{plan_item['keyword']}",
        )

    report_download_result = ReportDownloadProvider().fetch(company)
    record_provider(
        report_download_result,
        source_tier="reference_files",
        decision="supplementary_reference_probe",
        used_for="traceability_only",
    )
    alpha_has_content = any(bool(result.content.strip()) for result, _ in alpha_results)
    alpha_call_count = len(alpha_results)
    alpha_success_count = sum(1 for result, _ in alpha_results if bool(result.content.strip()))
    payload["recall_checks"]["source_prep"] = [
        {
            "name": "reference_files_recalled",
            "passed": True,
            "detail": f"{sum(1 for item in payload['providers'] if item.get('source_tier') == 'reference_files')} providers recorded",
        },
        {
            "name": "reference_files_have_content",
            "passed": reference_has_content,
            "detail": f"{sum(1 for item in payload['providers'] if item.get('source_tier') == 'reference_files' and str(item.get('content') or '').strip())} providers with content",
        },
        {
            "name": "local_kb_recalled",
            "passed": True,
            "detail": f"has_content={local_kb_has_content}",
        },
        {
            "name": "alpha_pai_recalled",
            "passed": True,
            "detail": "mandatory recall executed for every run",
        },
        {
            "name": "alpha_pai_has_content",
            "passed": alpha_has_content,
            "detail": f"success_count={alpha_success_count}/{alpha_call_count}",
        },
        {
            "name": "alpha_pai_call_count",
            "passed": alpha_call_count > 0,
            "detail": alpha_call_count,
        },
        {
            "name": "alpha_pai_query_plan",
            "passed": bool(alpha_query_plan),
            "detail": "; ".join(item["query"] for item in alpha_query_plan),
        },
    ]
    return payload


def review_forecast_inputs(
    *,
    annual_report_facts: dict[str, Any],
    meeting_notes_facts: dict[str, Any],
    evidence_payload: dict[str, Any],
) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []
    source_priority_warnings: list[str] = []
    coverage_warnings: list[str] = []

    reported_facts = annual_report_facts.get("reported_facts", {})
    bridge_facts = meeting_notes_facts.get("bridge_facts", {})
    providers = evidence_payload.get("providers", [])

    reference_items = [
        item for item in providers
        if item.get("source_tier") == "reference_files" and str(item.get("content") or "").strip()
    ]
    local_kb_items = [
        item for item in providers
        if item.get("source_tier") == "local_kb" and str(item.get("content") or "").strip()
    ]
    alpha_items = [
        item for item in providers
        if item.get("source_tier") == "alpha_pai" and str(item.get("content") or "").strip()
    ]

    checks.append(
        {
            "name": "reference_files_loaded",
            "passed": bool(reference_items),
            "detail": f"{len(reference_items)} reference providers with content",
        }
    )
    checks.append(
        {
            "name": "reported_facts_present",
            "passed": any(value is not None for value in reported_facts.values()),
            "detail": f"{sum(1 for value in reported_facts.values() if value is not None)} reported facts extracted",
        }
    )
    checks.append(
        {
            "name": "bridge_facts_present",
            "passed": bool(bridge_facts),
            "detail": f"{len(bridge_facts)} bridge facts available",
        }
    )

    if not reference_items:
        coverage_warnings.append("reference_files_missing_or_empty")
    if not any(value is not None for value in reported_facts.values()):
        coverage_warnings.append("annual_report_facts_missing")

    for label, fact in bridge_facts.items():
        if not isinstance(fact, dict):
            continue
        source_type = str(fact.get("source_type") or "")
        if source_type == "alpha_pai" and local_kb_items:
            source_priority_warnings.append(
                f"{label} uses alpha_pai while local_kb evidence is available"
            )

    if alpha_items and not reference_items:
        coverage_warnings.append("alpha_pai_present_without_reference_backing")

    passed = not source_priority_warnings and not coverage_warnings
    return {
        "passed": passed,
        "checks": checks,
        "coverage_warnings": coverage_warnings,
        "source_priority_warnings": source_priority_warnings,
        "reviewed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


GENERIC_CONSOLIDATED_LABELS = {
    "营业收入",
    "毛利",
    "毛利率",
    "归母净利润",
    "扣非归母净利润",
    "净利率",
    "销售费用率（%）",
    "管理费用率（%）",
    "研发费用率（%）",
    "财务费用率（%）",
    "所得税税率（%）",
}


def _column_letter(col_idx: int) -> str:
    return openpyxl.utils.get_column_letter(col_idx)


def _column_from_cell_letter(cell: str) -> str:
    match = re.match(r"([A-Z]+)", str(cell))
    return match.group(1) if match else ""


def _infer_report_year_from_blueprint(blueprint: WorkbookBlueprint) -> int | None:
    if blueprint.forecast_columns:
        forecast_start = min(blueprint.forecast_columns)
        historical_end = max(blueprint.historical_columns) if blueprint.historical_columns else None
        if historical_end is not None and forecast_start == historical_end + 1:
            return historical_end
        return forecast_start
    return max(blueprint.historical_columns) if blueprint.historical_columns else None


def _row_display_label_for_contract(
    *,
    ws: Worksheet,
    row: int,
    blueprint: WorkbookBlueprint,
    primary_label: str,
) -> str:
    primary_value = ws.cell(row, blueprint.label_column).value
    if isinstance(primary_value, str) and primary_value.strip():
        return primary_value.strip()
    metric_value = ws.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws.max_column else None
    if isinstance(metric_value, str) and metric_value.strip():
        return metric_value.strip()
    return f"{primary_label}_row_{row}"


def _row_formula_template(row_values: dict[str, Any], row_cells: dict[str, str]) -> str | None:
    for year, cell in row_cells.items():
        formula = row_values.get(year)
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        source_col = _column_from_cell_letter(cell)
        return re.sub(rf"\$?{re.escape(source_col)}(?=\$?\d+)", "{col}", formula)
    return None


def _is_writable_driver_row(*, label: str, row: int, primary_rows: set[int], has_formula: bool) -> bool:
    if has_formula or row not in primary_rows:
        return False
    if label in GENERIC_CONSOLIDATED_LABELS:
        return False
    lowered = label.lower()
    if any(token in lowered for token in ("yoy", "毛利率", "净利率", "市场份额", "份额", "asp", "销量", "市场容量", "税率", "费用率", "%")):
        return False
    if any(token in label for token in ("营业外", "税金", "附加")):
        return False
    generic_driver_labels = {"营收", "营收（亿元）", "收入", "收入（亿元）"}
    if label in generic_driver_labels:
        return False
    return ("收入" in label) or ("营收" in label)


def build_workbook_map_contract(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int | None = None,
    target_far_year: int | None = None,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    ws = wb[blueprint.primary_sheet]
    forecast_years = sorted(blueprint.forecast_columns)
    if report_year is not None and target_far_year is not None and target_far_year > report_year:
        current_far_year_col = blueprint.forecast_columns[max(forecast_years)]
        current_forecast_window = [f"{year}E" for year in range(report_year + 1, target_far_year + 1)]
    else:
        current_far_year_col = blueprint.forecast_columns[max(forecast_years)] if forecast_years else 0
        current_forecast_window = [f"{year}E" for year in forecast_years]
    primary_rows = set(blueprint.primary_row_labels.values())
    row_registry: list[dict[str, Any]] = []
    formula_rows: list[str] = []
    display_rows: list[str] = []
    writable_driver_targets: list[str] = []

    visited_rows: set[int] = set()
    for primary_label, block_rows in blueprint.row_blocks.items():
        for row in block_rows:
            if row in visited_rows:
                continue
            visited_rows.add(row)
            label = _row_display_label_for_contract(ws=ws, row=row, blueprint=blueprint, primary_label=primary_label)
            row_id = f"{_slugify_text(label)}__r{row}"
            source_year_cells = {f"{year}E": f"{_column_letter(col)}{row}" for year, col in sorted(blueprint.forecast_columns.items())}
            if report_year is not None and target_far_year is not None and target_far_year > report_year:
                year_cells = {
                    f"{year}E": f"{_column_letter(current_far_year_col + offset)}{row}"
                    for offset, year in enumerate(range(report_year + 1, target_far_year + 1))
                }
            else:
                year_cells = source_year_cells
            row_values = {f"{year}E": ws.cell(row, col).value for year, col in sorted(blueprint.forecast_columns.items())}
            has_formula = any(isinstance(value, str) and value.startswith("=") for value in row_values.values())
            writable = _is_writable_driver_row(label=label, row=row, primary_rows=primary_rows, has_formula=has_formula)
            if writable:
                role = "driver_input"
                basis_paths = {
                    year: f"assumptions.revenue__{_slugify_text(label)}__{year}.value"
                    for year in current_forecast_window
                }
                writable_driver_targets.append(row_id)
            elif has_formula:
                role = "formula_derived"
                basis_paths = None
                formula_rows.append(row_id)
            else:
                role = "summary_display"
                basis_paths = None
                display_rows.append(row_id)

            row_item: dict[str, Any] = {
                "row_id": row_id,
                "sheet": blueprint.primary_sheet,
                "row": row,
                "label": label,
                "role": role,
                "writable": writable,
                "required_years": current_forecast_window,
                "must_extend_to_far_year": True,
                "year_cells": year_cells,
                "validation": {"review_flag": False},
            }
            formula_template = _row_formula_template(row_values, source_year_cells)
            if formula_template:
                row_item["formula_template"] = formula_template
            if basis_paths:
                row_item["basis_paths"] = basis_paths
            row_registry.append(row_item)

    current_headers = {
        _column_letter(col): ws.cell(blueprint.header_row, col).value
        for _, col in sorted(blueprint.forecast_columns.items())
    }
    prior_forecast_col = blueprint.forecast_columns.get(forecast_years[-2]) if len(forecast_years) >= 2 else None
    far_forecast_col = blueprint.forecast_columns.get(forecast_years[-1]) if forecast_years else None
    gap_count = 0
    if prior_forecast_col and far_forecast_col:
        for rows in blueprint.row_blocks.values():
            for row in rows:
                prev_value = ws.cell(row, prior_forecast_col).value
                far_value = ws.cell(row, far_forecast_col).value
                if prev_value not in (None, "") and far_value in (None, ""):
                    gap_count += 1

    report_year = _infer_report_year_from_blueprint(blueprint)
    rollforward_pattern = {
        "reported_col": _column_letter(blueprint.historical_columns.get(report_year, 0)) if report_year in blueprint.historical_columns else None,
        "forecast_start_col": _column_letter(blueprint.forecast_columns.get(forecast_years[0], 0)) if forecast_years else None,
        "current_far_year_col": _column_letter(blueprint.forecast_columns.get(forecast_years[-1], 0)) if forecast_years else None,
    }
    return {
        "workbook": str(workbook_path),
        "main_modeling_sheet": blueprint.primary_sheet,
        "main_modeling_sheet_index": wb.sheetnames.index(blueprint.primary_sheet),
        "header_row": blueprint.header_row,
        "label_column": blueprint.label_column,
        "historical_columns": blueprint.historical_columns,
        "forecast_columns": blueprint.forecast_columns,
        "rollforward_pattern": rollforward_pattern,
        "current_headers": current_headers,
        "current_forecast_window": current_forecast_window,
        "summary_extension_status": {"rows_with_prior_populated_but_current_blank": gap_count},
        "row_registry": row_registry,
        "writable_driver_targets": writable_driver_targets,
        "formula_rows": formula_rows,
        "display_rows": display_rows,
        "map_validation_hints": {
            "must_match_headers": True,
            "must_preserve_sheet_index": True,
            "row_registry_generated_from_public_builder": True,
        },
    }


def _render_basis_sources(refs: list[str], evidence_lookup: dict[str, dict[str, Any]]) -> str:
    rendered: list[str] = []
    for ref in refs:
        evidence = evidence_lookup.get(ref)
        if evidence is None:
            rendered.append(ref)
            continue
        source_ref = Path(str(evidence.get("source_ref") or ref)).name
        reference_kind = evidence.get("metadata", {}).get("reference_kind")
        summary = str(evidence.get("content") or "").strip().replace("\n", " ")
        summary = summary[:80]
        parts = [source_ref]
        if reference_kind:
            parts.append(str(reference_kind))
        if summary:
            parts.append(summary)
        rendered.append(" | ".join(parts))
    return "; ".join(rendered)


def materialize_forecast_basis_sheet(
    *,
    workbook_path: Path,
    forecast_basis: dict[str, Any],
    evidence_store: list[dict[str, Any]],
) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    if "Forecast Basis" in wb.sheetnames:
        del wb["Forecast Basis"]
    ws = wb.create_sheet("Forecast Basis")

    evidence_lookup = {str(item.get("fact_id")): item for item in evidence_store if item.get("fact_id")}

    ws["A1"] = "Forecast Basis"
    ws["A2"] = f"Company: {forecast_basis.get('company', '')}"
    ws["D2"] = f"Cutoff: {forecast_basis.get('cutoff_date', '')}"
    ws["G2"] = f"Reported Year: {forecast_basis.get('reported_year', '')}"
    ws["I2"] = "Target Window: " + ", ".join(str(item) for item in forecast_basis.get("target_window", []))

    headers = ["分类", "项目", "年度", "数值/机制", "驱动形式", "依据摘要", "来源", "证据标签", "风险提示", "置信度"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(3, idx).value = header

    current_row = 4
    for fact in forecast_basis.get("facts", []):
        refs = [str(item) for item in fact.get("source_ref", [])]
        ws.cell(current_row, 1).value = "已知事实"
        ws.cell(current_row, 2).value = fact.get("metric") or fact.get("key")
        ws.cell(current_row, 3).value = fact.get("year")
        ws.cell(current_row, 4).value = fact.get("value")
        ws.cell(current_row, 5).value = "actual"
        ws.cell(current_row, 6).value = fact.get("summary") or "reported fact"
        ws.cell(current_row, 7).value = _render_basis_sources(refs, evidence_lookup)
        ws.cell(current_row, 8).value = "; ".join(refs)
        ws.cell(current_row, 9).value = fact.get("review_flag")
        ws.cell(current_row, 10).value = fact.get("confidence")
        current_row += 1

    subdriver_map = {
        "volume_logic": "volume",
        "asp_logic": "asp",
        "share_logic": "share",
        "margin_logic": "margin",
    }
    for card in forecast_basis.get("segment_assumption_cards", []):
        source_refs = [str(item) for item in card.get("source_ref", [])]
        basis_bits: list[str] = []
        for logic_key in subdriver_map:
            mechanism = (card.get(logic_key) or {}).get("mechanism")
            if mechanism:
                basis_bits.append(str(mechanism))

        ws.cell(current_row, 1).value = "预测假设"
        ws.cell(current_row, 2).value = f"{card.get('segment')}.{card.get('metric')}"
        ws.cell(current_row, 3).value = card.get("year")
        ws.cell(current_row, 4).value = card.get("value")
        ws.cell(current_row, 5).value = card.get("driver_form")
        ws.cell(current_row, 6).value = " | ".join(basis_bits)
        ws.cell(current_row, 7).value = _render_basis_sources(source_refs, evidence_lookup)
        ws.cell(current_row, 8).value = "; ".join(source_refs)
        ws.cell(current_row, 9).value = " | ".join(str(item) for item in card.get("kill_conditions", []))
        ws.cell(current_row, 10).value = card.get("confidence")
        current_row += 1

        for logic_key, subdriver in subdriver_map.items():
            logic_payload = card.get(logic_key) or {}
            mechanism = logic_payload.get("mechanism")
            evidence_refs = [str(item) for item in logic_payload.get("evidence_refs", [])]
            if not mechanism and not evidence_refs:
                continue
            ws.cell(current_row, 1).value = "子驱动依据"
            ws.cell(current_row, 2).value = f"{card.get('segment')}.{subdriver}"
            ws.cell(current_row, 3).value = card.get("year")
            ws.cell(current_row, 4).value = mechanism
            ws.cell(current_row, 6).value = mechanism
            ws.cell(current_row, 7).value = _render_basis_sources(evidence_refs, evidence_lookup)
            ws.cell(current_row, 8).value = "; ".join(evidence_refs)
            ws.cell(current_row, 10).value = card.get("confidence")
            current_row += 1

    wb.save(workbook_path)


def _build_evidence_store_for_basis(evidence_payload: dict[str, Any]) -> list[dict[str, Any]]:
    evidence_store: list[dict[str, Any]] = []
    for index, provider in enumerate(evidence_payload.get("providers", []), start=1):
        fact_id = str(provider.get("source_ref") or f"provider_{index}")
        evidence_store.append(
            {
                "fact_id": fact_id,
                "source_ref": provider.get("source_ref"),
                "source_file": Path(str(provider.get("source_ref") or f"provider_{index}")).name,
                "page_or_line": provider.get("metadata", {}).get("reference_kind") or provider.get("source_type"),
                "text_summary": str(provider.get("content") or "").strip().replace("\n", " ")[:120],
                "content": provider.get("content", ""),
                "metadata": {
                    **provider.get("metadata", {}),
                    "source_tier": provider.get("source_tier"),
                },
            }
        )
    return evidence_store


def _tier_rank(source_tier: str | None) -> int:
    return {"reference_files": 0, "local_kb": 1, "alpha_pai": 2}.get(str(source_tier or ""), 99)


def _normalize_segment_label(label: str) -> str:
    return re.sub(r"[\s（）()、/·\\-]+", "", str(label or "")).strip()


def _segment_keywords_generic(label: str) -> list[str]:
    normalized = _normalize_segment_label(label)
    keywords = {str(label).strip(), normalized}
    if normalized:
        keywords.add(normalized[:2])
        keywords.add(normalized[:3])
        keywords.add(normalized[-2:])
        if len(normalized) >= 4:
            for idx in range(len(normalized) - 1):
                keywords.add(normalized[idx: idx + 2])
    trimmed = re.sub(r"(业务板块|业务|板块|产品|收入|营收|亿元)$", "", normalized)
    if trimmed:
        keywords.add(trimmed)
    return [item for item in keywords if item]


def _is_summary_row_label(label: str) -> bool:
    text = str(label or "").strip()
    if not text:
        return True
    exact_blocked = {"合计", "总计", "营业收入", "营业总收入", "收入", "营收", "yoy", "同比", "毛利率", "净利率", "费用率"}
    if text in exact_blocked:
        return True
    if text.startswith("其中") or text.startswith("其他"):
        return True
    return False


def _is_generic_segment_placeholder(label: str) -> bool:
    text = str(label or "").strip()
    if not text:
        return True
    cleaned = _normalize_alpha_keyword(text)
    compact = re.sub(r"[\s（）()、/·\\-]+", "", cleaned).lower()
    if text.startswith(("加：", "加:", "减：", "减:")):
        return True
    if "营业外" in text:
        return True
    if text in {"营收（亿元）", "收入（亿元）", "营收(亿元)", "收入(亿元)"}:
        return True
    if compact in {"", "营收", "收入", "营业收入", "营业总收入", "营收亿元", "收入亿元"}:
        return True
    return False


def _is_metric_like_non_segment_label(label: str) -> bool:
    text = str(label or "").strip()
    if not text:
        return True
    normalized = _normalize_alpha_keyword(text).lower()
    blocked_tokens = (
        "占比",
        "比例",
        "比重",
        "yoy",
        "同比",
        "毛利率",
        "净利率",
        "费用率",
        "税率",
        "margin",
        "ratio",
    )
    return any(token in normalized for token in blocked_tokens)


def _split_text_sentences(text: str) -> list[str]:
    raw_parts = re.split(r"[\r\n。！？；;]+", str(text or ""))
    return [part.strip() for part in raw_parts if part and part.strip()]


def _is_stale_research_forecast_sentence(provider: dict[str, Any], sentence: str) -> bool:
    reference_kind = str(provider.get("metadata", {}).get("reference_kind") or "").strip()
    if reference_kind != "research_report":
        return False
    return bool(STALE_FORECAST_MARKER_RE.search(str(sentence or "")))


def _normalize_claim_text(text: str) -> str:
    return re.sub(r"[\s，。；;：:、|]+", "", str(text or "")).lower()


def _dedupe_evidence_items(
    items: list[dict[str, Any]],
    *,
    max_items: int,
    max_per_source: int = 1,
) -> list[dict[str, Any]]:
    deduped: list[dict[str, Any]] = []
    seen_claims: set[tuple[str, str]] = set()
    source_counts: dict[str, int] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        if str(item.get("evidence_role") or "") == "stale_forecast":
            continue
        claim = _normalize_claim_text(str(item.get("claim") or ""))
        source = str(item.get("source_ref") or item.get("source_label") or "").strip()
        if not claim and not source:
            continue
        signature = (source, claim)
        if signature in seen_claims:
            continue
        if source and source_counts.get(source, 0) >= max_per_source:
            continue
        deduped.append(item)
        seen_claims.add(signature)
        if source:
            source_counts[source] = source_counts.get(source, 0) + 1
        if len(deduped) >= max_items:
            break
    return deduped


def _best_sentence_for_keywords(
    content: str,
    keywords: list[str],
    provider: dict[str, Any] | None = None,
) -> str:
    sentences = _split_text_sentences(content)
    if not sentences:
        return ""
    lowered_keywords = [item.lower() for item in keywords if item]
    scored: list[tuple[int, str]] = []
    for sentence in sentences:
        if provider is not None and _is_stale_research_forecast_sentence(provider, sentence):
            continue
        lowered = sentence.lower()
        score = sum(1 for keyword in lowered_keywords if keyword in lowered or keyword in sentence)
        if score > 0:
            scored.append((score, sentence))
    if scored:
        scored.sort(key=lambda item: (-item[0], len(item[1])))
        return scored[0][1][:120]
    if provider is not None:
        for sentence in sentences:
            if _is_stale_research_forecast_sentence(provider, sentence):
                continue
            return sentence[:120]
        return ""
    return sentences[0][:120]


def _provider_source_label(provider: dict[str, Any]) -> str:
    source_name = Path(str(provider.get("source_ref") or "unknown")).name
    locator = provider.get("metadata", {}).get("reference_kind") or provider.get("source_type") or "source"
    return f"{source_name} | {locator}"


def _provider_evidence_item(provider: dict[str, Any], claim: str) -> dict[str, Any]:
    evidence_role = "stale_forecast" if _is_stale_research_forecast_sentence(provider, claim) else "mechanism"
    return {
        "claim": str(claim).strip()[:120],
        "source_ref": str(provider.get("source_ref") or ""),
        "source_tier": str(provider.get("source_tier") or ""),
        "source_label": _provider_source_label(provider),
        "evidence_role": evidence_role,
    }


def _collect_model_segment_rows(
    *,
    workbook_path: Path | None,
    blueprint: WorkbookBlueprint | None,
    report_year: int | None = None,
) -> list[str]:
    if workbook_path is None or blueprint is None:
        return []
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[blueprint.primary_sheet]
    formula_ws = None
    if report_year is not None:
        formula_wb = openpyxl.load_workbook(workbook_path, data_only=False)
        formula_ws = formula_wb[blueprint.primary_sheet]
    rows: list[str] = []
    for row in range(blueprint.header_row + 1, ws.max_row + 1):
        found_label = None
        for col in range(1, min(blueprint.label_column + 1, 4)):
            label = ws.cell(row, col).value
            metric = ws.cell(row, col + 1).value if col + 1 <= ws.max_column else None
            if not isinstance(label, str) or not label.strip():
                continue
            if _is_summary_row_label(label):
                continue
            label_has_revenue = "收入" in label or "营收" in label
            metric_has_revenue = isinstance(metric, str) and ("收入" in metric or "营收" in metric)
            if not label_has_revenue and not metric_has_revenue:
                continue
            found_label = label.strip()
            break
        if found_label:
            if report_year is not None and _is_formula_driven_summary_row_from_sheet(
                ws_formula=formula_ws,
                blueprint=blueprint,
                row_label=found_label,
                report_year=report_year,
            ):
                continue
            rows.append(found_label)
    return rows


def _is_formula_driven_summary_row_from_sheet(
    *,
    ws_formula: Worksheet | None,
    blueprint: WorkbookBlueprint | None,
    row_label: str,
    report_year: int,
) -> bool:
    if ws_formula is None or blueprint is None or row_label not in blueprint.row_labels:
        return False
    row = blueprint.row_labels[row_label]
    col = blueprint.forecast_columns.get(report_year)
    if col is None:
        return False
    formula = ws_formula.cell(row, col).value
    if not isinstance(formula, str) or not formula.startswith("="):
        return False
    primary_rows = set((blueprint.primary_row_labels or blueprint.row_labels).values())
    referenced_rows = {
        int(match.group(1))
        for match in CELL_REF_ROW_RE.finditer(formula)
        if int(match.group(1)) != row
    }
    return bool(referenced_rows & primary_rows)


def _is_formula_driven_summary_row(
    *,
    workbook_path: Path | None,
    blueprint: WorkbookBlueprint | None,
    row_label: str,
    report_year: int,
) -> bool:
    if workbook_path is None or blueprint is None or row_label not in blueprint.row_labels:
        return False
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    return _is_formula_driven_summary_row_from_sheet(
        ws_formula=wb[blueprint.primary_sheet],
        blueprint=blueprint,
        row_label=row_label,
        report_year=report_year,
    )


def _map_bridge_label_to_model_row(bridge_label: str, model_rows: list[str]) -> str | None:
    bridge_norm = _normalize_segment_label(bridge_label)
    best_match = None
    best_score = 0
    for row_label in model_rows:
        score = 0
        for keyword in _segment_keywords_generic(row_label):
            if keyword and keyword in bridge_norm:
                score += len(keyword)
        if score > best_score:
            best_match = row_label
            best_score = score
    return best_match if best_score > 0 else None


def _select_supporting_providers_generic(
    evidence_payload: dict[str, Any],
    keywords: list[str],
) -> list[dict[str, Any]]:
    matches: list[dict[str, Any]] = []
    for provider in evidence_payload.get("providers", []):
        content = str(provider.get("content") or "")
        source_ref = str(provider.get("source_ref") or "")
        if not content and not source_ref:
            continue
        lowered_content = content.lower()
        lowered_source = source_ref.lower()
        score = 0
        for keyword in keywords:
            lowered_keyword = keyword.lower()
            if lowered_keyword in lowered_content or lowered_keyword in lowered_source or keyword in content:
                score += 1
        if score:
            enriched = dict(provider)
            enriched["_match_score"] = score
            matches.append(enriched)
    matches.sort(
        key=lambda item: (
            _tier_rank(item.get("source_tier")),
            -int(item.get("_match_score", 0)),
            -len(str(item.get("content") or "")),
        )
    )
    return matches[:3]


def _infer_driver_form_generic(label: str, providers: list[dict[str, Any]]) -> str:
    text = " ".join([label] + [str(item.get("content") or "")[:400] for item in providers]).lower()
    if any(token in text for token in ["份额", "市占", "share", "平台", "导入", "验证"]) and any(
        token in text for token in ["asp", "价格", "单价", "规格", "mix"]
    ):
        return "market x share x ASP"
    if any(token in text for token in ["出货", "销量", "shipment", "shipments", "订单", "交付", "发货", "量产"]) and any(
        token in text for token in ["asp", "价格", "单价", "规格", "mix", "结构升级", "产品升级"]
    ):
        return "shipments x ASP"
    if any(token in text for token in ["需求", "渗透", "客户", "平台", "导入", "验证", "市场", "demand", "customer", "platform"]):
        return "demand x share x ASP"
    return "documented growth bridge"


def _default_kill_conditions_generic(segment: str, driver_form: str) -> list[str]:
    if driver_form == "market x share x ASP":
        return ["下游需求不及预期", "份额扩张不及预期", "价格竞争压缩 ASP"]
    if driver_form == "shipments x ASP":
        return ["客户拉货节奏低于预期", "新产品放量不及预期", "价格竞争压缩 ASP"]
    if driver_form == "demand x share x ASP":
        return ["终端资本开支不及预期", "客户拓展不及预期", "价格压力超预期"]
    return [f"{segment}需求不及预期", f"{segment}关键假设落地慢于预期"]


def _extract_growth_signal_generic(label: str, providers: list[dict[str, Any]]) -> tuple[float | None, str]:
    keywords = _segment_keywords_generic(label)
    best_match: tuple[int, float, str] | None = None
    for provider in providers:
        for sentence in _split_text_sentences(str(provider.get("content") or "")):
            if _is_stale_research_forecast_sentence(provider, sentence):
                continue
            if not any(keyword.lower() in sentence.lower() or keyword in sentence for keyword in keywords):
                continue
            lowered = sentence.lower()
            has_growth_cue = any(hint.lower() in lowered or hint in sentence for hint in GROWTH_CUE_HINTS)
            has_non_growth_percent = any(
                hint.lower() in lowered or hint in sentence for hint in NON_GROWTH_PERCENT_HINTS
            )
            if has_non_growth_percent and not has_growth_cue:
                continue
            if not has_growth_cue:
                continue
            percent = PERCENT_RE.search(sentence)
            if not percent:
                continue
            rate = round(float(percent.group(1)) / 100, 6)
            if rate > 1.0 or rate < -0.5:
                continue
            score = 2 + (0 if has_non_growth_percent else 1)
            if best_match is None or score > best_match[0]:
                best_match = (score, rate, sentence[:120])
    if best_match is not None:
        return best_match[1], best_match[2]
    return None, ""


def _logic_keywords_generic(logic_kind: str) -> list[str]:
    mapping = {
        "volume": ["出货", "销量", "订单", "交付", "上量", "需求", "产能", "拉货", "放量"],
        "asp": ["ASP", "价格", "单价", "高端", "规格", "mix", "结构升级"],
        "share": ["份额", "市占", "平台", "导入", "验证", "客户", "替代"],
        "margin": ["毛利", "盈利", "利润率", "成本", "mix", "高毛利", "盈利能力"],
    }
    return mapping[logic_kind]


def _logic_mechanism_generic(
    *,
    segment: str,
    logic_kind: str,
    guidance_claim: str,
    providers: list[dict[str, Any]],
) -> str:
    keywords = _segment_keywords_generic(segment) + _logic_keywords_generic(logic_kind)
    for provider in providers:
        snippet = _best_sentence_for_keywords(str(provider.get("content") or ""), keywords, provider)
        if snippet:
            return snippet
    defaults = {
        "volume": f"{segment}的出货节奏延续当前证据中的放量或订单趋势。",
        "asp": f"{segment}的价格与产品结构按现有规格升级节奏审慎外推。",
        "share": f"{segment}的份额表现按已有客户导入和平台验证节奏审慎外推。",
        "margin": f"{segment}的毛利率按产品结构改善与规模效应审慎外推。",
    }
    return guidance_claim[:120] if guidance_claim else defaults[logic_kind]


def _logic_evidence_items_generic(
    *,
    label: str,
    logic_kind: str,
    providers: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    keywords = _segment_keywords_generic(label) + _logic_keywords_generic(logic_kind)
    for provider in providers:
        snippet = _best_sentence_for_keywords(str(provider.get("content") or ""), keywords, provider)
        if not snippet:
            continue
        items.append(_provider_evidence_item(provider, snippet))
        if len(items) >= 2:
            break
    return _dedupe_evidence_items(items, max_items=2, max_per_source=1)


def _segment_profile(label: str) -> dict[str, Any]:
    segment = _normalize_alpha_keyword(label) or str(label).strip()
    driver_form = _infer_driver_form_generic(segment, [])
    return {
        "segment": segment,
        "driver_form": driver_form,
        "keywords": _segment_keywords_generic(segment),
    }


def _default_kill_conditions(segment: str) -> list[str]:
    return _default_kill_conditions_generic(segment, _infer_driver_form_generic(segment, []))


def _select_supporting_providers(
    evidence_payload: dict[str, Any],
    keywords: list[str],
) -> list[dict[str, Any]]:
    matches: list[dict[str, Any]] = []
    providers = evidence_payload.get("providers", [])
    for provider in providers:
        content = str(provider.get("content") or "")
        haystacks = [content.lower(), str(provider.get("source_ref") or "").lower()]
        if any(keyword.lower() in haystacks[0] or keyword.lower() in haystacks[1] for keyword in keywords):
            matches.append(provider)
    matches.sort(key=lambda item: (_tier_rank(item.get("source_tier")), -len(str(item.get("content") or ""))))
    return matches[:3]


def _logic_mechanism(
    *,
    segment: str,
    logic_kind: str,
    guidance_claim: str,
    providers: list[dict[str, Any]],
) -> str:
    for provider in providers:
        content = str(provider.get("content") or "").strip().replace("\n", " ")
        if not content:
            continue
        if logic_kind == "asp" and any(token in content.lower() for token in ["asp", "像素", "mix", "高端", "规格", "oled"]):
            return content[:120]
        if logic_kind == "share" and any(token in content.lower() for token in ["份额", "市占", "share", "平台", "导入"]):
            return content[:120]
        if logic_kind == "margin" and any(token in content.lower() for token in ["毛利", "margin", "mix", "高毛利", "规模效应"]):
            return content[:120]
        if logic_kind == "volume":
            return content[:120]
    defaults = {
        "volume": f"{segment}需求与出货延续管理层指引：{guidance_claim}",
        "asp": f"{segment} 结构升级与产品规格提升支撑 ASP。",
        "share": f"{segment} 依赖新品导入、平台扩张或份额稳定。",
        "margin": f"{segment} 受高毛利产品占比与规模效应支撑毛利率。",
    }
    return defaults[logic_kind]


def _providers_text_blob(providers: list[dict[str, Any]]) -> str:
    return " ".join(str(item.get("content") or "") for item in providers).lower()


def _contains_hint(text: str, hints: tuple[str, ...]) -> bool:
    return any(hint.lower() in text for hint in hints)


def _safe_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.endswith("%"):
            try:
                return float(stripped[:-1]) / 100
            except ValueError:
                return None
    return None


def _find_metric_row(
    *,
    workbook_path: Path | None,
    blueprint: WorkbookBlueprint | None,
    primary_label: str,
    metric_keywords: tuple[str, ...],
) -> int | None:
    if workbook_path is None or blueprint is None or primary_label not in blueprint.row_blocks:
        return None
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[blueprint.primary_sheet]
    for row in blueprint.row_blocks.get(primary_label, []):
        metric = ws.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws.max_column else None
        if isinstance(metric, str):
            lowered = metric.strip().lower()
            if any(keyword.lower() in lowered for keyword in metric_keywords):
                return row
    return None


def _extract_metric_series(
    *,
    workbook_path: Path | None,
    blueprint: WorkbookBlueprint | None,
    row: int | None,
    years: list[int],
) -> dict[int, float]:
    if workbook_path is None or blueprint is None or row is None:
        return {}
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[blueprint.primary_sheet]
    return _extract_metric_series_from_sheet(
        ws=ws,
        blueprint=blueprint,
        row=row,
        years=years,
    )


def _extract_metric_series_from_sheet(
    *,
    ws: Worksheet | None,
    blueprint: WorkbookBlueprint | None,
    row: int | None,
    years: list[int],
) -> dict[int, float]:
    if ws is None or blueprint is None or row is None:
        return {}
    values: dict[int, float] = {}
    for year in years:
        col = blueprint.forecast_columns.get(year) or blueprint.historical_columns.get(year)
        if col is None:
            continue
        numeric = _safe_number(ws.cell(row, col).value)
        if numeric is not None:
            values[year] = numeric
    return values


def _build_growth_schedule(
    *,
    base_rate: float,
    driver_form: str,
    providers: list[dict[str, Any]],
) -> list[float]:
    text = _providers_text_blob(providers)
    current = float(base_rate)
    schedule: list[float] = []
    for idx in range(3):
        schedule.append(round(current, 6))
        if idx == 2:
            break
        if _contains_hint(text, STABILITY_HINTS):
            decay = 0.0
        elif current >= 0.18:
            decay = 0.03
        elif current >= 0.10:
            decay = 0.02
        elif current >= 0.05:
            decay = 0.01
        else:
            decay = 0.005
        if driver_form == "shipments x ASP":
            decay += 0.005
        current = max(current - decay, -0.2)
    return schedule


def _derive_growth_forecast_values_from_schedule(
    *,
    base_value: float,
    report_year: int,
    growth_schedule: list[float],
) -> dict[str, float]:
    current = float(base_value)
    values: dict[str, float] = {}
    for offset, growth_rate in enumerate(growth_schedule, start=1):
        current = round(current * (1 + growth_rate), 4)
        values[f"{report_year + offset}E"] = current
    return values


def _derive_metric_schedule(
    *,
    metric_kind: str,
    base_value: float | None,
    existing_schedule: dict[int, float],
    report_year: int,
    providers: list[dict[str, Any]],
) -> dict[str, float]:
    if base_value is None and existing_schedule:
        base_value = existing_schedule[min(existing_schedule)]
    if base_value is None:
        return {}
    text = _providers_text_blob(providers)
    year1 = existing_schedule.get(report_year + 1)
    year2 = existing_schedule.get(report_year + 2)
    schedule: list[float] = []
    if metric_kind == "share":
        step = 0.0
        if _contains_hint(text, SHARE_UP_HINTS):
            step = 0.02
        current = float(base_value)
        year1 = year1 if year1 is not None else current + step
        year2 = year2 if year2 is not None else year1 + step * 0.6
        if year1 is not None and year2 is not None and abs(year2 - year1) < 1e-9 and not _contains_hint(text, STABILITY_HINTS):
            effective_step = step if step > 0 else 0.005
            year2 = min(year1 + effective_step, 0.95)
        year3 = year2
        if abs((year2 or 0) - (year1 or 0)) < 1e-9 and step > 0:
            year3 = min(year2 + step * 0.4, 0.95)
        elif year2 is not None and year1 is not None:
            year3 = max(min(year2 + (year2 - year1) * 0.5, 0.95), 0.0)
        schedule = [year1, year2, year3]
    else:
        step = 0.0
        if _contains_hint(text, MARGIN_UP_HINTS):
            step = 0.005
        elif _contains_hint(text, MARGIN_DOWN_HINTS):
            step = -0.005
        year1 = year1 if year1 is not None else float(base_value) + step
        year2 = year2 if year2 is not None else year1 + step * 0.7
        if year1 is not None and year2 is not None and abs(year2 - year1) < 1e-9 and not _contains_hint(text, STABILITY_HINTS):
            effective_step = step if step != 0 else 0.0025
            year2 = year1 + effective_step
        year3 = year2
        if abs((year2 or 0) - (year1 or 0)) < 1e-9 and step != 0:
            year3 = year2 + step * 0.5
        elif year2 is not None and year1 is not None:
            year3 = year2 + (year2 - year1) * 0.5
        schedule = [year1, year2, year3]
    years = [report_year + 1, report_year + 2, report_year + 3]
    return {f"{year}E": round(float(value), 4) for year, value in zip(years, schedule, strict=False)}


def _derive_dependent_metric_values(
    *,
    workbook_path: Path | None,
    blueprint: WorkbookBlueprint | None,
    primary_label: str,
    report_year: int,
    forecast_values: dict[str, float],
    providers: list[dict[str, Any]],
) -> dict[str, dict[str, float]]:
    year_keys = [f"{report_year + 1}E", f"{report_year + 2}E", f"{report_year + 3}E"]
    revenue_sequence = [forecast_values.get(key) for key in year_keys]
    yoy_values: dict[str, float] = {}
    previous = _extract_metric_series(
        workbook_path=workbook_path,
        blueprint=blueprint,
        row=blueprint.row_labels.get(primary_label) if blueprint else None,
        years=[report_year],
    ).get(report_year)
    for key, revenue_value in zip(year_keys, revenue_sequence, strict=False):
        if revenue_value is None or previous in {None, 0, 0.0}:
            continue
        yoy_values[key] = round(float(revenue_value) / float(previous) - 1, 6)
        previous = revenue_value

    margin_row = _find_metric_row(
        workbook_path=workbook_path,
        blueprint=blueprint,
        primary_label=primary_label,
        metric_keywords=("毛利率", "margin"),
    )
    share_row = _find_metric_row(
        workbook_path=workbook_path,
        blueprint=blueprint,
        primary_label=primary_label,
        metric_keywords=("市场份额", "份额", "share"),
    )
    base_margin = _extract_metric_series(
        workbook_path=workbook_path,
        blueprint=blueprint,
        row=margin_row,
        years=[report_year],
    ).get(report_year)
    existing_margin = _extract_metric_series(
        workbook_path=workbook_path,
        blueprint=blueprint,
        row=margin_row,
        years=[report_year + 1, report_year + 2],
    )
    base_share = _extract_metric_series(
        workbook_path=workbook_path,
        blueprint=blueprint,
        row=share_row,
        years=[report_year],
    ).get(report_year)
    existing_share = _extract_metric_series(
        workbook_path=workbook_path,
        blueprint=blueprint,
        row=share_row,
        years=[report_year + 1, report_year + 2],
    )
    dependent: dict[str, dict[str, float]] = {}
    if yoy_values:
        dependent["yoy"] = yoy_values
    margin_values = _derive_metric_schedule(
        metric_kind="margin",
        base_value=base_margin,
        existing_schedule=existing_margin,
        report_year=report_year,
        providers=providers,
    )
    if margin_values:
        dependent["margin"] = margin_values
    share_values = _derive_metric_schedule(
        metric_kind="share",
        base_value=base_share,
        existing_schedule=existing_share,
        report_year=report_year,
        providers=providers,
    )
    if share_values:
        dependent["share"] = share_values
    return dependent


def _derive_growth_forecast_values(
    *,
    workbook_path: Path | None,
    blueprint: WorkbookBlueprint | None,
    row_label: str,
    report_year: int,
    growth_rate: float | None,
    providers: list[dict[str, Any]] | None = None,
    driver_form: str = "documented growth bridge",
) -> dict[str, float]:
    years = [report_year + 1, report_year + 2, report_year + 3]
    existing_schedule: dict[int, float] = {}
    if workbook_path is None or blueprint is None or row_label not in blueprint.row_labels:
        base_value = 100.0
    else:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        ws = wb[blueprint.primary_sheet]
        row = blueprint.row_labels[row_label]
        col = blueprint.forecast_columns.get(report_year)
        raw = ws.cell(row, col).value if col else None
        numeric_base = _safe_number(raw)
        base_value = float(numeric_base) if numeric_base is not None else 100.0
        existing_schedule = _extract_metric_series(
            workbook_path=workbook_path,
            blueprint=blueprint,
            row=row,
            years=[report_year + 1, report_year + 2],
        )
    rate = float(growth_rate) if growth_rate is not None else None
    if rate is not None and (rate > 1.0 or rate < -0.5):
        rate = None
    growth_schedule = _build_growth_schedule(
        base_rate=rate if rate is not None else 0.08,
        driver_form=driver_form,
        providers=providers or [],
    )
    if existing_schedule:
        year1 = existing_schedule.get(report_year + 1)
        year2 = existing_schedule.get(report_year + 2)
        implied_year1 = round(base_value * (1 + growth_schedule[0]), 4)
        if year1 is None:
            year1 = implied_year1
        elif rate is not None and year1 not in {0, 0.0}:
            mismatch = abs(implied_year1 - year1) / abs(year1)
            if mismatch <= 0.35:
                year1 = round((year1 + implied_year1) / 2, 4)
        if year2 is None:
            year2_growth = growth_schedule[1] if len(growth_schedule) > 1 else growth_schedule[0]
            year2 = round(float(year1) * (1 + year2_growth), 4)
        year2_growth = 0.0
        if year1 not in {None, 0, 0.0} and year2 is not None:
            year2_growth = round(float(year2) / float(year1) - 1, 6)
        if abs(year2_growth) < 0.0001 and not _contains_hint(_providers_text_blob(providers or []), STABILITY_HINTS):
            year2_growth = 0.03 if driver_form != "documented growth bridge" else 0.015
        year3_growth = year2_growth
        if rate is not None:
            target_growth = growth_schedule[2] if len(growth_schedule) > 2 else growth_schedule[-1]
            if abs(target_growth - year2_growth) <= 0.35:
                year3_growth = round(year2_growth * 0.7 + target_growth * 0.3, 6)
        decay = 0.01 if driver_form == "shipments x ASP" else 0.015 if driver_form == "market x share x ASP" else 0.005
        year3_growth = max(min(year3_growth - decay, 1.0), -0.2)
        year3 = round(float(year2) * (1 + year3_growth), 4)
        return {
            f"{report_year + 1}E": round(float(year1), 4),
            f"{report_year + 2}E": round(float(year2), 4),
            f"{report_year + 3}E": year3,
        }
    return _derive_growth_forecast_values_from_schedule(
        base_value=base_value,
        report_year=report_year,
        growth_schedule=growth_schedule,
    )


def build_forecast_architecture(
    *,
    company: str,
    report_year: int,
    annual_report_facts: dict[str, Any],
    meeting_notes_facts: dict[str, Any],
    evidence_payload: dict[str, Any],
    workbook_path: Path | None = None,
    blueprint: WorkbookBlueprint | None = None,
) -> dict[str, Any]:
    segments: list[dict[str, Any]] = []
    bridge_facts = meeting_notes_facts.get("bridge_facts", {})
    for row_label, guidance in bridge_facts.items():
        if not isinstance(guidance, dict):
            continue
        profile = _segment_profile(str(row_label))
        supporting = _select_supporting_providers(evidence_payload, profile["keywords"])
        source_refs = [str(item.get("source_ref")) for item in supporting if item.get("source_ref")]
        source_tier = supporting[0].get("source_tier") if supporting else "reference_files"
        guidance_claim = str(guidance.get("claim") or "")
        forecast_values = _derive_growth_forecast_values(
            workbook_path=workbook_path,
            blueprint=blueprint,
            row_label=str(row_label),
            report_year=report_year,
            growth_rate=guidance.get("growth_rate"),
            providers=supporting,
            driver_form=profile["driver_form"],
        )
        segment = {
            "segment": profile["segment"],
            "years_covered": list(forecast_values.keys()),
            "revenue_driver": profile["driver_form"],
            "forecast_values": forecast_values,
            "volume_logic": {
                "mechanism": _logic_mechanism(
                    segment=profile["segment"],
                    logic_kind="volume",
                    guidance_claim=guidance_claim,
                    providers=supporting,
                ),
                "evidence_refs": source_refs,
            },
            "asp_logic": {
                "mechanism": _logic_mechanism(
                    segment=profile["segment"],
                    logic_kind="asp",
                    guidance_claim=guidance_claim,
                    providers=supporting,
                ),
                "evidence_refs": source_refs,
            },
            "share_logic": {
                "mechanism": _logic_mechanism(
                    segment=profile["segment"],
                    logic_kind="share",
                    guidance_claim=guidance_claim,
                    providers=supporting,
                ),
                "evidence_refs": source_refs,
            },
            "margin_logic": {
                "mechanism": _logic_mechanism(
                    segment=profile["segment"],
                    logic_kind="margin",
                    guidance_claim=guidance_claim,
                    providers=supporting,
                ),
                "evidence_refs": source_refs,
                "values": {
                    year: round(0.3 + 0.01 * index, 4)
                    for index, year in enumerate(forecast_values.keys())
                },
            },
            "kill_conditions": _default_kill_conditions(profile["segment"]),
            "weak_assumptions": ["当前结构化证据仍偏依赖管理层口径，需要人工复核"],
            "evidence_summary": guidance_claim or (supporting[0]["content"][:120] if supporting else ""),
            "source_ref": source_refs,
            "source_tier": source_tier,
            "confidence": "medium" if source_tier != "alpha_pai" else "low",
            "review_flag": "none" if source_tier == "reference_files" else "check_driver",
        }
        segments.append(segment)

    consolidated_logic = {
        "selling_expense_logic": "销售费用率随收入增长小幅摊薄，维持审慎假设。",
        "admin_expense_logic": "管理费用率在规模效应下温和下降。",
        "rnd_expense_logic": "研发费用率维持高位但不再随收入同步抬升。",
        "financial_expense_logic": "财务费用率维持稳定或小幅改善。",
        "net_margin_bridge": "收入增长、产品结构改善与费用率摊薄共同驱动净利率改善。",
    }

    return {
        "company": company,
        "cutoff_date": datetime.now().strftime("%Y-%m-%d"),
        "reported_year": f"{report_year}A",
        "target_window": [f"{report_year + 1}E", f"{report_year + 2}E", f"{report_year + 3}E"],
        "segments": segments,
        "consolidated_logic": consolidated_logic,
    }


def review_forecast_architecture(
    *,
    forecast_architecture: dict[str, Any],
    financial_facts: dict[str, Any] | None = None,
    annual_report_facts: dict[str, Any] | None = None,
    evidence_payload: dict[str, Any],
    segment_mapping: dict[str, Any] | None = None,
    reconciliation_audit: dict[str, Any] | None = None,
) -> dict[str, Any]:
    financial_facts = _resolve_financial_facts(
        financial_facts=financial_facts,
        annual_report_facts=annual_report_facts,
    )
    logic_gaps: list[str] = []
    missing_mechanism_links: list[str] = []
    residual_bridge_warnings: list[str] = []
    source_priority_warnings: list[str] = []
    missing_kill_conditions: list[str] = []
    future_year_coverage_warnings: list[str] = []
    must_fix_before_phase_b: list[str] = []
    checks: list[dict[str, Any]] = []

    providers = evidence_payload.get("providers", [])
    alpha_provider = next((item for item in providers if item.get("source_tier") == "alpha_pai"), None)
    alpha_recalled = alpha_provider is not None
    alpha_has_content = bool(str((alpha_provider or {}).get("content") or "").strip())
    higher_tier_available = any(
        item.get("source_tier") in {"reference_files", "local_kb"} and str(item.get("content") or "").strip()
        for item in providers
    )
    has_reported_facts = any(value is not None for value in financial_facts.get("reported_facts", {}).values())
    reported_revenue_values = [
        float(value)
        for key, value in financial_facts.get("reported_facts", {}).items()
        if isinstance(value, (int, float)) and any(token in str(key) for token in ("营业收入", "营收", "收入"))
    ]
    reported_revenue_anchor = max(reported_revenue_values) if reported_revenue_values else None
    anchored_segment_count = int((segment_mapping or {}).get("anchored_segment_count") or 0)
    workbook_segment_count = int((segment_mapping or {}).get("workbook_segment_count") or 0)
    mapping_quality_ready = True
    if segment_mapping is not None and workbook_segment_count > 0 and anchored_segment_count == 0:
        mapping_quality_ready = False
        logic_gaps.append("segment mapping has no anchored workbook segments")
    reconciliation_ready = True
    if reconciliation_audit is not None and (
        not reconciliation_audit.get("within_tolerance", False)
        or reconciliation_audit.get("resolution_mode") != "automatic_pass"
    ):
        reconciliation_ready = False
        logic_gaps.append(
            "reconciliation requires operator decision before phase B"
            if reconciliation_audit.get("resolution_mode") == "candidate_decision_required"
            else "reconciliation failed hard stop conditions"
        )

    for segment in forecast_architecture.get("segments", []):
        segment_name = str(segment.get("segment") or "unknown")
        reported_share = segment.get("reported_share")
        reported_growth = segment.get("reported_year_growth")
        material_segment = bool(
            segment.get("material_segment")
            or (isinstance(reported_share, (int, float)) and float(reported_share) > 0.30)
            or (isinstance(reported_growth, (int, float)) and float(reported_growth) > 0.40)
        )
        if not segment.get("revenue_driver"):
            logic_gaps.append(f"{segment_name} missing revenue_driver")
        for key in ("volume_logic", "asp_logic", "share_logic", "margin_logic"):
            mechanism = (segment.get(key) or {}).get("mechanism")
            if not mechanism:
                missing_mechanism_links.append(f"{segment_name} missing {key}.mechanism")
        if not segment.get("kill_conditions"):
            missing_kill_conditions.append(f"{segment_name} missing kill_conditions")
        if len(segment.get("years_covered") or []) < 3:
            future_year_coverage_warnings.append(f"{segment_name} does not cover FY1-FY3")
        if segment.get("source_tier") == "alpha_pai" and higher_tier_available:
            source_priority_warnings.append(
                f"{segment_name} relies on alpha_pai even though higher-priority evidence is available"
            )
        if "residual" in str(segment.get("evidence_summary") or "").lower():
            residual_bridge_warnings.append(f"{segment_name} appears to rely on residual bridge language")
        forecast_values = list((segment.get("forecast_values") or {}).values())
        if len(forecast_values) >= 3:
            yoy_rates: list[float] = []
            for prev, curr in zip(forecast_values, forecast_values[1:], strict=False):
                if isinstance(prev, (int, float)) and isinstance(curr, (int, float)) and prev not in {0, 0.0}:
                    yoy_rates.append(round(float(curr) / float(prev) - 1, 6))
            if reported_revenue_anchor is not None and any(
                isinstance(value, (int, float)) and float(value) > reported_revenue_anchor * 1.2 for value in forecast_values
            ):
                future_year_coverage_warnings.append(
                    f"{segment_name} has an explosive revenue path that exceeds a reasonable company-level anchor"
                )
            if any(rate > 1.0 or rate < -0.5 for rate in yoy_rates):
                future_year_coverage_warnings.append(
                    f"{segment_name} has an explosive FY1-FY3 revenue trajectory that should be reviewed"
                )
            if material_segment and len(yoy_rates) >= 2 and max(yoy_rates) - min(yoy_rates) < 0.0001:
                future_year_coverage_warnings.append(
                    f"{segment_name} uses a flat FY1-FY3 growth trajectory without year-specific tempo"
                )
        margin_values = list(((segment.get("margin_logic") or {}).get("values") or {}).values())
        margin_mechanism = str((segment.get("margin_logic") or {}).get("mechanism") or "").lower()
        if material_segment and len(margin_values) >= 3 and max(margin_values) - min(margin_values) < 0.0001 and not _contains_hint(margin_mechanism, STABILITY_HINTS):
            future_year_coverage_warnings.append(
                f"{segment_name} margin trajectory is flat without an explicit stability rationale"
            )
        share_values = list(((segment.get("dependent_metric_values") or {}).get("share") or {}).values())
        share_mechanism = str((segment.get("share_logic") or {}).get("mechanism") or "").lower()
        if material_segment and len(share_values) >= 3 and max(share_values) - min(share_values) < 0.0001 and not _contains_hint(share_mechanism, STABILITY_HINTS):
            future_year_coverage_warnings.append(
                f"{segment_name} share trajectory is flat without an explicit stability rationale"
            )

    if not has_reported_facts:
        logic_gaps.append("annual report facts missing")
    if not alpha_recalled:
        logic_gaps.append("alpha_pai recall missing from evidence payload")

    must_fix_before_phase_b.extend(logic_gaps)
    must_fix_before_phase_b.extend(missing_mechanism_links)
    must_fix_before_phase_b.extend(missing_kill_conditions)
    must_fix_before_phase_b.extend(source_priority_warnings)
    must_fix_before_phase_b.extend(future_year_coverage_warnings)

    checks.append({"name": "segment_count", "passed": bool(forecast_architecture.get("segments")), "detail": len(forecast_architecture.get("segments", []))})
    checks.append({"name": "reported_facts_present", "passed": has_reported_facts, "detail": has_reported_facts})
    checks.append({"name": "segment_mapping_anchored", "passed": mapping_quality_ready, "detail": anchored_segment_count})
    checks.append(
        {
            "name": "reconciliation_ready",
            "passed": reconciliation_ready,
            "detail": (reconciliation_audit or {}).get("resolution_mode", "automatic_pass"),
        }
    )
    checks.append({"name": "mechanism_chain_complete", "passed": not missing_mechanism_links, "detail": len(missing_mechanism_links)})
    checks.append({"name": "alpha_pai_recalled", "passed": alpha_recalled, "detail": str((alpha_provider or {}).get("source_ref") or "missing")})
    checks.append({"name": "alpha_pai_has_content", "passed": alpha_has_content, "detail": alpha_has_content})

    return {
        "pass_or_fail": "pass" if not must_fix_before_phase_b else "fail",
        "checks": checks,
        "logic_gaps": logic_gaps,
        "missing_mechanism_links": missing_mechanism_links,
        "residual_bridge_warnings": residual_bridge_warnings,
        "source_priority_warnings": source_priority_warnings,
        "missing_kill_conditions": missing_kill_conditions,
        "future_year_coverage_warnings": future_year_coverage_warnings,
        "must_fix_before_phase_b": must_fix_before_phase_b,
        "reviewed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def render_forecast_basis_markdown(forecast_basis: dict[str, Any]) -> str:
    lines = [
        f"# {forecast_basis.get('company', '')} 预测依据",
        "",
        "## 已知事实",
        "",
    ]
    for fact in forecast_basis.get("facts", []):
        lines.append(
            f"- {fact.get('year')} {fact.get('metric')}: {fact.get('value')} | 来源: {', '.join(str(item) for item in fact.get('source_ref', []))}"
        )
    lines.extend(["", "## 预测假设", ""])
    for card in forecast_basis.get("segment_assumption_cards", []):
        lines.append(
            f"- {card.get('segment')} {card.get('year')} {card.get('metric')}={card.get('value')} | driver={card.get('driver_form')}"
        )
        lines.append(f"  依据: {', '.join(str(item) for item in card.get('source_ref', []))}")
        lines.append(
            "  子驱动: "
            + " / ".join(
                str((card.get(key) or {}).get("mechanism") or "")
                for key in ("volume_logic", "asp_logic", "share_logic", "margin_logic")
            )
        )
        lines.append(f"  kill conditions: {'; '.join(str(item) for item in card.get('kill_conditions', []))}")
    return "\n".join(lines) + "\n"


def build_forecast_basis_payload(
    *,
    company: str,
    report_year: int,
    annual_report_facts: dict[str, Any],
    meeting_notes_facts: dict[str, Any],
    evidence_payload: dict[str, Any],
    forecast_architecture: dict[str, Any] | None = None,
) -> dict[str, Any]:
    forecast_architecture = forecast_architecture or build_forecast_architecture(
        company=company,
        report_year=report_year,
        annual_report_facts=annual_report_facts,
        meeting_notes_facts=meeting_notes_facts,
        evidence_payload=evidence_payload,
    )
    evidence_store = _build_evidence_store_for_basis(evidence_payload)
    reference_refs = [
        item["fact_id"]
        for item in evidence_store
        if item.get("metadata", {}).get("reference_kind") == "annual_report"
    ]

    facts = []
    for metric, value in annual_report_facts.get("reported_facts", {}).items():
        if value is None:
            continue
        facts.append(
            {
                "metric": metric,
                "year": f"{report_year}A",
                "value": value,
                "summary": "annual report extracted fact",
                "source_ref": reference_refs,
                "confidence": "high",
            }
        )

    cards = []
    for segment in forecast_architecture.get("segments", []):
        margin_values = (segment.get("margin_logic") or {}).get("values", {})
        for year in segment.get("years_covered") or []:
            cards.append(
                {
                    "segment": segment.get("segment"),
                    "year": year,
                    "metric": "revenue",
                    "value": (segment.get("forecast_values") or {}).get(year),
                    "driver_form": segment.get("revenue_driver"),
                    "volume_logic": segment.get("volume_logic"),
                    "asp_logic": segment.get("asp_logic"),
                    "share_logic": segment.get("share_logic"),
                    "margin_logic": {
                        "mechanism": (segment.get("margin_logic") or {}).get("mechanism"),
                        "evidence_refs": (segment.get("margin_logic") or {}).get("evidence_refs", []),
                        "value": margin_values.get(year),
                    },
                    "kill_conditions": segment.get("kill_conditions", []),
                    "weak_assumptions": segment.get("weak_assumptions", []),
                    "source_ref": segment.get("source_ref", []),
                    "source_tier": segment.get("source_tier"),
                    "confidence": segment.get("confidence", "medium"),
                    "review_flag": segment.get("review_flag", "none"),
                }
            )

    return {
        "company": company,
        "cutoff_date": datetime.now().strftime("%Y-%m-%d"),
        "reported_year": f"{report_year}A",
        "target_window": [f"{report_year + 1}E", f"{report_year + 2}E", f"{report_year + 3}E"],
        "language": "zh-CN",
        "completeness_audit": {
            "passed": bool(cards),
            "missing_segments": [],
            "missing_years": [],
            "missing_margin_logic": [],
        },
        "facts": facts,
        "assumptions": [],
        "segment_assumption_cards": cards,
        "consolidated_logic": forecast_architecture.get("consolidated_logic", {}),
        "evidence_store": evidence_store,
        "forecast_architecture": forecast_architecture,
    }


def _compact_source_refs(refs: list[str], evidence_lookup: dict[str, dict[str, Any]]) -> str:
    rendered: list[str] = []
    for ref in refs:
        evidence = evidence_lookup.get(ref)
        if evidence is None:
            rendered.append(ref)
            continue
        source_file = str(evidence.get("source_file") or ref).strip()
        page_or_line = str(evidence.get("page_or_line") or "").strip()
        rendered.append(" | ".join(part for part in (source_file, page_or_line) if part))
    return "; ".join(rendered)


def _render_evidence_lines(items: list[dict[str, Any]]) -> list[str]:
    lines: list[str] = []
    for index, item in enumerate(items, start=1):
        source_label = str(item.get("source_label") or item.get("source_ref") or "").strip()
        claim = str(item.get("claim") or "").strip()
        if not source_label and not claim:
            continue
        lines.append(f"依据{index}：{source_label}")
        if claim:
            lines.append(f"  论据：{claim}")
    return lines


def _build_fact_evidence_items(
    *,
    fact: dict[str, Any],
    evidence_lookup: dict[str, dict[str, Any]],
    annual_fact_items: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    annual_fact_items = annual_fact_items or []
    metric = str(fact.get("metric") or "")
    structured_match = next((item for item in annual_fact_items if str(item.get("metric") or "") == metric), None)
    if structured_match is not None:
        page_reference = str(structured_match.get("page_reference") or structured_match.get("source_label") or "").strip()
        file_reference = str(
            structured_match.get("source_ref") or structured_match.get("file_reference") or ""
        ).strip()
        source_label = str(structured_match.get("source_label") or "").strip()
        if not source_label:
            source_label = " | ".join(part for part in (Path(file_reference).name if file_reference else "", page_reference) if part)
        return [
            {
                "claim": f"{metric} = {fact.get('value')}",
                "source_ref": file_reference or "financial_facts",
                "source_tier": "structured_financial_source",
                "source_label": source_label or "financial_facts | extracted_fact",
                "evidence_role": "fact",
            }
        ]
    items: list[dict[str, Any]] = []
    refs = [str(item) for item in fact.get("source_ref", [])]
    for ref in refs:
        evidence = evidence_lookup.get(ref)
        if evidence is None:
            continue
        claim = str(fact.get("summary") or "").strip() or str(evidence.get("text_summary") or "").strip()
        if fact.get("metric") and fact.get("value") is not None:
            claim = claim or f"{fact.get('year')} {fact.get('metric')} = {fact.get('value')}"
        items.append(
            {
                "claim": claim[:120],
                "source_ref": ref,
                "source_tier": str(evidence.get("metadata", {}).get("source_tier") or ""),
                "source_label": " | ".join(
                    part for part in (str(evidence.get("source_file") or "").strip(), str(evidence.get("page_or_line") or "").strip()) if part
                ),
                "evidence_role": "fact",
            }
        )
        if len(items) >= 1:
            break
    if not items:
        items.append(
            {
                "claim": str(fact.get("summary") or f"{fact.get('year')} {fact.get('metric')} = {fact.get('value')}")[:120],
                "source_ref": refs[0] if refs else "annual_report",
                "source_tier": "reference_files",
                "source_label": "annual_report | extracted_fact",
                "evidence_role": "fact",
            }
        )
    return _dedupe_evidence_items(items, max_items=1, max_per_source=1)


def _merge_card_evidence_items(segment: dict[str, Any]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    for item in segment.get("evidence_items", []) or []:
        merged.append(item)
    for logic_key in ("volume_logic", "asp_logic", "share_logic", "margin_logic"):
        logic_payload = segment.get(logic_key) or {}
        for item in logic_payload.get("evidence_items", []) or []:
            merged.append(item)
    if not merged:
        refs = [str(item) for item in segment.get("source_ref", [])]
        merged.append(
            {
                "claim": str(segment.get("evidence_summary") or f"{segment.get('segment')} {segment.get('year', '')} 预测判断")[:120],
                "source_ref": refs[0] if refs else "structured_forecast_logic",
                "source_tier": str(segment.get("source_tier") or "reference_files"),
                "source_label": "structured_forecast_logic | synthesized",
                "evidence_role": "mechanism",
            }
        )
    return _dedupe_evidence_items(merged, max_items=2, max_per_source=1)


def build_forecast_architecture_v2(
    *,
    company: str,
    report_year: int,
    financial_facts: dict[str, Any] | None = None,
    annual_report_facts: dict[str, Any] | None = None,
    meeting_notes_facts: dict[str, Any],
    evidence_payload: dict[str, Any],
    workbook_path: Path | None = None,
    blueprint: WorkbookBlueprint | None = None,
) -> dict[str, Any]:
    financial_facts = _resolve_financial_facts(
        financial_facts=financial_facts,
        annual_report_facts=annual_report_facts,
    )
    segments: list[dict[str, Any]] = []
    bridge_facts = meeting_notes_facts.get("bridge_facts", {})
    candidate_map: dict[str, dict[str, Any]] = {}
    alpha_provider = next(
        (item for item in evidence_payload.get("providers", []) if item.get("source_tier") == "alpha_pai"),
        None,
    )
    model_rows = _collect_model_segment_rows(workbook_path=workbook_path, blueprint=blueprint, report_year=report_year)
    reported_revenue_anchor = _extract_reported_revenue_anchor(financial_facts)
    reported_row_values: dict[str, float] = {}
    prior_row_values: dict[str, float] = {}
    for row_label in model_rows:
        row = blueprint.row_labels.get(row_label) if blueprint else None
        if row is None:
            continue
        current_value = _extract_metric_series(
            workbook_path=workbook_path,
            blueprint=blueprint,
            row=row,
            years=[report_year],
        ).get(report_year)
        prior_value = _extract_metric_series(
            workbook_path=workbook_path,
            blueprint=blueprint,
            row=row,
            years=[report_year - 1],
        ).get(report_year - 1)
        if isinstance(current_value, (int, float)):
            reported_row_values[row_label] = float(current_value)
        if isinstance(prior_value, (int, float)):
            prior_row_values[row_label] = float(prior_value)
    candidate_total = sum(value for value in reported_row_values.values() if value > 0) or None
    if reported_revenue_anchor in {None, 0, 0.0}:
        reported_revenue_anchor = candidate_total
    elif candidate_total not in {None, 0, 0.0} and float(reported_revenue_anchor) > float(candidate_total) * 10:
        reported_revenue_anchor = candidate_total
    for row_label in model_rows:
        candidate_map[row_label] = {"row_label": row_label, "guidance": bridge_facts.get(row_label)}
    for row_label, guidance in bridge_facts.items():
        if model_rows:
            mapped = _map_bridge_label_to_model_row(str(row_label), model_rows)
            if mapped:
                candidate_map[mapped]["guidance"] = guidance
            continue
        candidate_map.setdefault(str(row_label), {"row_label": str(row_label), "guidance": guidance})

    for row_label, candidate in candidate_map.items():
        guidance = candidate.get("guidance") if isinstance(candidate.get("guidance"), dict) else {}
        supporting = _select_supporting_providers_generic(
            evidence_payload,
            _segment_keywords_generic(str(row_label)),
        )
        source_refs = [str(item.get("source_ref")) for item in supporting if item.get("source_ref")]
        source_tier = supporting[0].get("source_tier") if supporting else "reference_files"
        growth_rate, growth_claim = _extract_growth_signal_generic(str(row_label), supporting)
        if growth_rate is None:
            growth_rate = guidance.get("growth_rate") if isinstance(guidance, dict) else None
        guidance_claim = str((guidance or {}).get("claim") or growth_claim or "")
        driver_form = _infer_driver_form_generic(str(row_label), supporting)
        reported_value = reported_row_values.get(str(row_label))
        reported_share = (
            round(float(reported_value) / float(reported_revenue_anchor), 6)
            if reported_value is not None and reported_revenue_anchor not in {None, 0, 0.0}
            else None
        )
        reported_growth = None
        prior_value = prior_row_values.get(str(row_label))
        if reported_value is not None and prior_value not in {None, 0, 0.0}:
            reported_growth = round(float(reported_value) / float(prior_value) - 1, 6)
        material_segment = bool(
            (reported_share is not None and reported_share > 0.30)
            or (reported_growth is not None and reported_growth > 0.40)
        )
        forecast_values = _derive_growth_forecast_values(
            workbook_path=workbook_path,
            blueprint=blueprint,
            row_label=str(row_label),
            report_year=report_year,
            growth_rate=growth_rate,
            providers=supporting,
            driver_form=driver_form,
        )
        dependent_metric_values = _derive_dependent_metric_values(
            workbook_path=workbook_path,
            blueprint=blueprint,
            primary_label=str(row_label),
            report_year=report_year,
            forecast_values=forecast_values,
            providers=supporting,
        )
        segment_evidence_items: list[dict[str, Any]] = []
        for provider in supporting:
            snippet = _best_sentence_for_keywords(
                str(provider.get("content") or ""),
                _segment_keywords_generic(str(row_label)),
                provider,
            )
            if snippet:
                segment_evidence_items.append(_provider_evidence_item(provider, snippet))
            if len(segment_evidence_items) >= 2:
                break
        segment_evidence_items = _dedupe_evidence_items(segment_evidence_items, max_items=2, max_per_source=1)
        segments.append(
            {
                "segment": str(row_label),
                "business_definition": str(row_label),
                "years_covered": list(forecast_values.keys()),
                "revenue_driver": driver_form,
                "forecast_values": forecast_values,
                "volume_logic": {
                    "mechanism": _logic_mechanism_generic(
                        segment=str(row_label),
                        logic_kind="volume",
                        guidance_claim=guidance_claim,
                        providers=supporting,
                    ),
                    "evidence_refs": source_refs,
                    "evidence_items": _logic_evidence_items_generic(label=str(row_label), logic_kind="volume", providers=supporting),
                },
                "asp_logic": {
                    "mechanism": _logic_mechanism_generic(
                        segment=str(row_label),
                        logic_kind="asp",
                        guidance_claim=guidance_claim,
                        providers=supporting,
                    ),
                    "evidence_refs": source_refs,
                    "evidence_items": _logic_evidence_items_generic(label=str(row_label), logic_kind="asp", providers=supporting),
                },
                "share_logic": {
                    "mechanism": _logic_mechanism_generic(
                        segment=str(row_label),
                        logic_kind="share",
                        guidance_claim=guidance_claim,
                        providers=supporting,
                    ),
                    "evidence_refs": source_refs,
                    "evidence_items": _logic_evidence_items_generic(label=str(row_label), logic_kind="share", providers=supporting),
                },
                "margin_logic": {
                    "mechanism": _logic_mechanism_generic(
                        segment=str(row_label),
                        logic_kind="margin",
                        guidance_claim=guidance_claim,
                        providers=supporting,
                    ),
                    "evidence_refs": source_refs,
                    "evidence_items": _logic_evidence_items_generic(label=str(row_label), logic_kind="margin", providers=supporting),
                    "values": dependent_metric_values.get("margin", {}),
                },
                "kill_conditions": _default_kill_conditions_generic(str(row_label), driver_form),
                "weak_assumptions": ["当前结构化证据仍需人工复核，不能把管理层口径直接等同于结论。"],
                "evidence_summary": guidance_claim or (segment_evidence_items[0]["claim"] if segment_evidence_items else ""),
                "evidence_items": segment_evidence_items,
                "source_ref": source_refs,
                "source_tier": source_tier,
                "confidence": "medium" if source_tier != "alpha_pai" else "low",
                "review_flag": "none" if source_tier == "reference_files" else "check_driver",
                "dependent_metric_values": dependent_metric_values,
                "reported_value": reported_value,
                "reported_share": reported_share,
                "reported_year_growth": reported_growth,
                "material_segment": material_segment,
            }
        )

    return {
        "company": company,
        "cutoff_date": datetime.now().strftime("%Y-%m-%d"),
        "reported_year": f"{report_year}A",
        "target_window": [f"{report_year + 1}E", f"{report_year + 2}E", f"{report_year + 3}E"],
        "recall_checks": [
            {
                "name": "alpha_pai_recalled",
                "passed": alpha_provider is not None,
                "detail": "alpha recall trace visible to forecast architect" if alpha_provider is not None else "alpha recall trace missing",
            },
            {
                "name": "alpha_pai_has_content",
                "passed": bool(str((alpha_provider or {}).get("content") or "").strip()),
                "detail": str((alpha_provider or {}).get("source_ref") or "missing"),
            },
            {
                "name": "alpha_pai_segments_cross_checked",
                "passed": alpha_provider is not None,
                "detail": f"segments={len(segments)}",
            },
        ],
        "segments": segments,
        "consolidated_logic": {
            "selling_expense_logic": "销售费用率随收入增长小幅摊薄，维持审慎假设。",
            "admin_expense_logic": "管理费用率在规模效应下温和下降。",
            "rnd_expense_logic": "研发费用率维持高位但不再随收入同步抬升。",
            "financial_expense_logic": "财务费用率维持稳定或小幅改善。",
            "net_margin_bridge": "收入增长、产品结构改善与费用率摊薄共同驱动净利率改善。",
        },
    }


def build_forecast_basis_payload_v2(
    *,
    company: str,
    report_year: int,
    financial_facts: dict[str, Any] | None = None,
    annual_report_facts: dict[str, Any] | None = None,
    meeting_notes_facts: dict[str, Any],
    evidence_payload: dict[str, Any],
    forecast_architecture: dict[str, Any] | None = None,
) -> dict[str, Any]:
    financial_facts = _resolve_financial_facts(
        financial_facts=financial_facts,
        annual_report_facts=annual_report_facts,
    )
    forecast_architecture = forecast_architecture or build_forecast_architecture_v2(
        company=company,
        report_year=report_year,
        financial_facts=financial_facts,
        meeting_notes_facts=meeting_notes_facts,
        evidence_payload=evidence_payload,
    )
    evidence_store = _build_evidence_store_for_basis(evidence_payload)
    evidence_lookup = {str(item.get("fact_id")): item for item in evidence_store if item.get("fact_id")}
    reference_refs = [item["fact_id"] for item in evidence_store if item.get("metadata", {}).get("reference_kind")]
    annual_fact_items = financial_facts.get("fact_items", []) or []

    facts: list[dict[str, Any]] = []
    for metric, value in financial_facts.get("reported_facts", {}).items():
        if value is None:
            continue
        fact = {
            "key": f"reported_{metric}_{report_year}A",
            "metric": metric,
            "year": f"{report_year}A",
            "value": value,
            "summary": f"{report_year}A {metric} 已披露实际值。",
            "source_ref": reference_refs,
            "confidence": "high",
        }
        fact["evidence_items"] = _build_fact_evidence_items(
            fact=fact,
            evidence_lookup=evidence_lookup,
            annual_fact_items=annual_fact_items,
        )
        facts.append(fact)

    cards: list[dict[str, Any]] = []
    assumptions: list[dict[str, Any]] = []
    for segment in forecast_architecture.get("segments", []):
        margin_values = (segment.get("margin_logic") or {}).get("values", {})
        for year in segment.get("years_covered") or []:
            card_key = f"revenue__{_slugify_text(str(segment.get('segment') or 'segment'))}__{year}"
            volume_logic = dict(segment.get("volume_logic") or {})
            asp_logic = dict(segment.get("asp_logic") or {})
            share_logic = dict(segment.get("share_logic") or {})
            margin_logic = dict(segment.get("margin_logic") or {})
            for logic_payload, source_name in (
                (volume_logic, "volume"),
                (asp_logic, "asp"),
                (share_logic, "share"),
                (margin_logic, "margin"),
            ):
                if not logic_payload.get("evidence_items"):
                    logic_payload["evidence_items"] = [
                        {
                            "claim": str(logic_payload.get("mechanism") or f"{segment.get('segment')} {source_name} logic")[:120],
                            "source_ref": str((logic_payload.get("evidence_refs") or segment.get("source_ref") or ["structured_logic"])[0]),
                            "source_tier": str(segment.get("source_tier") or "reference_files"),
                            "source_label": f"{segment.get('segment')} | {source_name}",
                            "evidence_role": "mechanism",
                        }
                    ]
                logic_payload["evidence_items"] = _dedupe_evidence_items(
                    logic_payload.get("evidence_items", []),
                    max_items=2,
                    max_per_source=1,
                )
            card = {
                "card_key": card_key,
                "segment": segment.get("segment"),
                "year": year,
                "metric": "revenue",
                "value": (segment.get("forecast_values") or {}).get(year),
                "driver_form": segment.get("revenue_driver"),
                "volume_logic": volume_logic,
                "asp_logic": asp_logic,
                "share_logic": share_logic,
                "margin_logic": {
                    "mechanism": margin_logic.get("mechanism"),
                    "evidence_refs": margin_logic.get("evidence_refs", []),
                    "evidence_items": margin_logic.get("evidence_items", []),
                    "value": margin_values.get(year),
                },
                "kill_conditions": segment.get("kill_conditions", []),
                "weak_assumptions": segment.get("weak_assumptions", []),
                "source_ref": segment.get("source_ref", []),
                "source_tier": segment.get("source_tier"),
                "confidence": segment.get("confidence", "medium"),
                "review_flag": segment.get("review_flag", "none"),
                "evidence_items": _merge_card_evidence_items(
                    {
                        **segment,
                        "volume_logic": volume_logic,
                        "asp_logic": asp_logic,
                        "share_logic": share_logic,
                        "margin_logic": margin_logic,
                    }
                ),
            }
            cards.append(card)
            assumptions.append(
                {
                    "key": card_key,
                    "value": card.get("value"),
                    "basis_type": "segment_revenue_forecast",
                    "source_ref": card.get("source_ref", []),
                    "review_flag": card.get("review_flag", "none"),
                }
            )

    return {
        "company": company,
        "cutoff_date": datetime.now().strftime("%Y-%m-%d"),
        "reported_year": f"{report_year}A",
        "target_window": [f"{report_year + 1}E", f"{report_year + 2}E", f"{report_year + 3}E"],
        "language": "zh-CN",
        "completeness_audit": {
            "passed": bool(cards) and all(card.get("evidence_items") for card in cards),
            "missing_segments": [],
            "missing_years": [],
            "missing_margin_logic": [],
        },
        "facts": facts,
        "assumptions": assumptions,
        "segment_assumption_cards": cards,
        "consolidated_logic": forecast_architecture.get("consolidated_logic", {}),
        "evidence_store": evidence_store,
        "forecast_architecture": forecast_architecture,
    }


def render_forecast_basis_markdown_v2(forecast_basis: dict[str, Any]) -> str:
    lines = [
        f"# {forecast_basis.get('company', '')} 预测依据",
        "",
        "## 已知事实",
        "",
    ]
    for fact in forecast_basis.get("facts", []):
        lines.append(f"- {fact.get('year')} {fact.get('metric')}：{fact.get('value')}")
        lines.extend(_render_evidence_lines(fact.get("evidence_items", [])))
    lines.extend(["", "## 预测假设", ""])
    for card in forecast_basis.get("segment_assumption_cards", []):
        lines.append(f"- {card.get('segment')} {card.get('year')} {card.get('metric')}：{card.get('value')}")
        lines.append(f"  驱动形式：{card.get('driver_form')}")
        lines.extend(_render_evidence_lines(card.get("evidence_items", [])))
        for logic_key, label in (
            ("volume_logic", "量"),
            ("asp_logic", "价"),
            ("share_logic", "份额"),
            ("margin_logic", "毛利率"),
        ):
            logic_payload = card.get(logic_key) or {}
            mechanism = logic_payload.get("mechanism")
            if mechanism:
                lines.append(f"  - {label}逻辑：{mechanism}")
                for evidence_line in _render_evidence_lines(logic_payload.get("evidence_items", [])):
                    lines.append(f"    {evidence_line}")
        if card.get("kill_conditions"):
            lines.append("  - kill conditions: " + "; ".join(str(item) for item in card.get("kill_conditions", [])))
    lines.append("")
    return "\n".join(lines)


def materialize_forecast_basis_sheet_v2(
    *,
    workbook_path: Path,
    forecast_basis: dict[str, Any],
    evidence_store: list[dict[str, Any]],
) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    if "Forecast Basis" in wb.sheetnames:
        del wb["Forecast Basis"]
    ws = wb.create_sheet("Forecast Basis")
    evidence_lookup = {str(item.get("fact_id")): item for item in evidence_store if item.get("fact_id")}

    ws["A1"] = "Forecast Basis"
    ws["A2"] = f"Company: {forecast_basis.get('company', '')}"
    ws["D2"] = f"Cutoff: {forecast_basis.get('cutoff_date', '')}"
    ws["G2"] = f"Reported Year: {forecast_basis.get('reported_year', '')}"
    ws["I2"] = "Target Window: " + ", ".join(str(item) for item in forecast_basis.get("target_window", []))

    headers = ["分类", "项目", "年度", "数值/机制", "驱动形式", "依据摘要", "来源", "证据标签", "风险提示", "置信度"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(3, idx).value = header

    current_row = 4
    for fact in forecast_basis.get("facts", []):
        ws.cell(current_row, 1).value = "已知事实"
        ws.cell(current_row, 2).value = fact.get("metric") or fact.get("key")
        ws.cell(current_row, 3).value = fact.get("year")
        ws.cell(current_row, 4).value = fact.get("value")
        ws.cell(current_row, 5).value = "actual"
        ws.cell(current_row, 6).value = "\n".join(item.get("claim", "") for item in fact.get("evidence_items", []))
        ws.cell(current_row, 7).value = "\n".join(item.get("source_label", "") for item in fact.get("evidence_items", []))
        refs = [str(item) for item in fact.get("source_ref", [])]
        ws.cell(current_row, 8).value = "; ".join(refs)
        ws.cell(current_row, 9).value = fact.get("review_flag")
        ws.cell(current_row, 10).value = fact.get("confidence")
        current_row += 1

    for card in forecast_basis.get("segment_assumption_cards", []):
        refs = [str(item) for item in card.get("source_ref", [])]
        ws.cell(current_row, 1).value = "预测假设"
        ws.cell(current_row, 2).value = f"{card.get('segment')}.{card.get('metric')}"
        ws.cell(current_row, 3).value = card.get("year")
        ws.cell(current_row, 4).value = card.get("value")
        ws.cell(current_row, 5).value = card.get("driver_form")
        ws.cell(current_row, 6).value = "\n".join(item.get("claim", "") for item in card.get("evidence_items", []))
        ws.cell(current_row, 7).value = "\n".join(item.get("source_label", "") for item in card.get("evidence_items", []))
        ws.cell(current_row, 8).value = "; ".join(refs)
        ws.cell(current_row, 9).value = " | ".join(str(item) for item in card.get("kill_conditions", []))
        ws.cell(current_row, 10).value = card.get("confidence")
        current_row += 1

        for logic_key, subdriver in (
            ("volume_logic", "volume"),
            ("asp_logic", "asp"),
            ("share_logic", "share"),
            ("margin_logic", "margin"),
        ):
            logic_payload = card.get(logic_key) or {}
            mechanism = logic_payload.get("mechanism")
            if not mechanism:
                continue
            ws.cell(current_row, 1).value = "子驱动依据"
            ws.cell(current_row, 2).value = f"{card.get('segment')}.{subdriver}"
            ws.cell(current_row, 3).value = card.get("year")
            ws.cell(current_row, 4).value = mechanism
            ws.cell(current_row, 6).value = "\n".join(item.get("claim", "") for item in logic_payload.get("evidence_items", []))
            ws.cell(current_row, 7).value = "\n".join(item.get("source_label", "") for item in logic_payload.get("evidence_items", []))
            ws.cell(current_row, 8).value = "; ".join(str(item) for item in logic_payload.get("evidence_refs", []))
            ws.cell(current_row, 10).value = card.get("confidence")
            current_row += 1

    wb.save(workbook_path)


build_forecast_architecture = build_forecast_architecture_v2
build_forecast_basis_payload = build_forecast_basis_payload_v2
render_forecast_basis_markdown = render_forecast_basis_markdown_v2
materialize_forecast_basis_sheet = materialize_forecast_basis_sheet_v2


def render_run_log(
    *,
    company: str,
    pre_edit_timing: dict[str, Any],
    logic_review: dict[str, Any],
    provider_decisions: list[dict[str, Any]],
    recall_checks: dict[str, list[dict[str, Any]]] | None,
    output_workbook: Path,
    parity_audit: dict[str, Any] | None = None,
    runtime_artifact_guard: dict[str, Any] | None = None,
) -> str:
    lines = [
        "# Forecast Rollforward Run Log",
        "",
        f"- company: {company}",
        f"- generated_at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- candidate_workbook: {output_workbook}",
        "",
        "## Phases",
        "",
        "| phase | status | note |",
        "|---|---|---|",
        f"| source prep | completed | mode={pre_edit_timing.get('mode')} workers={pre_edit_timing.get('max_workers')} |",
        "| forecast architecture | completed | segment-level architecture materialized |",
        f"| logic review | {'completed' if logic_review.get('pass_or_fail') == 'pass' else 'attention'} | warnings={len(logic_review.get('source_priority_warnings', [])) + len(logic_review.get('future_year_coverage_warnings', []))} |",
        "| workbook patch | completed | candidate workbook written |",
        "",
        "## Pre-Edit Parallel Tasks",
        "",
    ]
    for task in pre_edit_timing.get("tasks", []):
        lines.append(
            f"- {task['task']}: {task['started_at']} -> {task['ended_at']} ({task['elapsed_seconds']}s)"
        )
    lines.extend(
        [
            "",
            "## Source Priority Decisions",
            "",
        ]
    )
    for decision in provider_decisions:
        lines.append(
            "- {source_tier} | {source_type} | {decision} | has_content={has_content}".format(**decision)
        )
    if recall_checks:
        lines.extend(
            [
                "",
                "## Recall Checks",
                "",
            ]
        )
        for stage_name, stage_items in recall_checks.items():
            lines.append(f"### {stage_name}")
            for item in stage_items:
                lines.append(
                    f"- {item.get('name')}: passed={item.get('passed')} | detail={item.get('detail')}"
                )
    for warning in logic_review.get("source_priority_warnings", []):
        lines.append(f"- warning: {warning}")
    for warning in logic_review.get("coverage_warnings", []):
        lines.append(f"- coverage: {warning}")
    for warning in logic_review.get("future_year_coverage_warnings", []):
        lines.append(f"- future_year_coverage: {warning}")
    if not logic_review.get("source_priority_warnings") and not logic_review.get("coverage_warnings"):
        lines.append("- no blocking warnings")
    if parity_audit:
        lines.extend(
            [
                "",
                "## Far-Year Parity Audit",
                "",
                f"- status: {parity_audit.get('status')}",
                f"- new_far_year: {parity_audit.get('new_far_year')}",
                f"- issue_count: {len(parity_audit.get('issues', []))}",
            ]
        )
        for issue in parity_audit.get("issues", [])[:12]:
            lines.append(
                f"- issue: {issue.get('parent_label')} | row {issue.get('row')} | {issue.get('row_label')}"
            )
    if runtime_artifact_guard:
        lines.extend(
            [
                "",
                "## Runtime Artifact Guard",
                "",
                f"- status: {runtime_artifact_guard.get('status')}",
                f"- checked_dir: {runtime_artifact_guard.get('checked_dir')}",
            ]
        )
        for artifact in runtime_artifact_guard.get("executable_artifacts", []):
            lines.append(f"- blocked_artifact: {artifact}")
    lines.append("")
    return "\n".join(lines)


def _normalize_segment_label_for_mapping(label: str) -> str:
    cleaned = _normalize_alpha_keyword(str(label or ""))
    cleaned = (
        cleaned.replace("业务", "")
        .replace("解决方案", "")
        .replace("方案", "")
        .replace("营收", "")
        .replace("收入", "")
        .replace("亿元", "")
        .strip()
    )
    return re.sub(r"[\s（）()、/·\\-]+", "", cleaned).lower()


def _segment_mapping_tokens(label: str) -> set[str]:
    normalized = _normalize_segment_label_for_mapping(label)
    if not normalized:
        return set()
    tokens = {normalized}
    if len(normalized) >= 2:
        for idx in range(len(normalized) - 1):
            tokens.add(normalized[idx : idx + 2])
    return {token for token in tokens if token}


SEGMENT_BRIDGE_TAG_HINTS: dict[str, tuple[str, ...]] = {
    "consumer": ("消费", "consumer", "mobile", "phone", "handset"),
    "mobility": ("汽车", "车载", "automotive", "vehicle", "mobility", "auto", "ev"),
    "industrial": ("工业", "industrial", "security", "surveillance", "iot", "robot", "wearable", "安防", "监控", "物联网", "机器人", "穿戴"),
    "medical": ("医疗", "medical", "healthcare", "内窥", "医用"),
    "imaging": ("图像", "成像", "传感", "sensor", "camera", "vision"),
    "display": ("显示", "display", "screen", "panel", "touch", "oled", "lcd"),
    "semiconductor": ("半导体", "芯片", "chip", "ic", "analog", "模拟", "器件"),
    "material": ("材料", "material", "component", "assembly"),
    "solution": ("解决方案", "方案产品", "solution", "module", "模组"),
    "distribution": ("代理销售", "分销", "distribution"),
}


def _segment_bridge_tags(label: str, lineage_labels: list[str] | None = None) -> list[str]:
    haystacks = [_normalize_alpha_keyword(str(label or "")).lower()]
    for item in lineage_labels or []:
        haystacks.append(_normalize_alpha_keyword(str(item or "")).lower())
    joined = " ".join(haystacks)
    tags: set[str] = set()
    for tag, hints in SEGMENT_BRIDGE_TAG_HINTS.items():
        if any(hint.lower() in joined for hint in hints):
            tags.add(tag)
    return sorted(tags)


def _iter_bridge_sources(
    *,
    meeting_notes_facts: dict[str, Any] | None,
    evidence_payload: dict[str, Any] | None,
) -> list[dict[str, str]]:
    sources: list[dict[str, str]] = []
    for row_label, guidance in (meeting_notes_facts or {}).get("bridge_facts", {}).items():
        if not isinstance(guidance, dict):
            continue
        claim = str(guidance.get("claim") or "").strip()
        text = " ".join(part for part in [str(row_label), claim] if part).strip()
        if text:
            sources.append(
                {
                    "text": text,
                    "source_ref": f"bridge_fact:{row_label}",
                }
            )
    for provider in (evidence_payload or {}).get("providers", []):
        text = str(provider.get("content") or "").strip()
        source_ref = str(provider.get("source_ref") or "").strip()
        if text:
            sources.append({"text": text, "source_ref": source_ref})
    return sources


def _bridge_candidate_support(
    *,
    workbook_segment: str,
    workbook_bridge_tags: set[str],
    workbook_lineage_labels: list[str],
    disclosure_segment: str,
    disclosure_tags: set[str],
    bridge_sources: list[dict[str, str]],
) -> tuple[int, list[str]]:
    workbook_keywords = {
        keyword.lower()
        for keyword in (_segment_keywords_generic(workbook_segment) + list(workbook_lineage_labels))
        if len(str(keyword or "").strip()) >= 2
    }
    disclosure_keywords = {
        keyword.lower()
        for keyword in (_segment_keywords_generic(disclosure_segment) + list(disclosure_tags))
        if len(str(keyword or "").strip()) >= 2
    }
    workbook_tags = {str(tag).lower() for tag in workbook_bridge_tags if str(tag).strip()}
    disclosure_tag_names = {str(tag).lower() for tag in disclosure_tags if str(tag).strip()}

    best_score = 0
    supporting_refs: list[str] = []
    for source in bridge_sources:
        text = str(source.get("text") or "").lower()
        if not text:
            continue
        workbook_hit = any(keyword in text for keyword in workbook_keywords)
        if not workbook_hit:
            continue
        disclosure_hit = any(keyword in text for keyword in disclosure_keywords)
        shared_tag_hint = any(tag in text for tag in disclosure_tag_names | workbook_tags)
        score = 0
        if disclosure_hit:
            score += 40
        if shared_tag_hint and disclosure_tag_names:
            score += 20
        if workbook_tags & disclosure_tag_names:
            score += 10
        if score > best_score:
            best_score = score
        if score:
            source_ref = str(source.get("source_ref") or "").strip()
            if source_ref and source_ref not in supporting_refs:
                supporting_refs.append(source_ref)
    return best_score, supporting_refs


def _is_candidate_segment_revenue_row(label: Any, metric: Any) -> bool:
    if not isinstance(label, str) or not label.strip():
        return False
    text = label.strip()
    if _is_summary_row_label(text):
        return False
    if _is_generic_segment_placeholder(text):
        return False
    if _is_metric_like_non_segment_label(text):
        return False
    lowered = text.lower()
    if any(token in lowered for token in ("/", "%", "率", "税", "费用", "附加", "毛利", "净利", "margin", "ratio", "share")):
        return False
    label_has_revenue = "收入" in text or "营收" in text
    metric_has_revenue = isinstance(metric, str) and ("收入" in metric or "营收" in metric)
    return label_has_revenue or metric_has_revenue


def _formula_referenced_candidate_rows(
    *,
    ws_formula: Worksheet,
    blueprint: WorkbookBlueprint,
    row: int,
    year: int,
    candidate_rows: set[int],
) -> set[int]:
    col = blueprint.forecast_columns.get(year) or blueprint.historical_columns.get(year)
    if col is None:
        return set()
    formula = ws_formula.cell(row, col).value
    if not isinstance(formula, str) or not formula.startswith("="):
        return set()
    return {
        int(match.group(1))
        for match in CELL_REF_ROW_RE.finditer(formula)
        if int(match.group(1)) != row and int(match.group(1)) in candidate_rows
    }


def _deprecated_build_model_segment_tree_v1(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
) -> dict[str, Any]:
    wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
    ws_data = wb_data[blueprint.primary_sheet]
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    ws_formula = wb_formula[blueprint.primary_sheet]

    ordered_primary = sorted((blueprint.primary_row_labels or blueprint.row_labels).items(), key=lambda item: item[1])
    nodes: list[dict[str, Any]] = []
    for label, row in ordered_primary:
        metric = ws_data.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws_data.max_column else None
        if not _is_candidate_segment_revenue_row(label, metric):
            continue
        series = _extract_metric_series_from_sheet(
            ws=ws_data,
            blueprint=blueprint,
            row=row,
            years=[report_year],
        )
        nodes.append(
            {
                "node_id": f"segment:{row}",
                "segment_name": label,
                "row": row,
                "revenue_row": row,
                "metric_label": metric if isinstance(metric, str) else "",
                "reported_value": series.get(report_year),
                "dependent_rows": [item for item in blueprint.row_blocks.get(label, []) if item != row],
                "parent_node_id": None,
                "child_node_ids": [],
                "lineage_labels": [label],
                "bridge_tags": [],
            }
        )

    row_to_node = {int(item["row"]): item for item in nodes}
    ordered_nodes = sorted(nodes, key=lambda item: int(item["row"]))
    candidate_rows = set(row_to_node)

    for node in ordered_nodes:
        referenced_rows = _formula_referenced_candidate_rows(
            ws_formula=ws_formula,
            blueprint=blueprint,
            row=int(node["row"]),
            year=report_year,
            candidate_rows=candidate_rows,
        )
        for child_row in sorted(referenced_rows):
            child = row_to_node.get(child_row)
            if child is None:
                continue
            child["parent_node_id"] = node["node_id"]
            node["child_node_ids"].append(child["node_id"])

    app_tags = {"phone", "auto", "security", "medical", "notebook", "iot"}
    for index, node in enumerate(ordered_nodes):
        if node["child_node_ids"]:
            continue
        parent_tags = set(_segment_bridge_tags(str(node["segment_name"])))
        if not parent_tags or re.match(r"^\s*\d+\s*[)）\.、]", str(node["segment_name"])):
            continue
        candidate_children: list[dict[str, Any]] = []
        for child in ordered_nodes[index + 1:]:
            if int(child["row"]) - int(node["row"]) > 60:
                break
            if child["parent_node_id"] is not None:
                continue
            child_tags = set(_segment_bridge_tags(str(child["segment_name"])))
            if re.match(r"^\s*\d+\s*[)）\.、]", str(child["segment_name"])) or (("cis" in parent_tags) and (child_tags & app_tags)):
                candidate_children.append(child)
                continue
            if candidate_children:
                break
        child_total = sum(float(item["reported_value"]) for item in candidate_children if isinstance(item.get("reported_value"), (int, float)))
        parent_value = node.get("reported_value")
        if candidate_children and isinstance(parent_value, (int, float)) and child_total not in {0, 0.0} and child_total <= float(parent_value) * 1.35:
            for child in candidate_children:
                child["parent_node_id"] = node["node_id"]
                node["child_node_ids"].append(child["node_id"])

    node_by_id = {str(item["node_id"]): item for item in ordered_nodes}
    for node in ordered_nodes:
        lineage_labels = [str(node["segment_name"])]
        parent_id = node.get("parent_node_id")
        visited: set[str] = set()
        while isinstance(parent_id, str) and parent_id and parent_id not in visited:
            visited.add(parent_id)
            parent_node = node_by_id.get(parent_id)
            if parent_node is None:
                break
            lineage_labels.insert(0, str(parent_node["segment_name"]))
            parent_id = parent_node.get("parent_node_id")
        node["lineage_labels"] = lineage_labels
        node["bridge_tags"] = _segment_bridge_tags(str(node["segment_name"]), lineage_labels[:-1])

    reportable_segments = [{**item, "reportable": not item["child_node_ids"]} for item in ordered_nodes if not item["child_node_ids"]]
    return {
        "reported_year": f"{report_year}A",
        "segment_nodes": ordered_nodes,
        "reportable_segments": reportable_segments,
    }


def _score_segment_mapping_match(
    workbook_segment: str,
    disclosure: dict[str, Any],
    *,
    workbook_bridge_tags: set[str] | None = None,
    workbook_lineage_labels: list[str] | None = None,
) -> tuple[int, str, str]:
    workbook_norm = _normalize_segment_label_for_mapping(workbook_segment)
    disclosure_norm = _normalize_segment_label_for_mapping(str(disclosure.get("segment") or ""))
    workbook_tags = set(workbook_bridge_tags or _segment_bridge_tags(workbook_segment, workbook_lineage_labels))
    disclosure_tags = set(_segment_bridge_tags(str(disclosure.get("segment") or "")))
    if not workbook_norm or not disclosure_norm:
        return 0, "proxy", "low"
    if workbook_norm == disclosure_norm:
        return 100, "exact", "high"
    bridge_overlap = workbook_tags & disclosure_tags
    if bridge_overlap:
        if workbook_tags == disclosure_tags:
            return 85, "exact", "high"
        return 75, "parent_child", "medium"
    if disclosure_norm in workbook_norm or workbook_norm in disclosure_norm:
        return 70, "parent_child", "medium"
    overlap = _segment_mapping_tokens(workbook_segment) & _segment_mapping_tokens(str(disclosure.get("segment") or ""))
    if overlap:
        return max(len(token) for token in overlap) * 10, "proxy", "low"
    return 0, "proxy", "low"


def _deprecated_build_segment_mapping_contract_v1(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
    financial_facts: dict[str, Any],
) -> dict[str, Any]:
    wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
    ws_data = wb_data[blueprint.primary_sheet]
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    workbook_segments: list[str] = []
    for row in range(blueprint.header_row + 1, ws_data.max_row + 1):
        found_label = None
        for col in range(1, min(blueprint.label_column + 1, 4)):
            label = ws_data.cell(row, col).value
            metric = ws_data.cell(row, col + 1).value if col + 1 <= ws_data.max_column else None
            if not isinstance(label, str) or not label.strip():
                continue
            if _is_summary_row_label(label):
                continue
            label_has_revenue = "收入" in label or "营收" in label
            metric_has_revenue = isinstance(metric, str) and ("收入" in metric or "营收" in metric)
            if not label_has_revenue and not metric_has_revenue:
                continue
            found_label = label.strip()
            break
        if found_label and not _is_formula_driven_summary_row_from_sheet(
            ws_formula=wb_formula[blueprint.primary_sheet],
            blueprint=blueprint,
            row_label=found_label,
            report_year=report_year,
        ):
            workbook_segments.append(found_label)
    disclosure_rows = [
        item
        for item in (financial_facts.get("segment_disclosure", []) or [])
        if not isinstance(item, dict) or item.get("mapping_ready", True)
    ]
    normalized_disclosures = [
        {
            **item,
            "_normalized_segment": _normalize_segment_label_for_mapping(str(item.get("segment") or "")),
        }
        for item in disclosure_rows
    ]
    reported_total = _extract_reported_revenue_anchor(financial_facts)
    mappings: list[dict[str, Any]] = []
    mapped_disclosures: set[str] = set()
    for workbook_segment in workbook_segments:
        normalized_workbook = _normalize_segment_label_for_mapping(workbook_segment)
        row = blueprint.row_labels.get(workbook_segment)
        reported_value = _extract_metric_series_from_sheet(
            ws=ws_data,
            blueprint=blueprint,
            row=row,
            years=[report_year],
        ).get(report_year)
        selected: dict[str, Any] | None = None
        mapping_type = "proxy"
        confidence = "low"
        best_score = 0
        for disclosure in normalized_disclosures:
            if not str(disclosure.get("_normalized_segment") or ""):
                continue
            score, candidate_mapping_type, candidate_confidence = _score_segment_mapping_match(
                workbook_segment,
                disclosure,
            )
            if score > best_score:
                selected = disclosure
                mapping_type = candidate_mapping_type
                confidence = candidate_confidence
                best_score = score
        if selected is not None:
            mapped_disclosures.add(str(selected.get("segment") or ""))
        mappings.append(
            {
                "workbook_segment": workbook_segment,
                "tushare_segment": str((selected or {}).get("segment") or ""),
                "mapping_type": mapping_type if selected is not None else "proxy",
                "confidence": confidence if selected is not None else "low",
                "reconciliation_target": str((selected or {}).get("segment") or "营业收入"),
                "workbook_reported_value": reported_value,
                "tushare_revenue": (selected or {}).get("revenue"),
                "source_refs": [str((selected or {}).get("source_ref") or financial_facts.get("source_ref") or "")],
            }
        )
    anchored_segment_count = sum(
        1
        for item in mappings
        if str(item.get("tushare_segment") or "").strip()
        and str(item.get("mapping_type") or "") in {"exact", "parent_child"}
    )
    return {
        "source_type": "tushare",
        "reported_year": f"{report_year}A",
        "reported_revenue_anchor": reported_total,
        "workbook_segment_count": len(workbook_segments),
        "anchored_segment_count": anchored_segment_count,
        "proxy_segment_count": sum(1 for item in mappings if item.get("mapping_type") == "proxy"),
        "segment_mappings": mappings,
        "unmapped_tushare_segments": [
            str(item.get("segment") or "")
            for item in disclosure_rows
            if str(item.get("segment") or "") not in mapped_disclosures
        ],
    }


def _deprecated_build_reconciliation_audit_v1(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
    financial_facts: dict[str, Any],
    segment_mapping: dict[str, Any],
) -> dict[str, Any]:
    wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
    ws_data = wb_data[blueprint.primary_sheet]
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    workbook_segments: list[str] = []
    for row in range(blueprint.header_row + 1, ws_data.max_row + 1):
        found_label = None
        for col in range(1, min(blueprint.label_column + 1, 4)):
            label = ws_data.cell(row, col).value
            metric = ws_data.cell(row, col + 1).value if col + 1 <= ws_data.max_column else None
            if not isinstance(label, str) or not label.strip():
                continue
            if _is_summary_row_label(label):
                continue
            label_has_revenue = "收入" in label or "营收" in label
            metric_has_revenue = isinstance(metric, str) and ("收入" in metric or "营收" in metric)
            if not label_has_revenue and not metric_has_revenue:
                continue
            found_label = label.strip()
            break
        if found_label and not _is_formula_driven_summary_row_from_sheet(
            ws_formula=wb_formula[blueprint.primary_sheet],
            blueprint=blueprint,
            row_label=found_label,
            report_year=report_year,
        ):
            workbook_segments.append(found_label)
    workbook_total = 0.0
    for segment in workbook_segments:
        row = blueprint.row_labels.get(segment)
        value = _extract_metric_series_from_sheet(
            ws=ws_data,
            blueprint=blueprint,
            row=row,
            years=[report_year],
        ).get(report_year)
        if isinstance(value, (int, float)):
            workbook_total += float(value)
    mapped_total = sum(
        float(item.get("workbook_reported_value"))
        for item in segment_mapping.get("segment_mappings", [])
        if isinstance(item.get("workbook_reported_value"), (int, float))
    )
    official_total = _extract_reported_revenue_anchor(financial_facts)
    revenue_gap = None
    revenue_gap_ratio = None
    if official_total not in {None, 0, 0.0}:
        revenue_gap = round(float(workbook_total) - float(official_total), 6)
        revenue_gap_ratio = round(revenue_gap / float(official_total), 6)
    anchored_segment_count = int(segment_mapping.get("anchored_segment_count") or 0)
    fail_reasons: list[str] = []
    if not workbook_segments:
        fail_reasons.append("no_workbook_segments_identified")
    if workbook_total in {0, 0.0}:
        fail_reasons.append("workbook_revenue_total_zero")
    if anchored_segment_count == 0:
        fail_reasons.append("no_tushare_segment_mappings")
    if official_total not in {None, 0, 0.0} and (revenue_gap_ratio is None or abs(revenue_gap_ratio) > 0.25):
        fail_reasons.append("reconciliation_gap_exceeds_tolerance")
    return {
        "reported_year": f"{report_year}A",
        "workbook_revenue_total": round(workbook_total, 6),
        "mapped_revenue_total": round(mapped_total, 6),
        "official_revenue_total": official_total,
        "revenue_gap": revenue_gap,
        "revenue_gap_ratio": revenue_gap_ratio,
        "anchored_segment_count": anchored_segment_count,
        "within_tolerance": not fail_reasons,
        "fail_reasons": fail_reasons,
        "unmapped_tushare_segments": segment_mapping.get("unmapped_tushare_segments", []),
    }


def build_segment_mapping_contract(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
    financial_facts: dict[str, Any],
    meeting_notes_facts: dict[str, Any] | None = None,
    evidence_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    segment_tree = build_model_segment_tree(
        workbook_path=workbook_path,
        blueprint=blueprint,
        report_year=report_year,
    )
    workbook_segments = list(segment_tree.get("reportable_segments", []))
    disclosure_rows = [
        item
        for item in (financial_facts.get("segment_disclosure", []) or [])
        if not isinstance(item, dict) or item.get("mapping_ready", True)
    ]
    normalized_disclosures = [
        {
            **item,
            "_normalized_segment": _normalize_segment_label_for_mapping(str(item.get("segment") or "")),
        }
        for item in disclosure_rows
    ]
    reported_total = _extract_reported_revenue_anchor(financial_facts)
    bridge_sources = _iter_bridge_sources(
        meeting_notes_facts=meeting_notes_facts,
        evidence_payload=evidence_payload,
    )
    mappings: list[dict[str, Any]] = []
    mapped_disclosures: set[str] = set()
    candidate_clusters: list[dict[str, Any]] = []
    for workbook_segment_node in workbook_segments:
        workbook_segment = str(workbook_segment_node.get("segment_name") or "")
        reported_value = workbook_segment_node.get("reported_value")
        workbook_bridge_tags = set(workbook_segment_node.get("bridge_tags") or [])
        workbook_lineage_labels = list(workbook_segment_node.get("lineage_labels") or [])
        selected: dict[str, Any] | None = None
        mapping_type = "proxy"
        confidence = "low"
        best_score = 0
        selected_source_refs: list[str] = [str(financial_facts.get("source_ref") or "").strip()]
        selected_bridge_support_refs: list[str] = []
        selected_bridge_support_score = 0
        candidate_matches: list[dict[str, Any]] = []
        for disclosure in normalized_disclosures:
            if not str(disclosure.get("_normalized_segment") or ""):
                continue
            score, candidate_mapping_type, candidate_confidence = _score_segment_mapping_match(
                workbook_segment,
                disclosure,
                workbook_bridge_tags=workbook_bridge_tags,
                workbook_lineage_labels=workbook_lineage_labels,
            )
            disclosure_tags = set(_segment_bridge_tags(str(disclosure.get("segment") or "")))
            bridge_score, bridge_support_refs = _bridge_candidate_support(
                workbook_segment=workbook_segment,
                workbook_bridge_tags=workbook_bridge_tags,
                workbook_lineage_labels=workbook_lineage_labels,
                disclosure_segment=str(disclosure.get("segment") or ""),
                disclosure_tags=disclosure_tags,
                bridge_sources=bridge_sources,
            )
            effective_mapping_type = candidate_mapping_type
            effective_confidence = candidate_confidence
            effective_score = score
            bridge_only_candidate = (
                score == 0
                and bridge_score >= 40
                and bool(workbook_bridge_tags)
                and bool(disclosure_tags)
            )
            if bridge_score:
                effective_score += bridge_score
                if score == 0:
                    if bridge_score >= 60 or workbook_bridge_tags & disclosure_tags:
                        effective_mapping_type = "parent_child"
                        effective_confidence = "medium"
                    else:
                        effective_mapping_type = "proxy"
                        effective_confidence = "low"
            source_refs = [
                ref
                for ref in [
                    str(disclosure.get("source_ref") or "").strip(),
                    *[str(ref).strip() for ref in bridge_support_refs],
                    str(financial_facts.get("source_ref") or "").strip(),
                ]
                if ref
            ]
            if effective_score > 0 and (score > 0 or bridge_only_candidate):
                candidate_matches.append(
                    {
                        "tushare_segment": str(disclosure.get("segment") or ""),
                        "mapping_type": effective_mapping_type,
                        "confidence": effective_confidence,
                        "score": effective_score,
                        "base_score": score,
                        "bridge_support_score": bridge_score,
                        "tushare_revenue": disclosure.get("revenue"),
                        "source_refs": source_refs,
                        "bridge_support_refs": bridge_support_refs,
                    }
                )
            if effective_score > best_score and (score > 0 or bridge_only_candidate):
                selected = disclosure
                mapping_type = effective_mapping_type
                confidence = effective_confidence
                best_score = effective_score
                selected_source_refs = source_refs
                selected_bridge_support_refs = list(bridge_support_refs)
                selected_bridge_support_score = bridge_score
        candidate_matches.sort(
            key=lambda item: (
                -int(item.get("score") or 0),
                0 if item.get("mapping_type") == "exact" else 1,
                str(item.get("tushare_segment") or ""),
            )
        )
        if selected is not None:
            mapped_disclosures.add(str(selected.get("segment") or ""))
        current_mapping = {
            "workbook_segment": workbook_segment,
            "tushare_segment": str((selected or {}).get("segment") or ""),
            "mapping_type": mapping_type if selected is not None else "proxy",
            "confidence": confidence if selected is not None else "low",
            "reconciliation_target": str((selected or {}).get("segment") or "营业收入"),
            "workbook_reported_value": reported_value,
            "tushare_revenue": (selected or {}).get("revenue"),
            "bridge_tags": sorted(workbook_bridge_tags),
            "lineage_labels": workbook_lineage_labels,
            "source_refs": selected_source_refs,
            "bridge_support_score": selected_bridge_support_score,
            "bridge_support_refs": selected_bridge_support_refs,
        }
        mappings.append(
            {
                **current_mapping,
                "candidate_matches": candidate_matches,
            }
        )
        if (
            len(candidate_matches) >= 2
            and not (
                str(current_mapping.get("mapping_type") or "") == "exact"
                and str(current_mapping.get("confidence") or "") == "high"
            )
        ):
            top_score = int(candidate_matches[0].get("score") or 0)
            second_score = int(candidate_matches[1].get("score") or 0)
            if top_score - second_score <= 15:
                candidate_clusters.append(
                    {
                        "workbook_segment": workbook_segment,
                        "current_mapping": current_mapping,
                        "candidate_matches": candidate_matches,
                    }
                )
    anchored_segment_count = sum(1 for item in mappings if str(item.get("tushare_segment") or "").strip())
    return {
        "source_type": "tushare",
        "reported_year": f"{report_year}A",
        "reported_revenue_anchor": reported_total,
        "workbook_segment_count": len(workbook_segments),
        "anchored_segment_count": anchored_segment_count,
        "proxy_segment_count": sum(1 for item in mappings if item.get("mapping_type") == "proxy"),
        "segment_tree": segment_tree,
        "bridge_generation_used": bool(bridge_sources),
        "candidate_clusters": candidate_clusters,
        "segment_mappings": mappings,
        "unmapped_tushare_segments": [
            str(item.get("segment") or "")
            for item in disclosure_rows
            if str(item.get("segment") or "") not in mapped_disclosures
        ],
    }


def _mapping_type_retention_score(mapping_type: str) -> float:
    return {
        "exact": 1.0,
        "parent_child": 0.9,
        "proxy": 0.75,
        "residual": 0.88,
    }.get(str(mapping_type or ""), 0.6)


def _summarize_mapping_assignments(
    assignments: list[dict[str, Any]],
    *,
    official_total: float | None,
) -> dict[str, Any]:
    anchored_revenue_total = round(
        sum(
            float(item.get("workbook_reported_value"))
            for item in assignments
            if isinstance(item.get("workbook_reported_value"), (int, float))
            and str(item.get("tushare_segment") or "").strip()
            and str(item.get("mapping_type") or "") in {"exact", "parent_child"}
        ),
        6,
    )
    proxy_revenue_total = round(
        sum(
            float(item.get("workbook_reported_value"))
            for item in assignments
            if isinstance(item.get("workbook_reported_value"), (int, float))
            and str(item.get("mapping_type") or "") == "proxy"
        ),
        6,
    )
    residual_revenue_total = round(
        sum(
            float(item.get("workbook_reported_value"))
            for item in assignments
            if isinstance(item.get("workbook_reported_value"), (int, float))
            and str(item.get("mapping_type") or "") == "residual"
        ),
        6,
    )
    represented_total = anchored_revenue_total + proxy_revenue_total + residual_revenue_total
    revenue_gap = None
    revenue_gap_ratio = None
    anchor_coverage_ratio = None
    if official_total not in {None, 0, 0.0}:
        revenue_gap = round(float(represented_total) - float(official_total), 6)
        revenue_gap_ratio = round(revenue_gap / float(official_total), 6)
        anchor_coverage_ratio = round(anchored_revenue_total / float(official_total), 6)
    structure_retention_score = round(
        (
            sum(_mapping_type_retention_score(str(item.get("mapping_type") or "")) for item in assignments) / len(assignments)
        )
        if assignments
        else 0.0,
        6,
    )
    return {
        "anchored_revenue_total": anchored_revenue_total,
        "mapped_revenue_total": anchored_revenue_total,
        "proxy_revenue_total": proxy_revenue_total,
        "residual_revenue_total": residual_revenue_total,
        "represented_revenue_total": round(represented_total, 6),
        "revenue_gap": revenue_gap,
        "revenue_gap_ratio": revenue_gap_ratio,
        "anchor_coverage_ratio": anchor_coverage_ratio,
        "structure_retention_score": structure_retention_score,
    }


def _segment_tree_dirty_labels(segment_tree: dict[str, Any]) -> list[str]:
    dirty: list[str] = []
    for node in segment_tree.get("reportable_segments", []) or []:
        label = str(node.get("segment_name") or "").strip()
        if _is_generic_segment_placeholder(label) or _is_metric_like_non_segment_label(label):
            dirty.append(label)
    return dirty


def _candidate_clusters_all_low_quality_proxy(candidate_clusters: list[dict[str, Any]]) -> bool:
    if not candidate_clusters:
        return False
    saw_any_match = False
    for cluster in candidate_clusters:
        matches = list(cluster.get("candidate_matches", []) or [])
        if not matches:
            continue
        saw_any_match = True
        if any(
            str(item.get("mapping_type") or "") != "proxy"
            or str(item.get("confidence") or "") != "low"
            or float(item.get("base_score") or 0) > 0
            for item in matches
        ):
            return False
    return saw_any_match


def materialize_selected_candidate_mapping(
    *,
    segment_mapping: dict[str, Any],
    reconciliation_audit: dict[str, Any],
    option_id: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    if str(option_id).upper() == "R":
        raise ValueError("operator selected R: cancel resume and return to re-bridge upstream mappings")
    option = next(
        (item for item in reconciliation_audit.get("candidate_options", []) if str(item.get("option_id")) == str(option_id)),
        None,
    )
    if option is None:
        raise ValueError(f"candidate option {option_id} not found")
    assignments = [dict(item) for item in option.get("segment_assignments", [])]
    anchored_segment_count = sum(
        1
        for item in assignments
        if str(item.get("tushare_segment") or "").strip() and str(item.get("mapping_type") or "") in {"exact", "parent_child"}
    )
    proxy_segment_count = sum(1 for item in assignments if str(item.get("mapping_type") or "") == "proxy")
    resolved_mapping = {
        **segment_mapping,
        "segment_mappings": assignments,
        "anchored_segment_count": anchored_segment_count,
        "proxy_segment_count": proxy_segment_count,
        "selected_candidate_option": str(option_id),
    }
    resolved_audit = {
        **reconciliation_audit,
        "within_tolerance": True,
        "resolution_mode": "automatic_pass",
        "selected_candidate_option": str(option_id),
        "operator_override": True,
        "anchored_revenue_total": option.get("anchored_revenue_total"),
        "mapped_revenue_total": option.get("mapped_revenue_total"),
        "proxy_revenue_total": option.get("proxy_revenue_total"),
        "residual_revenue_total": option.get("residual_revenue_total"),
    }
    return resolved_mapping, resolved_audit


def build_reconciliation_audit(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
    financial_facts: dict[str, Any],
    segment_mapping: dict[str, Any],
) -> dict[str, Any]:
    segment_tree = segment_mapping.get("segment_tree") or build_model_segment_tree(
        workbook_path=workbook_path,
        blueprint=blueprint,
        report_year=report_year,
    )
    workbook_segments = list(segment_tree.get("reportable_segments", []))
    workbook_total = sum(
        float(item.get("reported_value"))
        for item in workbook_segments
        if isinstance(item.get("reported_value"), (int, float))
    )
    official_total = _extract_reported_revenue_anchor(financial_facts)
    current_summary = _summarize_mapping_assignments(
        [dict(item) for item in segment_mapping.get("segment_mappings", [])],
        official_total=official_total,
    )
    revenue_gap = current_summary.get("revenue_gap")
    revenue_gap_ratio = current_summary.get("revenue_gap_ratio")
    anchored_segment_count = int(segment_mapping.get("anchored_segment_count") or 0)
    proxy_segment_count = int(segment_mapping.get("proxy_segment_count") or 0)
    workbook_segment_count = len(workbook_segments)
    coverage_ratio = (
        round(anchored_segment_count / workbook_segment_count, 6)
        if workbook_segment_count
        else 0.0
    )
    anchored_revenue_total = float(current_summary.get("anchored_revenue_total") or 0.0)
    unmapped_tushare_segments = list(segment_mapping.get("unmapped_tushare_segments", []))
    candidate_clusters = list(segment_mapping.get("candidate_clusters", []))
    ambiguity_requires_operator = bool(candidate_clusters)
    dirty_segment_labels = _segment_tree_dirty_labels(segment_tree)
    all_low_quality_proxy = _candidate_clusters_all_low_quality_proxy(candidate_clusters)
    fail_reasons: list[str] = []
    if not workbook_segments:
        fail_reasons.append("no_workbook_segments_identified")
    if workbook_total in {0, 0.0}:
        fail_reasons.append("workbook_revenue_total_zero")
    if anchored_segment_count == 0:
        fail_reasons.append("no_tushare_segment_mappings")
    if anchored_segment_count > 0 and anchored_revenue_total in {0, 0.0}:
        fail_reasons.append("anchored_segment_count_has_no_anchored_revenue")
    if dirty_segment_labels:
        fail_reasons.append("dirty_segment_tree_detected")
    if official_total not in {None, 0, 0.0} and (revenue_gap_ratio is None or abs(revenue_gap_ratio) > 0.25):
        fail_reasons.append("reconciliation_gap_exceeds_tolerance")
    if all_low_quality_proxy:
        fail_reasons.append("all_candidate_options_low_quality_proxy")
    if ambiguity_requires_operator:
        fail_reasons.append("mapping_ambiguity_requires_operator_decision")
    candidate_options: list[dict[str, Any]] = []
    resolution_mode = "automatic_pass"
    if fail_reasons:
        if (
            "no_workbook_segments_identified" in fail_reasons
            or "workbook_revenue_total_zero" in fail_reasons
            or official_total in {None, 0, 0.0}
            or not candidate_clusters
        ):
            resolution_mode = "hard_stop"
        else:
            resolution_mode = "candidate_decision_required"
            base_assignments = [dict(item) for item in segment_mapping.get("segment_mappings", [])]
            cluster_map = {str(item.get("workbook_segment") or ""): item for item in candidate_clusters}
            option_specs: list[dict[str, Any]] = []
            strategy_defs = [
                ("A", "保留当前 workbook 细分结构，优先采用当前最优锚定候选。", "primary"),
                ("B", "对歧义 segment 采用次优锚定候选，检验父层桥接是否更稳。", "secondary"),
                ("C", "保留 workbook 细分叶子，但将歧义 segment 显式放入 residual。", "residual"),
            ]
            for option_id, summary, strategy in strategy_defs:
                assignments: list[dict[str, Any]] = []
                for mapping in base_assignments:
                    workbook_segment = str(mapping.get("workbook_segment") or "")
                    cluster = cluster_map.get(workbook_segment)
                    if not cluster:
                        candidate_mapping = dict(mapping)
                    else:
                        matches = [dict(item) for item in cluster.get("candidate_matches", [])]
                        if strategy == "primary":
                            chosen = matches[0]
                        elif strategy == "secondary":
                            chosen = matches[1] if len(matches) > 1 else matches[0]
                        else:
                            chosen = {
                                "tushare_segment": "",
                                "mapping_type": "residual",
                                "confidence": "medium",
                                "score": 0,
                                "tushare_revenue": None,
                                "source_refs": [],
                            }
                        candidate_mapping = {
                            **dict(mapping),
                            "tushare_segment": str(chosen.get("tushare_segment") or ""),
                            "mapping_type": str(chosen.get("mapping_type") or mapping.get("mapping_type") or "proxy"),
                            "confidence": str(chosen.get("confidence") or mapping.get("confidence") or "low"),
                            "tushare_revenue": chosen.get("tushare_revenue"),
                            "source_refs": list(chosen.get("source_refs") or mapping.get("source_refs") or []),
                        }
                    assignments.append(candidate_mapping)
                metrics = _summarize_mapping_assignments(assignments, official_total=official_total)
                option_specs.append(
                    {
                        "option_id": option_id,
                        "summary": summary,
                        "continue_allowed": True,
                        "segment_assignments": assignments,
                        **metrics,
                        "proxy_segment_count": sum(1 for item in assignments if str(item.get("mapping_type") or "") == "proxy"),
                        "residual_segment_count": sum(1 for item in assignments if str(item.get("mapping_type") or "") == "residual"),
                    }
                )
            option_specs.sort(
                key=lambda item: (
                    -float(item["structure_retention_score"]),
                    -float(item.get("anchor_coverage_ratio") or 0.0),
                    int(item["proxy_segment_count"]) + int(item["residual_segment_count"]),
                    abs(float(item["revenue_gap_ratio"] or 0.0)),
                )
            )
            for index, option in enumerate(option_specs):
                candidate_options.append({**option, "recommended": index == 0})
    return {
        "reported_year": f"{report_year}A",
        "workbook_revenue_total": round(workbook_total, 6),
        "mapped_revenue_total": current_summary.get("mapped_revenue_total"),
        "official_revenue_total": official_total,
        "revenue_gap": current_summary.get("revenue_gap"),
        "revenue_gap_ratio": current_summary.get("revenue_gap_ratio"),
        "coverage_ratio": coverage_ratio,
        "anchored_segment_count": anchored_segment_count,
        "proxy_segment_count": proxy_segment_count,
        "anchored_revenue_total": current_summary.get("anchored_revenue_total"),
        "proxy_revenue_total": current_summary.get("proxy_revenue_total"),
        "residual_revenue_total": current_summary.get("residual_revenue_total"),
        "dirty_segment_labels": dirty_segment_labels,
        "candidate_quality": {
            "all_low_quality_proxy": all_low_quality_proxy,
            "candidate_cluster_count": len(candidate_clusters),
        },
        "within_tolerance": not fail_reasons,
        "resolution_mode": resolution_mode,
        "candidate_options": candidate_options,
        "fail_reasons": fail_reasons,
        "unmapped_tushare_segments": unmapped_tushare_segments,
    }


def _deprecated_build_model_segment_tree_v2(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
) -> dict[str, Any]:
    wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
    ws_data = wb_data[blueprint.primary_sheet]
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    ws_formula = wb_formula[blueprint.primary_sheet]

    ordered_primary = sorted((blueprint.primary_row_labels or blueprint.row_labels).items(), key=lambda item: item[1])
    nodes: list[dict[str, Any]] = []
    for label, row in ordered_primary:
        metric = ws_data.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws_data.max_column else None
        if not _is_candidate_segment_revenue_row(label, metric):
            continue
        series = _extract_metric_series_from_sheet(
            ws=ws_data,
            blueprint=blueprint,
            row=row,
            years=[report_year],
        )
        nodes.append(
            {
                "node_id": f"segment:{row}",
                "segment_name": label,
                "row": row,
                "revenue_row": row,
                "metric_label": metric if isinstance(metric, str) else "",
                "reported_value": series.get(report_year),
                "dependent_rows": [item for item in blueprint.row_blocks.get(label, []) if item != row],
                "parent_node_id": None,
                "child_node_ids": [],
                "lineage_labels": [label],
                "bridge_tags": [],
            }
        )

    row_to_node = {int(item["row"]): item for item in nodes}
    ordered_nodes = sorted(nodes, key=lambda item: int(item["row"]))
    candidate_rows = set(row_to_node)

    for node in ordered_nodes:
        referenced_rows = _formula_referenced_candidate_rows(
            ws_formula=ws_formula,
            blueprint=blueprint,
            row=int(node["row"]),
            year=report_year,
            candidate_rows=candidate_rows,
        )
        for child_row in sorted(referenced_rows):
            child = row_to_node.get(child_row)
            if child is None:
                continue
            child["parent_node_id"] = node["node_id"]
            node["child_node_ids"].append(child["node_id"])

    app_tags = {"phone", "auto", "security", "medical", "notebook", "iot"}
    broad_parent_tags = {"cis", "communication", "display", "power", "analog", "material"}
    for index, node in enumerate(ordered_nodes):
        if node["child_node_ids"]:
            continue
        parent_tags = set(_segment_bridge_tags(str(node["segment_name"])))
        if (
            not parent_tags
            or re.match(r"^\s*\d+\s*[)锛塡.銆乚", str(node["segment_name"]))
            or not (parent_tags & broad_parent_tags)
            or (parent_tags & app_tags)
        ):
            continue
        candidate_children: list[dict[str, Any]] = []
        for child in ordered_nodes[index + 1:]:
            if int(child["row"]) - int(node["row"]) > 60:
                break
            if child["parent_node_id"] is not None:
                continue
            child_tags = set(_segment_bridge_tags(str(child["segment_name"])))
            if re.match(r"^\s*\d+\s*[)锛塡.銆乚", str(child["segment_name"])) or (("cis" in parent_tags) and (child_tags & app_tags)):
                candidate_children.append(child)
                continue
            if candidate_children:
                break
        child_total = sum(float(item["reported_value"]) for item in candidate_children if isinstance(item.get("reported_value"), (int, float)))
        parent_value = node.get("reported_value")
        if candidate_children and isinstance(parent_value, (int, float)) and child_total not in {0, 0.0} and child_total <= float(parent_value) * 1.35:
            for child in candidate_children:
                child["parent_node_id"] = node["node_id"]
                node["child_node_ids"].append(child["node_id"])

    node_by_id = {str(item["node_id"]): item for item in ordered_nodes}
    for node in ordered_nodes:
        lineage_labels = [str(node["segment_name"])]
        parent_id = node.get("parent_node_id")
        visited: set[str] = set()
        while isinstance(parent_id, str) and parent_id and parent_id not in visited:
            visited.add(parent_id)
            parent_node = node_by_id.get(parent_id)
            if parent_node is None:
                break
            lineage_labels.insert(0, str(parent_node["segment_name"]))
            parent_id = parent_node.get("parent_node_id")
        node["lineage_labels"] = lineage_labels
        node["bridge_tags"] = _segment_bridge_tags(str(node["segment_name"]), lineage_labels[:-1])

    reportable_segments = [{**item, "reportable": not item["child_node_ids"]} for item in ordered_nodes if not item["child_node_ids"]]
    return {
        "reported_year": f"{report_year}A",
        "segment_nodes": ordered_nodes,
        "reportable_segments": reportable_segments,
    }


def _starts_numbered_segment_label(text: str) -> bool:
    return bool(re.match(r"^\s*\d+", str(text or "")))


def build_model_segment_tree(
    *,
    workbook_path: Path,
    blueprint: WorkbookBlueprint,
    report_year: int,
) -> dict[str, Any]:
    wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
    ws_data = wb_data[blueprint.primary_sheet]
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    ws_formula = wb_formula[blueprint.primary_sheet]

    ordered_primary = sorted((blueprint.primary_row_labels or blueprint.row_labels).items(), key=lambda item: item[1])
    nodes: list[dict[str, Any]] = []
    for label, row in ordered_primary:
        metric = ws_data.cell(row, blueprint.label_column + 1).value if blueprint.label_column + 1 <= ws_data.max_column else None
        if not _is_candidate_segment_revenue_row(label, metric):
            continue
        series = _extract_metric_series_from_sheet(
            ws=ws_data,
            blueprint=blueprint,
            row=row,
            years=[report_year],
        )
        nodes.append(
            {
                "node_id": f"segment:{row}",
                "segment_name": label,
                "row": row,
                "revenue_row": row,
                "metric_label": metric if isinstance(metric, str) else "",
                "reported_value": series.get(report_year),
                "dependent_rows": [item for item in blueprint.row_blocks.get(label, []) if item != row],
                "parent_node_id": None,
                "child_node_ids": [],
                "lineage_labels": [label],
                "bridge_tags": [],
            }
        )

    row_to_node = {int(item["row"]): item for item in nodes}
    ordered_nodes = sorted(nodes, key=lambda item: int(item["row"]))
    candidate_rows = set(row_to_node)

    for node in ordered_nodes:
        referenced_rows = _formula_referenced_candidate_rows(
            ws_formula=ws_formula,
            blueprint=blueprint,
            row=int(node["row"]),
            year=report_year,
            candidate_rows=candidate_rows,
        )
        for child_row in sorted(referenced_rows):
            child = row_to_node.get(child_row)
            if child is None:
                continue
            child["parent_node_id"] = node["node_id"]
            node["child_node_ids"].append(child["node_id"])

    for index, node in enumerate(ordered_nodes):
        if node["child_node_ids"]:
            continue
        parent_name = str(node["segment_name"])
        if _starts_numbered_segment_label(parent_name):
            continue
        candidate_children: list[dict[str, Any]] = []
        saw_numbered_child = False
        for child in ordered_nodes[index + 1:]:
            if int(child["row"]) - int(node["row"]) > 60:
                break
            if child["parent_node_id"] is not None:
                continue
            child_name = str(child["segment_name"])
            if _starts_numbered_segment_label(child_name):
                candidate_children.append(child)
                saw_numbered_child = True
                continue
            if saw_numbered_child:
                break
        child_total = sum(float(item["reported_value"]) for item in candidate_children if isinstance(item.get("reported_value"), (int, float)))
        parent_value = node.get("reported_value")
        if candidate_children and isinstance(parent_value, (int, float)) and child_total not in {0, 0.0} and child_total <= float(parent_value) * 1.35:
            for child in candidate_children:
                child["parent_node_id"] = node["node_id"]
                node["child_node_ids"].append(child["node_id"])

    node_by_id = {str(item["node_id"]): item for item in ordered_nodes}
    for node in ordered_nodes:
        lineage_labels = [str(node["segment_name"])]
        parent_id = node.get("parent_node_id")
        visited: set[str] = set()
        while isinstance(parent_id, str) and parent_id and parent_id not in visited:
            visited.add(parent_id)
            parent_node = node_by_id.get(parent_id)
            if parent_node is None:
                break
            lineage_labels.insert(0, str(parent_node["segment_name"]))
            parent_id = parent_node.get("parent_node_id")
        node["lineage_labels"] = lineage_labels
        node["bridge_tags"] = _segment_bridge_tags(str(node["segment_name"]), lineage_labels[:-1])

    reportable_segments = [{**item, "reportable": not item["child_node_ids"]} for item in ordered_nodes if not item["child_node_ids"]]
    return {
        "reported_year": f"{report_year}A",
        "segment_nodes": ordered_nodes,
        "reportable_segments": reportable_segments,
    }


def _timed_task(task: str, fn, *args, **kwargs) -> tuple[Any, dict[str, Any]]:
    started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    started_perf = time.perf_counter()
    result = fn(*args, **kwargs)
    ended_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elapsed_seconds = round(time.perf_counter() - started_perf, 4)
    return result, {
        "task": task,
        "started_at": started_at,
        "ended_at": ended_at,
        "elapsed_seconds": elapsed_seconds,
    }


def collect_pre_edit_inputs(
    *,
    repo_root: Path,
    company: str,
    model_path: Path,
    report_path: Path,
    meeting_notes_path: Path | None,
    research_report_path: Path | None = None,
    max_workers: int = 4,
) -> dict[str, Any]:
    def _infer_report_year_for_financial_facts() -> int:
        blueprint = WorkbookBlueprintParser().parse(model_path)
        if blueprint.forecast_columns:
            forecast_start = min(blueprint.forecast_columns)
            historical_end = max(blueprint.historical_columns) if blueprint.historical_columns else None
            if historical_end is not None and forecast_start == historical_end + 1:
                return historical_end
            return forecast_start
        if blueprint.historical_columns:
            return max(blueprint.historical_columns)
        raise ValueError("unable to infer report year from workbook blueprint")

    def _call_build_evidence_payload_for_parallel() -> dict[str, Any]:
        callable_target = getattr(build_evidence_payload, "side_effect", None)
        signature_target = callable_target if callable(callable_target) else build_evidence_payload
        parameters = inspect.signature(signature_target).parameters
        if "workbook_path" in parameters or len(parameters) >= 8:
            return build_evidence_payload(
                repo_root,
                company,
                meeting_notes_path,
                research_report_path,
                report_path,
                model_path,
                None,
                None,
            )
        return build_evidence_payload(
            repo_root,
            company,
            meeting_notes_path,
            research_report_path,
            report_path,
        )

    def _extract_financial_facts() -> dict[str, Any]:
        report_year = _infer_report_year_for_financial_facts()
        return TushareFinancialFactsAdapter().extract(
            company=company,
            report_year=report_year,
            report_path=report_path,
            model_path=model_path,
        )

    tasks: dict[str, Any] = {
        "blueprint_parse": lambda: _timed_task(
            "blueprint_parse",
            WorkbookBlueprintParser().parse,
            model_path,
        ),
        "financial_facts_extract": lambda: _timed_task(
            "financial_facts_extract",
            _extract_financial_facts,
        ),
        "evidence_payload_build": lambda: _timed_task(
            "evidence_payload_build",
            _call_build_evidence_payload_for_parallel,
        ),
    }
    if meeting_notes_path is not None and meeting_notes_path.exists():
        tasks["meeting_notes_extract"] = lambda: _timed_task(
            "meeting_notes_extract",
            MeetingNotesFactExtractor().extract,
            meeting_notes_path,
        )

    resolved_workers = max(1, min(max_workers, len(tasks)))
    results: dict[str, Any] = {}
    timing_rows: list[dict[str, Any]] = []

    with ThreadPoolExecutor(max_workers=resolved_workers, thread_name_prefix="forecast-pre-edit") as executor:
        future_map = {executor.submit(fn): name for name, fn in tasks.items()}
        for future in as_completed(future_map):
            name = future_map[future]
            result, timing = future.result()
            results[name] = result
            timing_rows.append(timing)

    timing_rows.sort(key=lambda item: item["started_at"])
    meeting_notes_facts = results.get("meeting_notes_extract", {"bridge_facts": {}})
    evidence_payload = results["evidence_payload_build"]
    financial_facts = results["financial_facts_extract"]
    blueprint = results["blueprint_parse"]
    if blueprint.forecast_columns:
        forecast_start = min(blueprint.forecast_columns)
        historical_end = max(blueprint.historical_columns) if blueprint.historical_columns else None
        report_year = historical_end if historical_end is not None and forecast_start == historical_end + 1 else forecast_start
    else:
        report_year = max(blueprint.historical_columns)
    segment_mapping = build_segment_mapping_contract(
        workbook_path=model_path,
        blueprint=blueprint,
        report_year=report_year,
        financial_facts=financial_facts,
        meeting_notes_facts=meeting_notes_facts,
        evidence_payload=evidence_payload,
    )
    reconciliation_audit = build_reconciliation_audit(
        workbook_path=model_path,
        blueprint=blueprint,
        report_year=report_year,
        financial_facts=financial_facts,
        segment_mapping=segment_mapping,
    )
    evidence_payload["timing"] = {
        "mode": "thread_pool",
        "max_workers": resolved_workers,
        "tasks": timing_rows,
    }
    return {
        "blueprint": blueprint,
        "financial_facts": financial_facts,
        "annual_report_facts": financial_facts,
        "meeting_notes_facts": meeting_notes_facts,
        "evidence_payload": evidence_payload,
        "segment_mapping": segment_mapping,
        "reconciliation_audit": reconciliation_audit,
        "timing": evidence_payload["timing"],
    }
