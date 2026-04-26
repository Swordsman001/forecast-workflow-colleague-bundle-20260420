# -*- coding: utf-8 -*-
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl

from .build_cell_instructions import build_cell_instructions, load_json, validate_forecast_basis, validate_workbook_map
from .contract_validators import validate_cell_instructions_payload, validate_evidence_store_payload, validate_patch_log_payload
from .patch_executor import execute_patch_from_instructions
from .verification import verify_contract_patch
from .artifact_utils import load_jsonl, sha256_file, sha256_json, sha256_jsonl


def normalize_evidence_store(
    evidence_store: list[dict[str, Any]],
    *,
    company: str | None,
    default_confidence: str = "medium",
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for line in evidence_store:
        current = dict(line)
        if company and not current.get("company"):
            current["company"] = company
        if not current.get("confidence"):
            current["confidence"] = default_confidence
        normalized.append(current)
    return normalized


def _sheet_scalar(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, list):
        return "; ".join("" if item is None else str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def _coerce_str_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if item is not None]
    if value is None:
        return []
    return [str(value)]


def _render_source_refs(refs: Any, evidence_lookup: dict[str, dict[str, Any]]) -> str:
    rendered: list[str] = []
    for ref in _coerce_str_list(refs):
        evidence = evidence_lookup.get(ref)
        if evidence is None:
            rendered.append(ref)
            continue
        parts = [
            str(evidence.get("source_file") or "").strip(),
            str(evidence.get("page_or_line") or "").strip(),
            str(evidence.get("text_summary") or "").strip(),
        ]
        rendered.append(" | ".join(part for part in parts if part))
    return "; ".join(rendered)


def _materialize_forecast_basis_sheet(
    *,
    workbook_path: Path,
    forecast_basis: dict[str, Any],
    evidence_store: list[dict[str, Any]],
) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    if "Forecast Basis" in wb.sheetnames:
        del wb["Forecast Basis"]
    ws = wb.create_sheet("Forecast Basis")
    evidence_lookup = {
        str(item.get("fact_id")): item
        for item in evidence_store
        if item.get("fact_id") is not None
    }

    ws["A1"] = "预测依据"
    ws["A2"] = f"公司：{_sheet_scalar(forecast_basis.get('company'))}"
    ws["D2"] = f"截止日期：{_sheet_scalar(forecast_basis.get('cutoff_date'))}"
    ws["G2"] = f"报告年份：{_sheet_scalar(forecast_basis.get('reported_year'))}"
    ws["I2"] = f"目标区间：{', '.join(forecast_basis.get('target_window', []))}"

    headers = ["分类", "项目", "年度", "数值", "驱动形式", "依据摘要", "来源", "证据标签", "风险提示", "置信度"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(3, idx).value = header

    current_row = 4
    for fact in forecast_basis.get("facts", []):
        ws[f"A{current_row}"] = "已知事实"
        ws[f"B{current_row}"] = _sheet_scalar(fact.get("key") or fact.get("metric"))
        ws[f"C{current_row}"] = _sheet_scalar(fact.get("year"))
        ws[f"D{current_row}"] = _sheet_scalar(fact.get("value"))
        ws[f"E{current_row}"] = "actual"
        ws[f"F{current_row}"] = "已披露实际值"
        ws[f"G{current_row}"] = _sheet_scalar(_render_source_refs(fact.get("source_ref"), evidence_lookup))
        ws[f"H{current_row}"] = _sheet_scalar("; ".join(_coerce_str_list(fact.get("source_ref") or fact.get("evidence_refs"))))
        ws[f"I{current_row}"] = _sheet_scalar(fact.get("review_flag"))
        ws[f"J{current_row}"] = _sheet_scalar(fact.get("confidence"))
        current_row += 1

    for card in forecast_basis.get("segment_assumption_cards", []):
        evidence_refs: list[str] = []
        for logic_key in ("volume_logic", "asp_logic", "share_logic", "margin_logic"):
            logic_payload = card.get(logic_key) or {}
            refs = logic_payload.get("evidence_refs") if isinstance(logic_payload, dict) else None
            if isinstance(refs, list):
                evidence_refs.extend(str(item) for item in refs if item is not None)
            elif refs:
                evidence_refs.append(str(refs))

        basis_summary_parts = []
        for label, logic_key in (
            ("量", "volume_logic"),
            ("价", "asp_logic"),
            ("份额", "share_logic"),
            ("毛利率", "margin_logic"),
        ):
            logic_payload = card.get(logic_key) or {}
            mechanism = logic_payload.get("mechanism") if isinstance(logic_payload, dict) else None
            if mechanism:
                basis_summary_parts.append(f"{label}：{mechanism}")

        risk_notes = []
        kill_conditions = card.get("kill_conditions") or []
        weak_assumptions = card.get("weak_assumptions") or []
        if kill_conditions:
            risk_notes.append("kill: " + "；".join(str(item) for item in kill_conditions))
        if weak_assumptions:
            risk_notes.append("weak: " + "；".join(str(item) for item in weak_assumptions))

        ws[f"A{current_row}"] = "预测假设"
        ws[f"B{current_row}"] = _sheet_scalar(card.get("segment"))
        ws[f"C{current_row}"] = _sheet_scalar(card.get("year"))
        ws[f"D{current_row}"] = _sheet_scalar(card.get("value"))
        ws[f"E{current_row}"] = _sheet_scalar(card.get("driver_form"))
        ws[f"F{current_row}"] = _sheet_scalar(" | ".join(basis_summary_parts))
        ws[f"G{current_row}"] = _sheet_scalar(card.get("source_ref"))
        source_ref_items = _coerce_str_list(card.get("source_ref"))
        evidence_tag_items = sorted(dict.fromkeys(evidence_refs or source_ref_items))
        ws[f"G{current_row}"] = _sheet_scalar(_render_source_refs(card.get("source_ref"), evidence_lookup))
        ws[f"H{current_row}"] = _sheet_scalar("; ".join(evidence_tag_items))
        ws[f"I{current_row}"] = _sheet_scalar(" | ".join(risk_notes))
        ws[f"J{current_row}"] = _sheet_scalar(card.get("confidence"))
        current_row += 1

    consolidated_logic = forecast_basis.get("consolidated_logic", {})
    if consolidated_logic:
        for key, value in consolidated_logic.items():
            ws[f"A{current_row}"] = "综合逻辑"
            ws[f"B{current_row}"] = _sheet_scalar(key)
            ws[f"F{current_row}"] = _sheet_scalar(value)
            current_row += 1

    wb.save(workbook_path)


def run_contract_workflow(
    *,
    workbook_map_path: Path,
    forecast_basis_path: Path,
    evidence_store_path: Path,
    workbook_path: Path,
    output_dir: Path,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook_map = load_json(workbook_map_path)
    forecast_basis = load_json(forecast_basis_path)
    evidence_store = normalize_evidence_store(
        load_jsonl(evidence_store_path),
        company=forecast_basis.get("company"),
    )
    validate_workbook_map(workbook_map)
    validate_forecast_basis(forecast_basis)
    validate_evidence_store_payload(evidence_store)

    artifact_hashes = {
        "source_workbook_hash": sha256_file(workbook_path),
        "workbook_map_hash": sha256_json(workbook_map),
        "forecast_basis_hash": sha256_json(forecast_basis),
        "evidence_store_hash": sha256_jsonl(evidence_store),
    }

    cell_instructions = build_cell_instructions(
        workbook_map,
        forecast_basis,
        source_workbook_hash=artifact_hashes["source_workbook_hash"],
        workbook_map_hash=artifact_hashes["workbook_map_hash"],
        forecast_basis_hash=artifact_hashes["forecast_basis_hash"],
        evidence_store_hash=artifact_hashes["evidence_store_hash"],
    )
    artifact_hashes["cell_instructions_hash"] = sha256_json(cell_instructions)
    cell_instructions["cell_instructions_hash"] = artifact_hashes["cell_instructions_hash"]
    validate_cell_instructions_payload(cell_instructions)

    artifact_hashes_path = output_dir / "artifact_hashes.json"
    cell_instructions_path = output_dir / "cell_instructions.json"
    candidate_workbook_path = output_dir / "candidate.xlsx"
    patch_log_path = output_dir / "patch_log.json"
    verification_report_path = output_dir / "verification_report.json"

    artifact_hashes_path.write_text(json.dumps(artifact_hashes, ensure_ascii=False, indent=2), encoding="utf-8")
    cell_instructions_path.write_text(json.dumps(cell_instructions, ensure_ascii=False, indent=2), encoding="utf-8")

    patch_log = execute_patch_from_instructions(
        workbook_path=workbook_path,
        workbook_map=workbook_map,
        cell_instructions=cell_instructions,
        output_workbook=candidate_workbook_path,
        patch_log_path=patch_log_path,
    )

    _materialize_forecast_basis_sheet(
        workbook_path=candidate_workbook_path,
        forecast_basis=forecast_basis,
        evidence_store=evidence_store,
    )
    final_hash = sha256_file(candidate_workbook_path)
    for entry in patch_log:
        entry["output_hash"] = final_hash
    validate_patch_log_payload(patch_log)
    patch_log_path.write_text(json.dumps(patch_log, ensure_ascii=False, indent=2), encoding="utf-8")

    verify_contract_patch(
        cell_instructions=cell_instructions,
        patch_log=patch_log,
        candidate_workbook_path=candidate_workbook_path,
        report_path=verification_report_path,
    )

    return {
        "artifact_hashes_path": artifact_hashes_path,
        "cell_instructions_path": cell_instructions_path,
        "candidate_workbook_path": candidate_workbook_path,
        "patch_log_path": patch_log_path,
        "verification_report_path": verification_report_path,
    }
