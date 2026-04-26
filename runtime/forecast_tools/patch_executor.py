# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import os
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

import openpyxl

from .build_cell_instructions import ContractValidationError
from .contract_validators import validate_cell_instructions_payload, validate_patch_log_payload
from .artifact_utils import sha256_file


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=path.parent, suffix=".tmp") as handle:
        handle.write(text)
        temp_path = Path(handle.name)
    os.replace(temp_path, path)


def execute_patch_from_instructions(
    *,
    workbook_path: Path,
    workbook_map: dict[str, Any],
    cell_instructions: dict[str, Any],
    output_workbook: Path,
    patch_log_path: Path,
) -> list[dict[str, Any]]:
    validate_cell_instructions_payload(cell_instructions)

    main_sheet = workbook_map.get("main_modeling_sheet")
    if not main_sheet:
        raise ContractValidationError("workbook_map missing main_modeling_sheet")
    if cell_instructions.get("main_modeling_sheet") != main_sheet:
        raise ContractValidationError("cell_instructions.main_modeling_sheet does not match workbook_map")

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise ContractValidationError(f"workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path)
    if main_sheet not in wb.sheetnames:
        raise ContractValidationError(f"main modeling sheet not found: {main_sheet}")

    patch_log: list[dict[str, Any]] = []
    lineage_instruction_hash = cell_instructions.get("cell_instructions_hash")
    lineage_basis_hash = cell_instructions.get("forecast_basis_hash")
    lineage_map_hash = cell_instructions.get("workbook_map_hash")
    lineage_evidence_hash = cell_instructions.get("evidence_store_hash")
    for idx, instruction in enumerate(cell_instructions["instructions"]):
        sheet_name = instruction["sheet"]
        if sheet_name not in wb.sheetnames:
            raise ContractValidationError(f"instruction {idx} targets missing sheet: {sheet_name}")

        ws = wb[sheet_name]
        cell_ref = instruction["cell"]
        before = ws[cell_ref].value
        write_type = instruction["write_type"]
        if write_type == "value":
            after = instruction["value"]
        elif write_type in {"formula", "verification_target"}:
            after = instruction["formula_template"]
        else:
            raise ContractValidationError(f"unsupported write_type: {write_type}")

        ws[cell_ref] = after
        patch_log.append(
            {
                "sheet": sheet_name,
                "cell": cell_ref,
                "row_id": instruction["row_id"],
                "year": instruction["year"],
                "before": before,
                "after": after,
                "write_type": write_type,
                "instruction_id": instruction["instruction_id"],
                "basis_ref": instruction["source_basis_ref"],
                "review_flag": instruction.get("review_flag"),
                "formula_preserved": instruction.get("formula_preserved"),
                "parity_audit_status": "pending",
                "instruction_hash": lineage_instruction_hash,
                "basis_hash": lineage_basis_hash,
                "map_hash": lineage_map_hash,
                "evidence_hash": lineage_evidence_hash,
            }
        )

    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(delete=False, dir=output_workbook.parent, suffix=".xlsx") as handle:
        temp_workbook = Path(handle.name)
    try:
        wb.save(temp_workbook)
        os.replace(temp_workbook, output_workbook)
    finally:
        if temp_workbook.exists():
            temp_workbook.unlink()

    final_hash = sha256_file(output_workbook)
    for entry in patch_log:
        entry["output_hash"] = final_hash

    validate_patch_log_payload(patch_log)
    _atomic_write_text(patch_log_path, json.dumps(patch_log, ensure_ascii=False, indent=2))
    return patch_log
